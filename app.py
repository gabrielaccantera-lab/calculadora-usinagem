import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import math
import re
import unicodedata

st.set_page_config(page_title="Calculadora de Recursos — Usinagem", layout="wide", page_icon="🏭")

JD_VERDE       = "#367C2B"
JD_VERDE_ESC   = "#1F4D19"
JD_AMARELO     = "#FFDE00"
JD_AMARELO_ESC = "#C9A800"
JD_TEXTO       = "#1A1A1A"

st.markdown("""
<style>
.jd-header{background:#1F4D19;padding:16px 24px;border-radius:10px;border-left:6px solid #FFDE00;margin-bottom:20px;}
.jd-header h1{color:#FFFFFF;margin:0;font-size:21px;font-weight:700;}
.jd-header p{color:#b8d4b4;margin:4px 0 0;font-size:12px;}
.jd-section{font-size:15px;font-weight:700;color:#FFDE00;border-left:4px solid #FFDE00;padding-left:10px;margin:22px 0 10px;}
.jd-sub{font-size:12px;font-weight:600;color:#7BC67A;margin:12px 0 4px;text-transform:uppercase;letter-spacing:.04em;}
.aviso-erro{background:#3D0000;border-left:4px solid #FF5252;border-radius:6px;padding:9px 13px;margin:5px 0;font-size:12px;color:#FF8A80;}
.aviso-warn{background:#3D2D00;border-left:4px solid #FFDE00;border-radius:6px;padding:9px 13px;margin:5px 0;font-size:12px;color:#FFE57F;}
.aviso-ok{background:#003D10;border-left:4px solid #69F0AE;border-radius:6px;padding:9px 13px;margin:5px 0;font-size:12px;color:#B9F6CA;}
.formula-box{background:#0D1117;color:#A8E6A3;font-family:monospace;padding:10px 14px;border-radius:6px;font-size:12px;line-height:1.8;border-left:3px solid #FFDE00;}
.mem-step{background:#1A1A2E;border:1px solid #444;border-radius:8px;padding:12px 16px;margin:6px 0;color:#FAFAFA;}
.mem-step b{color:#FFDE00;}
.mem-step .step-num{background:#FFDE00;color:#1F4D19;border-radius:50%;width:24px;height:24px;display:inline-flex;align-items:center;justify-content:center;font-size:12px;font-weight:700;margin-right:8px;}
.log-line{font-family:monospace;font-size:11px;color:#AAAAAA;padding:1px 0;}
.log-ok{color:#69F0AE;} .log-warn{color:#FFDE00;} .log-err{color:#FF5252;}
.cenario-card{background:#1A1A2E;border:1.5px solid #333;border-radius:10px;padding:14px 16px;margin:6px 0;color:#FAFAFA;}
.cenario-card b{color:#FFDE00;}
.kpi-card{background:linear-gradient(135deg,#1A2E1A 0%,#0D1F0D 100%);border:1px solid #2A4A2A;border-radius:12px;padding:16px 18px;margin:4px 0;position:relative;overflow:hidden;}
.kpi-card::before{content:'';position:absolute;top:0;left:0;width:4px;height:100%;background:#FFDE00;}
.kpi-card .kpi-icon{font-size:22px;margin-bottom:4px;}
.kpi-card .kpi-label{font-size:10px;color:#7BC67A;text-transform:uppercase;letter-spacing:.06em;font-weight:600;}
.kpi-card .kpi-value{font-size:26px;font-weight:800;color:#FFFFFF;line-height:1.1;}
.kpi-card .kpi-sub{font-size:11px;color:#AAAAAA;margin-top:2px;}
.kpi-card.destaque{border-color:#FFDE00;}
.kpi-card.destaque::before{background:#367C2B;width:100%;height:3px;top:0;left:0;}
.mes-row{display:flex;align-items:center;gap:10px;padding:7px 12px;border-radius:8px;margin:3px 0;background:#131313;border:1px solid #222;}
.mes-nome{font-size:12px;font-weight:700;color:#FFDE00;min-width:36px;}
.mes-bar{flex:1;height:8px;background:#1A2E1A;border-radius:4px;overflow:hidden;}
.mes-bar-fill{height:100%;border-radius:4px;background:linear-gradient(90deg,#367C2B,#FFDE00);}
.mes-num{font-size:12px;color:#FFFFFF;font-weight:700;min-width:28px;text-align:right;}
.mes-labor{font-size:11px;min-width:44px;text-align:right;}
.tp{display:inline-block;border-radius:20px;padding:1px 8px;font-size:10px;font-weight:700;margin:0 1px;}
.tpA{background:#1F4D19;color:#92D050;} .tpB{background:#3D2D00;color:#FFDE00;} .tpC{background:#0D2040;color:#00B0F0;}
.gauge-wrap{text-align:center;padding:6px 0;}
/* ── Tab-nav via radio ── */
div[data-testid="stRadio"]:has(div[role="radiogroup"]>label:nth-child(5)){margin-bottom:0!important;}
div[data-testid="stRadio"]:has(div[role="radiogroup"]>label:nth-child(5)) div[role="radiogroup"]{gap:0!important;flex-wrap:nowrap;border-bottom:2px solid #2a2a2a;padding-bottom:0;}
div[data-testid="stRadio"]:has(div[role="radiogroup"]>label:nth-child(5)) div[role="radiogroup"]>label{background:transparent;border:1px solid transparent;border-bottom:2px solid transparent;border-radius:4px 4px 0 0;padding:9px 15px!important;margin-bottom:-2px;cursor:pointer;font-size:13px;color:#777;transition:color .15s;}
div[data-testid="stRadio"]:has(div[role="radiogroup"]>label:nth-child(5)) div[role="radiogroup"]>label:hover{color:#ccc;background:#151515;}
div[data-testid="stRadio"]:has(div[role="radiogroup"]>label:nth-child(5)) div[role="radiogroup"]>label:has(input:checked){color:#FFDE00;border-color:#2a2a2a #2a2a2a transparent;border-bottom:2px solid #FFDE00;background:#0d1a0d;}
div[data-testid="stRadio"]:has(div[role="radiogroup"]>label:nth-child(5)) div[role="radiogroup"]>label>div:first-child{display:none!important;}
/* Prevent content dimming during tab switches */
.main .block-container{opacity:1!important;transition:none!important;}
.element-container{opacity:1!important;transition:none!important;}
</style>
""", unsafe_allow_html=True)

MESES       = ["Novembro","Dezembro","Janeiro","Fevereiro","Março","Abril",
               "Maio","Junho","Julho","Agosto","Setembro","Outubro"]
MESES_ABREV = ["NOV","DEZ","JAN","FEV","MAR","ABR","MAI","JUN","JUL","AGO","SET","OUT"]

COL_MAP = {
    "INPUTTEMPO": {
        "centro":   ["Máquina","MÁQUINA","maquina","Maquina","Centro","CENTRO","centro",
                     "Estação","Estacao","ESTACAO","Posto","POSTO","Célula","Celula","CELULA",
                     "Centro de Trabalho","Centro de Usinagem","CEN","Cen"],
        "peca":     ["PEÇA","Peça","peca","PECA","REFERÊNCIA","Referência","referencia","REFERENCIA",
                     "REF","Ref.","Ref","N. PEÇA","N. PECA","N.PECA","Código","Codigo","CODIGO",
                     "COD. PEÇA","COD PECA","Cod. Peça","Part Number","PN","SKU","ITEM"],
        "pc_trat":  ["Quantidade por Veículo","Quantidade por Veiculo","QTD POR VEÍCULO","QTD POR VEICULO",
                     "QTD/VEÍCULO","QTD/VEICULO","QUANTIDADE POR VEICULO","Qtd/Veículo","Qtd por Veículo",
                     "PEÇA/TRATOR","PEÇA TRATOR","PECA TRATOR","PEÇA\nTRATOR","PECA\nTRATOR",
                     "Quantidade Trator","QTD TRATOR","QTD/TRAT","PEÇA POR TRATOR","PCS POR TRATOR",
                     "PÇ/TRAT","PC/TRAT","pc_trat","Qtd Veículo","Qtd Veiculo"],
        "t_ciclo":  ["Tempo\nCiclo\n(min)","Tempo Ciclo (min)","Tempo Ciclo","Tempo de Ciclo",
                     "TEMPO CICLO","TEMPO DE CICLO","T.CICLO","T_CICLO","T CICLO","t_ciclo",
                     "Ciclo (min)","Ciclo (min.)","CICLO","TC","TC (min)","Cycle Time","Cycle Time (min)"],
        "t_labor":  ["Tempo\nLabor\n(min)","Tempo Labor (min)","Tempo Labor","Tempo de Labor",
                     "TEMPO LABOR","TEMPO DE LABOR","T.LABOR","T_LABOR","T LABOR","t_labor",
                     "Labor (min)","Labor (min.)","LABOR","TL","TL (min)",
                     "Tempo Operador","Tempo Mão de Obra","Tempo Mao de Obra",
                     "Manuseio (min)","Tempo Manuseio","Tempo Assistência","Tempo Assistencia",
                     "Handling Time","Operator Time"],
    },
    "INPUTDISTRIBUIÇÃO": {
        "centro":     ["Máquina","MÁQUINA","maquina","Maquina","Centro","CENTRO","centro",
                       "Estação","Estacao","ESTACAO","Posto","POSTO","Célula","Celula","CELULA",
                       "Centro de Trabalho","Centro de Usinagem","CEN","Cen"],
        "peca":       ["PEÇA","Peça","peca","PECA","REFERÊNCIA","Referência","referencia","REFERENCIA",
                       "REF","Ref.","Ref","N. PEÇA","N. PECA","N.PECA","Código","Codigo","CODIGO",
                       "COD. PEÇA","COD PECA","Cod. Peça","Part Number","PN","SKU","ITEM"],
        "div_carga":  ["Divisão\nCarga\nENTRE\nMÁQUINAS","Div Carga","DIV_CARGA","div_carga",
                       "Divisão de Carga","Divisão Carga","DIVISAO CARGA","Divisão Entre Máquinas",
                       "Divisão Entre Maquinas","Distribuição Carga","Distribuicao Carga",
                       "DIV CARGA","DC","Fator Carga","% Carga","Proporção Carga"],
        "vol_int":    ["Vol.\nInterna","Vol. Interna","VOL_INT","vol_int","VOL. INTERNA",
                       "Volume Interna","Volume Interno","Vol. Interno","Vol Interno","VOLUME INTERNO",
                       "Volume de \nProdução\nInterna","Volume de\nProdução\nInterna",
                       "Volume Produção Interna","Vol Produção Interna",
                       "Produção Interna","VI","% Volume Interno","% Vol. Interno"],
        "div_volume": ["Divisão \nde\nVolume\nENTRE\nPEÇAS","Divisão\nde\nVolume\nENTRE\nPEÇAS",
                       "Div Volume","DIV_VOLUME","div_volume","Divisão de Volume","Divisão Volume",
                       "DIVISAO VOLUME","Divisão Volume Peças","Distribuição de Volume",
                       "Divisão Entre Peças","Divisão Entre Pecas","DIV VOLUME","DV",
                       "Proporção de Volume","% Volume por Peça","% Vol. Peça"],
        "disponib":   ["Disponi-\nbilidade","Disponibilidade","DISPONIB","disponib",
                       "DISPONIBILIDADE","Disponib.","Disp.","DISP","% Disponibilidade",
                       "Tempo Disponível","Tempo Disponivel","Availability","Avail."],
        "perf_op":    ["Performance\nOperador X\nMáquina","Performance Operador X Máquina",
                       "Performance\nOperador X\nMaquina","Performance Operador X Maquina",
                       "Performance\nOperador X\n Máquina","Performance\nOperador X\n Maquina",
                       "Performance\nOperador\nMáquina","Performance Operador Máquina",
                       "Performance\nOperador\nMaquina","Performance Operador Maquina",
                       "Performance Operador X Maq","Performance Op X Maq",
                       "Perf. Operador Máquina","Perf. Operador Maquina",
                       "Perf Operador Máquina","Perf Operador Maquina",
                       "Perf. Op. Máquina","Perf. Op. Maquina",
                       "Performance Operador","PERF_OP","perf_op","PERFORMANCE",
                       "Fator Operador","Fator Performance","% Performance",
                       "Eficiência Operador","Eficiencia Operador","Índice Operador",
                       "NOVO","Performance\nOperador\nX\nMáquina","Performance\nOperador\nX\nMaquina"],
    },
}

ABA_FORMATOS = {
    "INPUTPMP": "**INPUTPMP** — Linha 1: dias trabalhados (colunas B→M = Nov→Out). Linhas 3+: modelos, colunas B→M = qtd peças.",
    "INPUTTEMPO": "**INPUTTEMPO** — Cabeçalho linha 1. Colunas obrigatórias: `Máquina`, `PEÇA` (ou `REFERÊNCIA`), `PEÇA/TRATOR` (ou `Quantidade por Veículo`), `Tempo Ciclo (min)`, `Tempo Labor (min)`. Coluna opcional: `UM`.",
    "INPUTDISTRIBUIÇÃO": "**INPUTDISTRIBUIÇÃO** — Cabeçalho linha 1. Colunas obrigatórias: `Máquina`, `PEÇA` (ou `REFERÊNCIA`), `Div Carga`, `Vol. Interna`, `Div Volume`, `Disponibilidade`, `Performance Operador X Máquina`. Colunas opcionais: `PEÇA/TRATOR`, `UM`, `Tempo Ciclo (min)`, `Tempo Labor (min)`, `Índice Ciclo`.",
    "INPUTAPLICAÇÃO": "**INPUTAPLICAÇÃO** — Cabeçalho linha 1. Col A=Centro, Col B=PEÇA (ou REFERÊNCIA), Col C=Descrição, Col D=PÇ/TRAT, Col E=UM (opcional), Col F+=modelos (valor 1=ativo, 0=inativo).",
    "INPUTTURNOS": "**INPUTTURNOS** — Linha 1: horas acumuladas. B1=Turno A, C1=Turno B, D1=Turno C.",
}

def _norm(s):
    s = re.sub(r'\s+', ' ', str(s).lower().replace("\n"," ").replace("\r"," ")).strip()
    return unicodedata.normalize('NFD', s).encode('ascii', 'ignore').decode('ascii')

def _find_header_row(df, all_candidates, max_scan=10):
    """
    Varre as primeiras linhas do DataFrame (lido com header=None) e retorna
    o índice da primeira linha que contém pelo menos uma célula reconhecida
    como nome de coluna esperado. Ignora linhas completamente em branco.
    """
    known = {_norm(c) for cands in all_candidates for c in cands}
    for r in range(min(max_scan, len(df))):
        row_vals = [str(v) for v in df.iloc[r] if pd.notna(v) and str(v).strip() != ""]
        if not row_vals:
            continue  # linha em branco — pula
        row_norms = {_norm(v) for v in row_vals}
        if row_norms & known:
            return r
    return 0  # fallback: assume linha 0

def _apply_header(df, hdr_row, log, aba):
    """Usa hdr_row como cabeçalho, descarta linhas anteriores, loga se pulou algo."""
    if hdr_row > 0:
        log.append(f"   ℹ️ [{aba}] {hdr_row} linha(s) em branco/extras ignorada(s) antes do cabeçalho")
    df.columns = df.iloc[hdr_row].astype(str)
    df = df.iloc[hdr_row + 1:].reset_index(drop=True)
    return df

def find_col(df, candidates, aba, campo):
    # 1) busca exata
    for c in candidates:
        if c in df.columns:
            return c
    # 2) busca normalizada
    norm_map = {_norm(col): col for col in df.columns}
    for c in candidates:
        nc = _norm(c)
        if nc in norm_map:
            st.session_state.setdefault("log_leitura",[]).append(
                f"ℹ️ [{aba}] Campo '{campo}' encontrado via normalização: '{norm_map[nc]}'")
            return norm_map[nc]
    # 3) fallback por índice posicional
    idx_fallback = {
        "centro":0,"peca":1,"pc_trat":3,"t_ciclo":5,"t_labor":6,
        "div_carga":7,"vol_int":8,"div_volume":9,"disponib":10,"perf_op":11
    }
    if campo in idx_fallback:
        idx = idx_fallback[campo]
        if idx < len(df.columns):
            st.session_state.setdefault("log_leitura",[]).append(
                f"⚠️ [{aba}] Campo '{campo}' não encontrado pelo nome — usando coluna {idx+1} ({df.columns[idx]}) como fallback")
            return df.columns[idx]
    # 4) Erro detalhado com sugestão
    nomes_esperados = " ou ".join(f'"{c}"' for c in candidates[:3])
    colunas_disponiveis = ", ".join(f'"{c}"' for c in df.columns)
    raise ValueError(
        f"\n🔴 [{aba}] Campo obrigatório '{campo}' não encontrado!\n\n"
        f"   O app esperava uma coluna chamada {nomes_esperados} (entre outras variações).\n\n"
        f"   Colunas encontradas na aba '{aba}':\n"
        f"   {colunas_disponiveis}\n\n"
        f"   👉 Renomeie a coluna correspondente para um dos nomes esperados:\n"
        f"   {', '.join(f'{c}' for c in candidates)}"
    )

def verificar_prefixo_centro(df, aba, log):
    """Avisa se algum centro não começa com 'CEN', mas deixa passar."""
    if "centro" not in df.columns:
        return
    centros_fora = df[~df["centro"].astype(str).str.upper().str.startswith("CEN")]["centro"].dropna().unique()
    if len(centros_fora) > 0:
        exemplos = ", ".join(f'"{c}"' for c in centros_fora[:5])
        log.append(
            f"⚠️ [{aba}] {len(centros_fora)} centro(s) não iniciam com 'CEN' — "
            f"verifique se estão corretos: {exemplos}"
            f"{'...' if len(centros_fora) > 5 else ''}"
        )

def find_aba(sheetnames, candidatos):
    """Busca aba por nome exato, depois por similaridade (ignora underline, espaço, maiúsculas, acento)."""
    import unicodedata
    def _norm_aba(s):
        s = str(s).upper().strip()
        s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
        for ch in [' ', '_', '-', '.']: s = s.replace(ch, '')
        return s
    # 1) busca exata
    for c in candidatos:
        if c in sheetnames:
            return c
    # 2) busca normalizada
    norm_map = {_norm_aba(a): a for a in sheetnames}
    for c in candidatos:
        nc = _norm_aba(c)
        if nc in norm_map:
            return norm_map[nc]
    return None

@st.cache_data(show_spinner=False)
def verificar_abas(fb):
    try:
        wb = openpyxl.load_workbook(BytesIO(fb), read_only=True, data_only=True)
        sheetnames = wb.sheetnames; wb.close()
    except: sheetnames = []

    CANDIDATOS = {
        "INPUTPMP":         ["INPUTPMP","INPUT_PMP","INPUT PMP","PMP"],
        "INPUTTEMPO":        ["INPUTTEMPO","IMPUTTEMPO","INPUT_TEMPO","INPUT TEMPO","TEMPO"],
        "INPUTDISTRIBUIÇÃO": ["INPUTDISTRIBUIÇÃO","INPUTDISTRIBUICAO","IMPUTDISTRIBUIÇÃO",
                              "IMPUTDISTRIBUICAO","INPUT_DISTRIBUIÇÃO","INPUT DISTRIBUIÇÃO",
                              "INPUT_DISTRIBUICAO","DISTRIBUIÇÃO","DISTRIBUICAO"],
        "INPUTAPLICAÇÃO":    ["INPUTAPLICAÇÃO","INPUTAPLICACAO","IMPUTAPLICAÇÃO",
                              "IMPUTAPLICACAO","INPUT_APLICAÇÃO","INPUT APLICAÇÃO",
                              "INPUT_APLICACAO","APLICAÇÃO","APLICACAO"],
        "INPUTTURNOS":       ["INPUTTURNOS","IMPUTTURNOS","INPUT_TURNOS","INPUT TURNOS","TURNOS"],
    }

    resultado = {}
    for aba_padrao, candidatos in CANDIDATOS.items():
        encontrada = find_aba(sheetnames, candidatos)
        resultado[aba_padrao] = encontrada  # None se não encontrou, nome real se encontrou
    return resultado

def _get_aba(chave):
    """Retorna o nome real da aba encontrada no arquivo, ou o nome padrão como fallback."""
    return st.session_state.get("_abas_map", {}).get(chave) or chave

def read_pmp(fb, log):
    aba = _get_aba("INPUTPMP")
    try:
        df = pd.read_excel(BytesIO(fb), sheet_name=aba, header=None)
    except Exception as e:
        raise ValueError(f"Não foi possível ler '{aba}': {e}\n\n{ABA_FORMATOS['INPUTPMP']}")
    log.append(f"✅ INPUTPMP lido (aba: '{aba}'): {df.shape[0]}L × {df.shape[1]}C")
    def _is_num(v):
        try: int(float(v)); return True
        except: return False
    def _row_has_content(r):
        return any(pd.notna(df.iloc[r, c]) and str(df.iloc[r, c]).strip() != "" for c in range(df.shape[1]))
    # Pular linhas em branco no topo — encontrar primeira linha com conteúdo
    first_content = 0
    for r in range(min(10, len(df))):
        if _row_has_content(r):
            first_content = r; break
    # A partir daí, detectar qual linha tem os dias trabalhados
    dias_row, data_start = first_content, first_content + 2
    for r in range(first_content, min(first_content + 5, len(df))):
        if df.shape[1] > 1 and _is_num(df.iloc[r, 1]):
            dias_row = r
            data_start = r + 2
            break
        elif df.shape[1] > 1 and not _is_num(df.iloc[r, 1]) and r + 1 < len(df) and _is_num(df.iloc[r + 1, 1]):
            dias_row = r + 1
            data_start = r + 3
            break
    if first_content > 0:
        log.append(f"   ℹ️ [INPUTPMP] {first_content} linha(s) em branco ignorada(s) no topo")
    log.append(f"   Layout INPUTPMP: linha de dias={dias_row+1}, dados a partir da linha={data_start+1}")
    # Mapear colunas pelos nomes dos meses (se existir linha de cabeçalho entre dias e dados)
    _MES_NORMS = {}
    for m, abr in zip(MESES, MESES_ABREV):
        for v in [m, m.upper(), abr, abr.lower()]:
            _MES_NORMS[_norm(v)] = m
    mes_col = {}  # mes → índice de coluna
    if data_start > dias_row + 1:
        hdr_r = data_start - 1
        for c in range(df.shape[1]):
            v = df.iloc[hdr_r, c]
            if pd.notna(v):
                k = _norm(str(v))
                if k in _MES_NORMS:
                    mes_col[_MES_NORMS[k]] = c
    if mes_col:
        log.append(f"   Colunas de mês encontradas por nome: {list(mes_col.keys())[:4]}...")
    else:
        # Fallback: posicional — meses nas colunas 1..12 em ordem de MESES
        mes_col = {m: i for i, m in enumerate(MESES, 1)}
    dias = {}
    for m, col in mes_col.items():
        v = df.iloc[dias_row, col] if col < df.shape[1] else None
        try: dias[m] = int(float(v)) if pd.notna(v) else 0
        except: dias[m] = 0
    log.append(f"   Dias: { {m:d for m,d in dias.items() if d>0} }")
    # Detectar coluna de modelos (primeira coluna, ou pela label "MODELO")
    col_modelo = 0
    if data_start > 0:
        for c in range(min(3, df.shape[1])):
            v = df.iloc[data_start - 1, c]
            if pd.notna(v) and _norm(str(v)) in {"modelo","modelos","model","models","equipamento","trator"}:
                col_modelo = c; break
    rows = []
    for r in range(data_start, len(df)):
        modelo = df.iloc[r, col_modelo]
        if pd.isna(modelo): continue
        if str(modelo).strip().upper() in ("MODELO", "TOTAL", ""): continue
        for m, col in mes_col.items():
            v = df.iloc[r, col] if col < df.shape[1] else None
            try: qtd = int(float(v)) if pd.notna(v) else 0
            except: qtd = 0
            if qtd > 0:
                rows.append({"modelo": str(modelo).strip(), "mes": m, "qtd": qtd})
    log.append(f"   {len(rows)} registros com qtd>0")
    return pd.DataFrame(rows), dias

def read_turnos(fb):
    aba = _get_aba("INPUTTURNOS")
    _DEFAULTS = {"A": 7.5, "B": 14.25, "C": 19.5}
    try:
        df = pd.read_excel(BytesIO(fb), sheet_name=aba, header=None)
    except:
        return _DEFAULTS, False
    try:
        # 1) Busca por label de turno: procura linha com "TURNO A" ou "A" nas colunas
        _CANDS_A = {"turno a","a","ta","t.a","turno_a","horas a","h a","h.a"}
        _CANDS_B = {"turno b","b","tb","t.b","turno_b","horas b","h b","h.b"}
        _CANDS_C = {"turno c","c","tc","t.c","turno_c","horas c","h c","h.c"}
        col_A = col_B = col_C = None
        hdr_row = None
        for r in range(min(5, len(df))):
            for c in range(min(10, df.shape[1])):
                v = _norm(str(df.iloc[r, c])) if pd.notna(df.iloc[r, c]) else ""
                if v in _CANDS_A and col_A is None: col_A = c; hdr_row = r
                elif v in _CANDS_B and col_B is None: col_B = c
                elif v in _CANDS_C and col_C is None: col_C = c
        if col_A is not None and col_B is not None and col_C is not None and hdr_row is not None:
            val_row = hdr_row + 1 if hdr_row + 1 < len(df) else hdr_row
            def _rv(r, c):
                v = df.iloc[r, c]
                return float(v) if pd.notna(v) else None
            hA = _rv(val_row, col_A) or _DEFAULTS["A"]
            hB = _rv(val_row, col_B) or _DEFAULTS["B"]
            hC = _rv(val_row, col_C) or _DEFAULTS["C"]
            return {"A": hA, "B": hB, "C": hC}, True
        # 2) Fallback posicional: procura primeira linha com 3+ valores numéricos >= 1
        for r in range(min(5, len(df))):
            nums = []
            for c in range(1, min(8, df.shape[1])):
                v = df.iloc[r, c]
                if pd.notna(v):
                    try:
                        f = float(v)
                        if f >= 1: nums.append((c, f))
                    except: pass
            if len(nums) >= 3:
                return {"A": nums[0][1], "B": nums[1][1], "C": nums[2][1]}, True
        # 3) Fallback original: linha 0, colunas 1/2/3
        hA = float(df.iloc[0,1]) if df.shape[1] > 1 and pd.notna(df.iloc[0,1]) else _DEFAULTS["A"]
        hB = float(df.iloc[0,2]) if df.shape[1] > 2 and pd.notna(df.iloc[0,2]) else _DEFAULTS["B"]
        hC = float(df.iloc[0,3]) if df.shape[1] > 3 and pd.notna(df.iloc[0,3]) else _DEFAULTS["C"]
        return {"A": hA, "B": hB, "C": hC}, True
    except:
        return _DEFAULTS, False

def read_tempo(fb, log):
    aba = _get_aba("INPUTTEMPO")
    try:
        df = pd.read_excel(BytesIO(fb), sheet_name=aba, header=None)
    except Exception as e:
        raise ValueError(f"Não foi possível ler '{aba}': {e}\n\n{ABA_FORMATOS['INPUTTEMPO']}")
    mp = COL_MAP["INPUTTEMPO"]
    hdr = _find_header_row(df, list(mp.values()))
    df = _apply_header(df, hdr, log, aba)
    log.append(f"✅ INPUTTEMPO lido (aba: '{aba}'): {df.shape[0]}L")
    mp = COL_MAP["INPUTTEMPO"]
    # pc_trat is optional — try to find it, fall back to None
    try:
        pc_col = find_col(df, mp["pc_trat"], aba, "pc_trat")
    except (ValueError, KeyError):
        pc_col = None
    c = {k: find_col(df, v, aba, k) for k, v in mp.items() if k != "pc_trat"}
    out = df[[c["centro"], c["peca"], c["t_ciclo"], c["t_labor"]]].copy()
    out.columns = ["centro", "peca", "t_ciclo", "t_labor"]
    if pc_col is not None:
        _raw_pc = pd.to_numeric(df[pc_col], errors="coerce")
        _blank_pc = _raw_pc.isna().sum()
        out["pc_trat"] = _raw_pc.fillna(1.0).clip(lower=1.0)
        _acima1 = (out["pc_trat"] > 1.0).sum()
        log.append(f"   PÇ/TRAT lido de '{pc_col}' (INPUTTEMPO): {_acima1} linha(s) com valor >1" +
                   (f" ⚠️ {_blank_pc} célula(s) em branco → padrão 1.0" if _blank_pc else ""))
    out = out.dropna(subset=["centro"])
    verificar_prefixo_centro(out, aba, log)
    log.append(f"   {len(out)} combinações centro+peça")
    return out.copy()

def read_dist(fb, log):
    aba = _get_aba("INPUTDISTRIBUIÇÃO")
    try:
        df = pd.read_excel(BytesIO(fb), sheet_name=aba, header=None)
    except Exception as e:
        raise ValueError(f"Não foi possível ler '{aba}': {e}\n\n{ABA_FORMATOS['INPUTDISTRIBUIÇÃO']}")
    mp = COL_MAP["INPUTDISTRIBUIÇÃO"]
    hdr = _find_header_row(df, list(mp.values()))
    df = _apply_header(df, hdr, log, aba)
    log.append(f"✅ INPUTDISTRIBUIÇÃO lido (aba: '{aba}'): {df.shape[0]}L")
    log.append(f"   Colunas brutas: {list(df.columns)}")
    mp = COL_MAP["INPUTDISTRIBUIÇÃO"]
    c = {k: find_col(df, v, aba, k) for k,v in mp.items()}
    log.append(f"   Mapeamento: { {k: v for k,v in c.items()} }")
    out = df[[c["centro"],c["peca"],c["div_carga"],c["vol_int"],c["div_volume"],c["disponib"],c["perf_op"]]].copy()
    out.columns = ["centro","peca","div_carga","vol_int","div_volume","disponib","perf_op"]
    out["vol_int"]    = pd.to_numeric(out["vol_int"],    errors="coerce").fillna(1.0)
    out["div_carga"]  = pd.to_numeric(out["div_carga"],  errors="coerce").fillna(0.0)
    out["div_volume"] = pd.to_numeric(out["div_volume"], errors="coerce").fillna(0.0)
    out["disponib"]   = pd.to_numeric(out["disponib"],   errors="coerce").fillna(1.0)
    out["perf_op"]    = pd.to_numeric(out["perf_op"],    errors="coerce").fillna(1.0)
    out = out.dropna(subset=["centro"])
    verificar_prefixo_centro(out, aba, log)
    amostra = out[["centro","peca","div_carga","vol_int","div_volume","disponib","perf_op"]].head(3)
    for _, r in amostra.iterrows():
        log.append(f"   ✔ {r.centro}/{r.peca}: div_carga={r.div_carga}, vol_int={r.vol_int}, div_volume={r.div_volume}, disponib={r.disponib}, perf_op={r.perf_op}")
    log.append(f"   {len(out)} combinações")
    return out.copy()

def read_aplic(fb, log):
    aba = _get_aba("INPUTAPLICAÇÃO")
    try:
        df = pd.read_excel(BytesIO(fb), sheet_name=aba, header=None)
    except Exception as e:
        raise ValueError(f"Não foi possível ler '{aba}': {e}\n\n{ABA_FORMATOS['INPUTAPLICAÇÃO']}")
    # Para INPUTAPLICAÇÃO os candidatos de cabeçalho são: centro, peça e pc_trat
    _aplic_cands = COL_MAP["INPUTTEMPO"]["centro"] + COL_MAP["INPUTTEMPO"]["peca"] + \
                   ["PÇ/TRAT","PC/TRAT","PEÇA\nTRATOR","PEÇA TRATOR","Descrição","Descricao","DESCRICAO"]
    hdr = _find_header_row(df, [_aplic_cands])
    df = _apply_header(df, hdr, log, aba)
    log.append(f"✅ INPUTAPLICAÇÃO lido (aba: '{aba}'): {df.shape[0]}L")
    # Busca centro e peça por nome primeiro; fallback posicional (col 0 / col 1)
    _cen_cands = COL_MAP["INPUTTEMPO"]["centro"]
    _pec_cands = COL_MAP["INPUTTEMPO"]["peca"]
    _norm_map = {_norm(str(c)): c for c in df.columns}
    _col_cen = next((c for c in _cen_cands if c in df.columns), None) or \
               next((_norm_map[_norm(c)] for c in _cen_cands if _norm(c) in _norm_map), None) or \
               df.columns[0]
    _col_pec = next((c for c in _pec_cands if c in df.columns and c != _col_cen), None) or \
               next((_norm_map[_norm(c)] for c in _pec_cands if _norm(c) in _norm_map and _norm_map[_norm(c)] != _col_cen), None) or \
               df.columns[1]
    if _col_cen != df.columns[0] or _col_pec != df.columns[1]:
        log.append(f"   Centro='{_col_cen}' Peça='{_col_pec}' encontrados por nome")
    df = df.rename(columns={_col_cen: "centro", _col_pec: "peca"})

    pc_trat_candidates = ["PEÇA\nTRATOR","PÇ/TRAT","PC/TRAT","PCTRAT","Peça Trator","pc_trat",
                          "ERA PEÇA\nTRATOR","ERA PEÇA TRATOR","ERA PECA TRATOR","ERA PECA\nTRATOR",
                          "QUANTIDADE\nE\nPOR\nVEÍCULO","QUANTIDADE E POR VEÍCULO",
                          "QUANTIDADE E POR VEICULO","QUANTIDADE\nE\nPOR\nVEICULO",
                          "QTD POR VEÍCULO","QTD POR VEICULO","QTD/VEÍCULO","QTD/VEICULO"]
    pc_trat_col = next((c for c in pc_trat_candidates if c in df.columns), None)
    if pc_trat_col is None and len(df.columns) > 3:
        pc_trat_col = df.columns[3]

    colunas_ignorar = {"centro", "peca"}
    if pc_trat_col:
        colunas_ignorar.add(pc_trat_col)

    # Ignorar colunas "UM" / unidade de medida — não são colunas de modelo
    _UM_NORMS = {"um","u.m.","unidade","unid","unidade de medida","unit","und"}
    for _c in df.columns:
        if _norm(str(_c)) in _UM_NORMS:
            colunas_ignorar.add(_c)

    # Aceita qualquer coluna a partir do índice 4 como modelo (qualquer nome)
    # Também ignora colunas de tempo/distribuição que não pertencem a INPUTAPLICAÇÃO
    _NAO_MODELO_NORMS = {
        "tempo ciclo (min)","tempo labor (min)","t.ciclo","t.labor","ciclo","labor",
        "div carga","vol. interna","div volume","disponibilidade","performance",
        "descricao","descrição","description",
    }
    for _c in df.columns:
        if _norm(str(_c)) in _NAO_MODELO_NORMS:
            colunas_ignorar.add(_c)

    mcols = [
        c for i, c in enumerate(df.columns)
        if i >= 4
        and c not in colunas_ignorar
        and not str(c).startswith("Unnamed")
        and str(c).strip() not in ("", "nan", "None")
    ]

    if not mcols:
        raise ValueError(
            f"\n🔴 [INPUTAPLICAÇÃO] Nenhuma coluna de modelo encontrada a partir da coluna 5!\n\n"
            f"   Colunas encontradas: {', '.join(str(c) for c in df.columns)}\n\n"
            f"   👉 Verifique se as colunas de modelo estão a partir da 5ª coluna da aba INPUTAPLICAÇÃO.\n"
            f"   Os modelos podem ter qualquer nome — não precisam seguir um padrão específico."
        )

    log.append(f"   {len(mcols)} modelos: {mcols[:3]}{'...' if len(mcols)>3 else ''}")

    id_vars = ["centro", "peca", "pc_trat"] if pc_trat_col else ["centro", "peca"]
    if pc_trat_col:
        df["pc_trat"] = pd.to_numeric(df[pc_trat_col], errors="coerce").fillna(1.0).clip(lower=1.0)
        log.append(f"   PÇ/TRAT lido de '{pc_trat_col}'")
    else:
        df["pc_trat"] = 1.0

    melted = df[id_vars + mcols].melt(id_vars=id_vars, var_name="modelo", value_name="ativo")
    out = melted[melted["ativo"] == 1][id_vars + ["modelo"]].reset_index(drop=True)
    verificar_prefixo_centro(out, "INPUTAPLICAÇÃO", log)
    log.append(f"   {len(out)} combinações ativas")
    return out

def validar(pmp, tempo, dist, aplic, dias):
    erros, alertas, oks = [], [], []
    chaves_tempo = set(zip(tempo.centro, tempo.peca))
    chaves_dist  = set(zip(dist.centro,  dist.peca))
    chaves_aplic = set(zip(aplic.centro, aplic.peca))

    zero_disp = dist[dist.disponib == 0]
    if len(zero_disp):
        exemplos = zero_disp[["centro","peca"]].head(3).apply(lambda r: f"{r.centro}/{r.peca}", axis=1).tolist()
        erros.append(f"Disponibilidade=0 em {len(zero_disp)} linha(s) — causa divisão por zero no índice de ciclo. Ex: {', '.join(exemplos)}")

    zero_perf = dist[pd.to_numeric(dist.perf_op, errors="coerce").fillna(0) == 0]
    if len(zero_perf):
        exemplos = zero_perf[["centro","peca"]].head(3).apply(lambda r: f"{r.centro}/{r.peca}", axis=1).tolist()
        erros.append(f"Performance Operador=0 em {len(zero_perf)} linha(s) — causa divisão por zero no índice de ciclo. Ex: {', '.join(exemplos)}")

    diff_td = chaves_tempo - chaves_dist
    if diff_td:
        exemplos = list(diff_td)[:3]
        erros.append(f"{len(diff_td)} combinações em INPUTTEMPO sem INPUTDISTRIBUIÇÃO — não terão carga calculada. Ex: {exemplos}")

    tempo["t_ciclo"] = pd.to_numeric(tempo["t_ciclo"], errors="coerce")
    tempo["t_labor"] = pd.to_numeric(tempo["t_labor"], errors="coerce")
    t_invalidos = tempo[(tempo.t_ciclo.fillna(0) <= 0) | (tempo.t_labor.fillna(0) < 0)]
    if len(t_invalidos):
        exemplos = t_invalidos[["centro","peca","t_ciclo","t_labor"]].head(3).to_dict("records")
        erros.append(f"Tempo de ciclo ≤0 ou labor <0 em {len(t_invalidos)} linha(s) — verifique INPUTTEMPO. Ex: {exemplos[0]}")

    dist_num = dist.copy()
    dist_num["div_carga"]  = pd.to_numeric(dist_num["div_carga"],  errors="coerce").fillna(0)
    dist_num["div_volume"] = pd.to_numeric(dist_num["div_volume"], errors="coerce").fillna(0)
    zero_carga  = dist_num[dist_num["div_carga"]  == 0]
    zero_volume = dist_num[dist_num["div_volume"] == 0]
    if len(zero_carga):
        erros.append(f"div_carga=0 em {len(zero_carga)} linha(s) — zera completamente a carga daquele centro/peça")
    if len(zero_volume):
        erros.append(f"div_volume=0 em {len(zero_volume)} linha(s) — zera completamente a carga daquele centro/peça")

    sem_aplic = chaves_tempo - chaves_aplic
    if sem_aplic:
        exemplos = list(sem_aplic)[:3]
        alertas.append(f"{len(sem_aplic)} centro+peça sem modelo em INPUTAPLICAÇÃO — não entrarão no cálculo de carga. Ex: {exemplos}")

    modelos_sem = set(pmp.modelo.unique()) - set(aplic.modelo.unique())
    if modelos_sem:
        exemplos = list(modelos_sem)[:5]
        alertas.append(f"{len(modelos_sem)} modelo(s) com demanda no PMP mas sem aplicação em nenhuma máquina: {exemplos}")

    merged = tempo.merge(dist, on=["centro","peca"], how="inner")
    labor_maior = merged[merged.t_labor > merged.t_ciclo]
    if len(labor_maior):
        exemplos = labor_maior[["centro","peca"]].head(3).apply(lambda r: f"{r.centro}/{r.peca}", axis=1).tolist()
        alertas.append(f"{len(labor_maior)} linha(s) com t_labor > t_ciclo — operador mais ocupado que a máquina, verifique se é intencional. Ex: {', '.join(exemplos)}")

    for m in MESES:
        qtd_m = pmp[pmp.mes==m].qtd.sum() if len(pmp[pmp.mes==m]) else 0
        if qtd_m > 0 and dias.get(m,0) == 0:
            alertas.append(f"Mês '{m}' tem {int(qtd_m)} peças no PMP mas dias trabalhados=0 — mês será ignorado no cálculo")

    for m in MESES:
        qtd_m = pmp[pmp.mes==m].qtd.sum() if len(pmp[pmp.mes==m]) else 0
        if qtd_m == 0 and dias.get(m,0) > 0:
            alertas.append(f"Mês '{m}' tem {dias[m]} dias configurados mas nenhuma demanda no PMP — headcount será zero")

    dist_num2 = dist.copy()
    dist_num2["div_carga"] = pd.to_numeric(dist_num2["div_carga"], errors="coerce").fillna(0)
    soma_carga = dist_num2.groupby("centro")["div_carga"].sum()
    sobrecarga = soma_carga[soma_carga > 1.001]
    if len(sobrecarga):
        detalhes = [f"{c}={v:.2f}" for c,v in sobrecarga.items()][:4]
        alertas.append(f"div_carga soma >1 em {len(sobrecarga)} centro(s) — a carga está sendo multiplicada, verifique a distribuição. Ex: {', '.join(detalhes)}")

    disp_alta = dist[pd.to_numeric(dist.disponib, errors="coerce").fillna(0) > 1]
    if len(disp_alta):
        exemplos = disp_alta[["centro","peca","disponib"]].head(3).apply(lambda r: f"{r.centro} disp={r.disponib}", axis=1).tolist()
        alertas.append(f"Disponibilidade >1 (>100%) em {len(disp_alta)} linha(s) — verifique se está em decimal (ex: 0.85 = 85%). Ex: {', '.join(exemplos)}")

    pecas_com_demanda = set(pmp.merge(aplic, on="modelo")["peca"].unique()) if len(pmp) > 0 else set()
    centros_sem_demanda = set(tempo.centro.unique()) - set(
        tempo[tempo.peca.isin(pecas_com_demanda)].centro.unique()
    )
    if centros_sem_demanda:
        alertas.append(f"{len(centros_sem_demanda)} centro(s) sem nenhuma demanda ativa — aparecem em INPUTTEMPO mas nenhuma peça deles tem produção no PMP: {sorted(centros_sem_demanda)[:5]}")

    dist_vi = dist.copy()
    dist_vi["vol_int"] = pd.to_numeric(dist_vi["vol_int"], errors="coerce")
    vi_alto = dist_vi[dist_vi["vol_int"] > 5]
    if len(vi_alto):
        alertas.append(f"vol_interna > 5 em {len(vi_alto)} linha(s) — valor incomum, verifique se não é um erro de digitação")

    if not erros and not alertas:
        oks.append("Todos os inputs validados sem inconsistências.")
    return erros, alertas, oks

def calcular(pmp, tempo, dist, aplic, dias, horas_turno, thresholds, suporte_cfg,
             horas_efetivas=None, overrides=None, retornar_intermediarios=False):
    if horas_efetivas is None:
        horas_efetivas = horas_turno
    df = (aplic.merge(pmp, on="modelo")
               .merge(tempo, on=["centro","peca"])
               .merge(dist,  on=["centro","peca"]))
    # Resolve pc_trat: tempo's value (from "Quantidade por Veículo") takes precedence over aplic's
    if "pc_trat_x" in df.columns and "pc_trat_y" in df.columns:
        df["pc_trat"] = df["pc_trat_y"].fillna(df["pc_trat_x"]).fillna(1.0)
        df.drop(columns=["pc_trat_x","pc_trat_y"], inplace=True)
    elif "pc_trat_y" in df.columns:
        df.rename(columns={"pc_trat_y": "pc_trat"}, inplace=True)
    elif "pc_trat_x" in df.columns:
        df.rename(columns={"pc_trat_x": "pc_trat"}, inplace=True)
    if "pc_trat" not in df.columns:
        df["pc_trat"] = 1.0
    df["pc_trat"] = pd.to_numeric(df["pc_trat"], errors="coerce").fillna(1.0).clip(lower=1.0)
    if "vol_int" not in df.columns: df["vol_int"] = 1.0
    df["vol_int"]    = pd.to_numeric(df["vol_int"],    errors="coerce").fillna(1.0)
    df["div_carga"]  = pd.to_numeric(df["div_carga"],  errors="coerce").fillna(0.0)
    df["div_volume"] = pd.to_numeric(df["div_volume"], errors="coerce").fillna(0.0)
    df["disponib"]   = pd.to_numeric(df["disponib"],   errors="coerce").fillna(1.0)
    df["perf_op"]    = pd.to_numeric(df["perf_op"],    errors="coerce").fillna(1.0) if "perf_op" in df.columns else 1.0
    df["indice_ciclo"] = (df.t_ciclo * df.div_carga * df.div_volume * df.vol_int) / (df.disponib * df.perf_op)
    df["min_ciclo"]    = df.indice_ciclo * df.qtd
    df["min_labor"]    = df.t_labor * df.div_carga * df.qtd * df.pc_trat
    agg = df.groupby(["centro","mes"])[["min_ciclo","min_labor"]].sum().reset_index()
    thr_A = thresholds["A"] / 100
    thr_B = thresholds["B"] / 100
    thr_C = thresholds["C"] / 100
    hA, hB, hC = horas_turno["A"], horas_turno["B"], horas_turno["C"]
    heA, heB, heC = horas_efetivas["A"], horas_efetivas["B"], horas_efetivas["C"]
    resultados = {}
    for mes in MESES:
        d = dias.get(mes, 0)
        if d == 0: resultados[mes] = None; continue
        sub = agg[agg.mes == mes].copy()
        if sub.empty: resultados[mes] = None; continue
        minA = d*hA*60; minB = d*hB*60; minC = d*hC*60
        centros = []
        for _, row in sub.iterrows():
            cen = row.centro; mc = row.min_ciclo; ml = row.min_labor
            pA = mc/minA if minA>0 else 0
            pB = mc/minB if minB>0 else 0
            pC = mc/minC if minC>0 else 0
            aA = 1 if pA > thr_A else 0
            aB = 1 if pA > thr_B else 0
            aC = 1 if pB > thr_C else 0
            nA=aA; nB=aB; nC=aC
            if overrides and mes in overrides and cen in overrides[mes]:
                ov = overrides[mes][cen]
                if "A" in ov: nA=int(ov["A"]); aA = 1 if nA > 0 else 0
                if "B" in ov: nB=int(ov["B"]); aB = 1 if nB > 0 else 0
                if "C" in ov: nC=int(ov["C"]); aC = 1 if nC > 0 else 0
            centros.append({
                "centro":cen,"ocup_A":pA,"ocup_B":pB,"ocup_C":pC,
                "ativo_A":aA,"ativo_B":aB,"ativo_C":aC,
                "num_A":nA,"num_B":nB,"num_C":nC,
                "min_ciclo_total":mc,"min_labor_total":ml,
                "min_disp_A":minA,"min_disp_B":minB,"min_disp_C":minC,
                "horas_ciclo":mc/60,"horas_labor":ml/60,
                "horas_disp_A":d*heA*nA,"horas_disp_B":d*heB*nB,"horas_disp_C":d*heC*nC,
            })
        df_c = pd.DataFrame(centros)
        op_A = int(df_c.num_A.sum()); op_B = int(df_c.num_B.sum()); op_C = int(df_c.num_C.sum())
        def get_sup(key, t, op_count):
            cfg = suporte_cfg[key]
            if op_count == 0: return 0
            if cfg["modo"] == "auto":
                defaults = {"lavadora":{"A":1,"B":1,"C":0},"gravacao":{"A":1,"B":1,"C":0},
                            "preset":{"A":2,"B":1,"C":1},"coringa":{"A":1,"B":0,"C":0},
                            "facilitador":{"A":1,"B":1,"C":0}}
                return defaults[key][t]
            return cfg[t] if op_count > 0 else 0
        lav={t:get_sup("lavadora",t,[op_A,op_B,op_C]["ABC".index(t)]) for t in "ABC"}
        gra={t:get_sup("gravacao",t,[op_A,op_B,op_C]["ABC".index(t)]) for t in "ABC"}
        pre={t:get_sup("preset",t,[op_A,op_B,op_C]["ABC".index(t)]) for t in "ABC"}
        cor={t:get_sup("coringa",t,[op_A,op_B,op_C]["ABC".index(t)]) for t in "ABC"}
        fac={t:get_sup("facilitador",t,[op_A,op_B,op_C]["ABC".index(t)]) for t in "ABC"}
        tot_A = op_A+lav["A"]+gra["A"]+pre["A"]+cor["A"]+fac["A"]
        tot_B = op_B+lav["B"]+gra["B"]+pre["B"]+cor["B"]+fac["B"]
        tot_C = op_C+lav["C"]+gra["C"]+pre["C"]+cor["C"]+fac["C"]
        total = tot_A+tot_B+tot_C
        h_ciclo = float(df_c.horas_ciclo.sum()); h_labor = float(df_c.horas_labor.sum())
        h_ativos = float((df_c.horas_disp_A+df_c.horas_disp_B+df_c.horas_disp_C).sum())
        h_todos  = tot_A*d*heA+tot_B*d*heB+tot_C*d*heC
        resultados[mes] = {
            "centros":df_c,
            "op_A":op_A,"op_B":op_B,"op_C":op_C,
            "tot_A":tot_A,"tot_B":tot_B,"tot_C":tot_C,"total":total,
            "suporte":{"lavadora":lav,"gravacao":gra,"preset":pre,"coringa":cor,"facilitador":fac},
            "h_ciclo":h_ciclo,"h_labor":h_labor,"h_ativos":h_ativos,"h_todos":h_todos,
            "prod_ciclo_op":  h_ciclo/h_ativos if h_ativos>0 else 0,
            "prod_ciclo_tot": h_ciclo/h_todos  if h_todos>0 else 0,
            "prod_labor_op":  h_labor/h_ativos if h_ativos>0 else 0,
            "prod_labor_tot": h_labor/h_todos  if h_todos>0 else 0,
            "dias":d,"hA":hA,"hB":hB,"hC":hC,"heA":heA,"heB":heB,"heC":heC,
            "thr_A":thr_A,"thr_B":thr_B,"thr_C":thr_C,
            "minA":d*hA*60,"minB":d*hB*60,"minC":d*hC*60,
        }
    if retornar_intermediarios:
        return resultados, df, agg
    return resultados

def agregar_ano(res_dict, meses_lista):
    rr = [res_dict.get(m) for m in meses_lista if res_dict.get(m)]
    if not rr: return None
    n = len(rr)
    sh_ciclo = sum(r["h_ciclo"]  for r in rr)
    sh_labor = sum(r["h_labor"]  for r in rr)
    sh_ativos= sum(r["h_ativos"] for r in rr)
    sh_todos = sum(r["h_todos"]  for r in rr)
    return {
        "tot_A":  round(sum(r["tot_A"] for r in rr)/n, 1),
        "tot_B":  round(sum(r["tot_B"] for r in rr)/n, 1),
        "tot_C":  round(sum(r["tot_C"] for r in rr)/n, 1),
        "total":  round(sum(r["total"]  for r in rr)/n, 1),
        "prod_labor_tot": sh_labor/sh_todos  if sh_todos>0  else 0,
        "prod_ciclo_tot": sh_ciclo/sh_todos  if sh_todos>0  else 0,
        "prod_labor_op":  sh_labor/sh_ativos if sh_ativos>0 else 0,
        "prod_ciclo_op":  sh_ciclo/sh_ativos if sh_ativos>0 else 0,
    }

def show_tabela(r):
    dias=r["dias"]; hA,hB,hC=r["hA"],r["hB"],r["hC"]
    heA,heB,heC=r.get("heA",hA),r.get("heB",hB),r.get("heC",hC)
    rows=[]
    for _,row in r["centros"].iterrows():
        rows.append({"Centro":row.centro,
            "Ocup. A":row.ocup_A,"Ocup. B":row.ocup_B,"Ocup. C":row.ocup_C,
            "Ativo A":int(row.num_A) if hasattr(row,"num_A") else int(row.ativo_A),
            "Ativo B":int(row.num_B) if hasattr(row,"num_B") else int(row.ativo_B),
            "Ativo C":int(row.num_C) if hasattr(row,"num_C") else int(row.ativo_C),
            "Horas A":round(row.horas_disp_A,2),"Horas B":round(row.horas_disp_B,2),"Horas C":round(row.horas_disp_C,2)})
    df=pd.DataFrame(rows)
    def sr(row):
        s=[""]*len(row)
        for i,col in enumerate(df.columns):
            v=row.iloc[i]
            if col in("Ocup. A","Ocup. B","Ocup. C"):
                if v>1.0: s[i]="background-color:#FFCDD2;color:#B71C1C;font-weight:600"
                elif v>=0.85: s[i]="background-color:#FFFDE7;color:#7B5800;font-weight:600"
                else: s[i]=f"background-color:#E8F5E9;color:{JD_VERDE_ESC};font-weight:600"
            elif col in("Ativo A","Ativo B","Ativo C"):
                s[i]="background-color:#B3E5FC;color:#01579B;font-weight:700" if v else "background-color:#FFFDE7;color:#888"
            elif col in("Horas A","Horas B","Horas C"):
                s[i]="background-color:#B3E5FC;color:#01579B" if v>0 else "background-color:#F5F5F5;color:#AAA"
        return s
    st.dataframe(df.style.apply(sr,axis=1).format({
        "Ocup. A":"{:.0%}","Ocup. B":"{:.0%}","Ocup. C":"{:.0%}",
        "Horas A":"{:.1f}","Horas B":"{:.1f}","Horas C":"{:.1f}"}),
        use_container_width=True,hide_index=True)
    sup=r["suporte"]
    srows=[]
    for nm,k in [("Lavadora e Inspeção","lavadora"),("Gravação e Estanqueidade","gravacao"),
                 ("Preset","preset"),("Coringa","coringa"),("Facilitador","facilitador")]:
        s=sup[k]
        srows.append({"Função":nm,"Qtd A":s["A"],"Qtd B":s["B"],"Qtd C":s["C"],
            "Horas A":round(s["A"]*heA*dias,1),"Horas B":round(s["B"]*heB*dias,1),"Horas C":round(s["C"]*heC*dias,1)})
    srows.append({"Função":"▶ TOTAL POR TURNO",
        "Qtd A":r["tot_A"],"Qtd B":r["tot_B"],"Qtd C":r["tot_C"],
        "Horas A":round(r["tot_A"]*heA*dias,1),"Horas B":round(r["tot_B"]*heB*dias,1),"Horas C":round(r["tot_C"]*heC*dias,1)})
    df_s=pd.DataFrame(srows)
    def ss(row):
        is_t="TOTAL" in str(row["Função"])
        return [f"background-color:{JD_AMARELO};color:{JD_VERDE_ESC};font-weight:700" if is_t else ""]*len(row)
    st.dataframe(df_s.style.apply(ss,axis=1).format({"Horas A":"{:.1f}","Horas B":"{:.1f}","Horas C":"{:.1f}"}),
        use_container_width=True,hide_index=True)
    c1,c2,c3,c4=st.columns(4)
    c1.metric("Total funcionários",r["total"])
    c2.metric("Ciclo operacional",f"{r['prod_ciclo_op']:.0%}")
    c3.metric("Labor operacional",f"{r['prod_labor_op']:.0%}")
    c4.metric("⭐ Labor total",f"{r['prod_labor_tot']:.0%}")

def grafico_cenarios(cenarios_dict):
    fig=make_subplots(specs=[[{"secondary_y":True}]])
    cores_A=[JD_VERDE,"#66BB6A","#A5D6A7","#1B5E20"]
    cores_B=[JD_AMARELO_ESC,"#FFD54F","#FFE082","#FF6F00"]
    cores_C=["#1565C0","#64B5F6","#BBDEFB","#0D47A1"]
    cores_p=["#C62828","#FF6D00","#7B1FA2","#00695C"]
    for i,(nome,res) in enumerate(cenarios_dict.items()):
        mv,tA,tB,tC,prod=[],[],[],[],[]
        for m,abr in zip(MESES,MESES_ABREV):
            r=res.get(m)
            if not r: continue
            mv.append(abr); tA.append(r["tot_A"]); tB.append(r["tot_B"])
            tC.append(r["tot_C"]); prod.append(r["prod_labor_tot"]*100)
        if res.get("__ANO__"):
            r=res["__ANO__"]
            mv.append("ANO"); tA.append(r["tot_A"]); tB.append(r["tot_B"])
            tC.append(r["tot_C"]); prod.append(r["prod_labor_tot"]*100)
        op=0.9 if i==0 else 0.65
        fig.add_trace(go.Bar(name=f"A—{nome}",x=mv,y=tA,marker_color=cores_A[i%4],opacity=op,
            offsetgroup=i,legendgroup=nome,text=tA,textposition="inside",textfont=dict(color="white",size=9)),secondary_y=False)
        fig.add_trace(go.Bar(name=f"B—{nome}",x=mv,y=tB,marker_color=cores_B[i%4],opacity=op,
            offsetgroup=i,legendgroup=nome,base=tA,text=tB,textposition="inside",textfont=dict(size=9)),secondary_y=False)
        fig.add_trace(go.Bar(name=f"C—{nome}",x=mv,y=tC,marker_color=cores_C[i%4],opacity=op,
            offsetgroup=i,legendgroup=nome,base=[a+b for a,b in zip(tA,tB)],
            text=tC,textposition="inside",textfont=dict(color="white",size=9)),secondary_y=False)
        fig.add_trace(go.Scatter(name=f"Labor—{nome}",x=mv,y=prod,mode="lines+markers+text",
            marker=dict(color=cores_p[i%4],size=10,symbol="circle" if i==0 else "diamond"),
            line=dict(color=cores_p[i%4],width=2,dash="solid" if i==0 else "dot"),
            text=[f"{p:.0f}%" for p in prod],textposition="top center",
            textfont=dict(color=cores_p[i%4],size=11)),secondary_y=True)
    fig.update_layout(
        barmode="stack",
        title=dict(text="MÃO-DE-OBRA POR TURNO",font=dict(size=14,color=JD_VERDE_ESC)),
        legend=dict(orientation="h",y=-0.32,x=0,font=dict(size=10,color="#000000"),
                    bgcolor="rgba(255,255,255,0.95)",bordercolor="#AAAAAA",borderwidth=1),
        height=480,plot_bgcolor="white",paper_bgcolor="white",
        xaxis=dict(showgrid=False,tickfont=dict(size=11,color="#1A1A1A")),
        yaxis=dict(showgrid=True,gridcolor="#E8E8E8",tickfont=dict(size=11,color="#1A1A1A"),
                   title="Nº Funcionários",title_font=dict(size=12,color="#1A1A1A")),
        yaxis2=dict(title="Labor Total (%)",tickformat=".0f",ticksuffix="%",range=[0,100],
                    tickfont=dict(size=11,color="#1A1A1A"),title_font=dict(size=12,color="#1A1A1A")))
    return fig

def safe_int(v):
    try: return int(float(v)) if v is not None else 0
    except: return 0

def safe_float(v):
    try: return float(v) if v is not None else 0.0
    except: return 0.0

_BRD_DIAG = Border(left=Side(style='thin',color='CCCCCC'),right=Side(style='thin',color='CCCCCC'),
                   top=Side(style='thin',color='CCCCCC'),bottom=Side(style='thin',color='CCCCCC'))
VERDE_HEADER    = PatternFill("solid", fgColor="1F4D19")
VERMELHO_HEADER = PatternFill("solid", fgColor="C62828")
CINZA_HEADER    = PatternFill("solid", fgColor="555555")
VERMELHO_CLARO  = PatternFill("solid", fgColor="FFCDD2")
VERMELHO_ESCURO = PatternFill("solid", fgColor="C62828")
AMARELO         = PatternFill("solid", fgColor="FFDE00")
VERDE           = PatternFill("solid", fgColor="E8F5E9")
F_HDR_VERD  = PatternFill("solid", fgColor="1F4D19")
F_HDR_AMAR  = PatternFill("solid", fgColor="FFDE00")
F_HDR_CINZA = PatternFill("solid", fgColor="555555")
F_CINZA_CLR = PatternFill("solid", fgColor="F4F4F4")
F_BRANCO    = PatternFill("solid", fgColor="FFFFFF")
F_VERM_CLR  = PatternFill("solid", fgColor="FFCDD2")
F_AMAR      = PatternFill("solid", fgColor="FFDE00")
F_VERDE     = PatternFill("solid", fgColor="E8F5E9")
F_VERDE_MED = PatternFill("solid", fgColor="C8E6C9")
F_AZUL_CLR  = PatternFill("solid", fgColor="E3F2FD")
C_VERDE     = PatternFill("solid", fgColor="92D050")
C_AMAR      = PatternFill("solid", fgColor="FFFF00")
C_AZUL      = PatternFill("solid", fgColor="00B0F0")
C_PRETO     = PatternFill("solid", fgColor="000000")
C_CINZA     = PatternFill("solid", fgColor="D9D9D9")
C_BRANCO    = PatternFill("solid", fgColor="FFFFFF")
C_ROSA      = PatternFill("solid", fgColor="FFB6C1")
C_VERM_DIV  = PatternFill("solid", fgColor="FF0000")

def cor_ocup(pct):
    try:
        v = float(pct)
        if v >= 1.06: return PatternFill("solid", fgColor="FF0000")
        if v >= 1.00: return PatternFill("solid", fgColor="FFFF00")
        if v >= 0.40: return PatternFill("solid", fgColor="92D050")
        return PatternFill("solid", fgColor="FFFFFF")
    except: return PatternFill("solid", fgColor="FFFFFF")

def ec(ws, r, c, val, fill=None, bold=False, color="000000", size=9, center=True, italic=False, comment_text=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = Font(name="Arial", bold=bold, color=color, size=size, italic=italic)
    cell.fill = fill or F_BRANCO
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center", wrap_text=True)
    cell.border = _BRD_DIAG
    return cell

def cell_style(ws, r, c, val, fill=None, bold=False, color="000000", font_size=9, center=True, italic=False, comment_text=None):
    return ec(ws, r, c, val, fill, bold, color, font_size, center, italic, comment_text)

# ─────────────────────────────────────────
# TABELONA PURA — sem comparação com referência
# ─────────────────────────────────────────
def gerar_tabelona_pura(resultados, tempo, dist, aplic, pmp, dias, horas_turno, horas_efetivas, thresholds):
    import openpyxl as _opx
    from openpyxl.styles import PatternFill as _PF, Font as _Ft, Alignment as _Al, Border as _Bd, Side as _Sd
    _F_VERDE=_PF("solid",fgColor="92D050"); _F_AMAR=_PF("solid",fgColor="FFDE00")
    _F_AZUL=_PF("solid",fgColor="00B0F0"); _F_PRETO=_PF("solid",fgColor="000000")
    _F_CINZA=_PF("solid",fgColor="D9D9D9"); _F_CINZA2=_PF("solid",fgColor="BFBFBF")
    _F_BRANCO=_PF("solid",fgColor="FFFFFF"); _F_VERDE_JD=_PF("solid",fgColor="1F4D19"); _F_VERM=_PF("solid",fgColor="FF0000"); _F_VERM_S=_PF("solid",fgColor="FF9999")
    _F_AMAR_JD=_PF("solid",fgColor="FFDE00")
    _BRD=_Bd(left=_Sd(style="thin",color="AAAAAA"),right=_Sd(style="thin",color="AAAAAA"),
             top=_Sd(style="thin",color="AAAAAA"),bottom=_Sd(style="thin",color="AAAAAA"))
    def _ec(ws,r,c,val,fill=None,bold=False,color="000000",size=8,center=True,wrap=False):
        cell=ws.cell(row=r,column=c,value=val)
        cell.font=_Ft(name="Arial",bold=bold,color=color,size=size)
        cell.fill=fill or _F_BRANCO
        cell.alignment=_Al(horizontal="center" if center else "left",vertical="center",wrap_text=wrap)
        cell.border=_BRD
        return cell
    def _ec_pct(ws,r,c,val,fill=None,bold=False,color="000000",size=8):
        cell=_ec(ws,r,c,val,fill,bold,color,size)
        cell.number_format="0.0000000000%"
        return cell
    def _cor_pct(v):
        try:
            f=float(v)
            if f>=1.06: return _PF("solid",fgColor="FF0000")
            if f>=1.00: return _PF("solid",fgColor="FFFF00")
            if f>=0.40: return _PF("solid",fgColor="92D050")
            return _F_BRANCO
        except: return _F_BRANCO
    hA_t=horas_turno["A"]; hB_t=horas_turno["B"]; hC_t=horas_turno["C"]
    heA_t=horas_efetivas["A"]; heB_t=horas_efetivas["B"]; heC_t=horas_efetivas["C"]
    try:
        df_all_t=(aplic.merge(pmp,on="modelo").merge(tempo,on=["centro","peca"]).merge(dist,on=["centro","peca"]))
        if "pc_trat_x" in df_all_t.columns and "pc_trat_y" in df_all_t.columns:
            df_all_t["pc_trat"]=df_all_t["pc_trat_y"].fillna(df_all_t["pc_trat_x"]).fillna(1.0)
            df_all_t.drop(columns=["pc_trat_x","pc_trat_y"],inplace=True)
        elif "pc_trat_y" in df_all_t.columns: df_all_t.rename(columns={"pc_trat_y":"pc_trat"},inplace=True)
        elif "pc_trat_x" in df_all_t.columns: df_all_t.rename(columns={"pc_trat_x":"pc_trat"},inplace=True)
        if "pc_trat" not in df_all_t.columns: df_all_t["pc_trat"]=1.0
        df_all_t["pc_trat"]=pd.to_numeric(df_all_t["pc_trat"],errors="coerce").fillna(1.0).clip(lower=1.0)
        if "vol_int" not in df_all_t.columns: df_all_t["vol_int"] = 1.0
        df_all_t["vol_int"]    = pd.to_numeric(df_all_t["vol_int"],    errors="coerce").fillna(1.0)
        df_all_t["div_carga"]  = pd.to_numeric(df_all_t["div_carga"],  errors="coerce").fillna(0.0)
        df_all_t["div_volume"] = pd.to_numeric(df_all_t["div_volume"], errors="coerce").fillna(0.0)
        df_all_t["disponib"]   = pd.to_numeric(df_all_t["disponib"],   errors="coerce").fillna(1.0)
        df_all_t["perf_op"]    = pd.to_numeric(df_all_t["perf_op"],    errors="coerce").fillna(1.0) if "perf_op" in df_all_t.columns else 1.0
        df_all_t["indice_ciclo"]=(df_all_t.t_ciclo*df_all_t.div_carga*df_all_t.div_volume*df_all_t.vol_int)/(df_all_t.disponib*df_all_t.perf_op)
        df_all_t["min_ciclo"]=df_all_t.indice_ciclo*df_all_t.qtd
        df_all_t["min_labor"]=df_all_t.t_labor*df_all_t.div_carga*df_all_t.qtd*df_all_t.pc_trat
        agg_cp_t=df_all_t.groupby(["centro","peca","mes"])[["min_ciclo","min_labor"]].sum()
    except: agg_cp_t=pd.DataFrame()
    pares_cp = list(dist[["centro","peca"]].drop_duplicates().itertuples(index=False, name=None))
    modelos_lista = sorted(pmp["modelo"].unique().tolist())

    _dist_idx = {(r.centro, r.peca): r for r in dist.itertuples()}
    _tempo_idx = {(r.centro, r.peca): r for r in tempo.itertuples()}
    _aplic_set = set(zip(aplic.centro, aplic.peca, aplic.modelo))
    _pmp_pivot = pmp.pivot_table(index=["modelo","mes"], values="qtd", aggfunc="sum").to_dict()
    # Prefer pc_trat from tempo ("Quantidade por Veículo"), fall back to aplic
    _tempo_pc_idx = {(r.centro, r.peca): float(r.pc_trat) for r in tempo.itertuples() if hasattr(r, "pc_trat") and pd.notna(r.pc_trat)}
    _aplic_pc_idx = {(r.centro, r.peca): _tempo_pc_idx.get((r.centro, r.peca), float(r.pc_trat) if hasattr(r, "pc_trat") else 1.0) for r in aplic.drop_duplicates(["centro","peca"]).itertuples()}

    wb_out = _opx.Workbook(); primeira_t = True
    for mes_t in MESES:
        d_t = dias.get(mes_t, 0)
        if d_t == 0: continue
        r_auto = resultados.get(mes_t)
        if not r_auto: continue
        minA_t=d_t*hA_t*60; minB_t=d_t*hB_t*60; minC_t=d_t*hC_t*60
        if primeira_t: ws_out=wb_out.active; ws_out.title=mes_t[:10]; primeira_t=False
        else: ws_out=wb_out.create_sheet(mes_t[:10])
        ws_out.freeze_panes="F7"
        _F_CINZA_H=_PF("solid",fgColor="D9D9D9"); _F_VD_H=_PF("solid",fgColor="1F4D19")
        ws_out.merge_cells("A1:O1")
        _ec(ws_out,1,1,f"TOTAIS — {mes_t.upper()}",_F_VD_H,True,"FFFFFF",9,True)
        for ci_h,txt_h,f_h in [(16,"TURNO A",_F_VERDE),(17,"TURNO B",_F_AMAR),(18,"TURNO C",_F_AZUL)]:
            _ec(ws_out,1,ci_h,txt_h,f_h,True,"000000",8,True)
        ws_out.row_dimensions[1].height=14
        ws_out.merge_cells("A2:O2"); _ec(ws_out,2,1,"TOTAL DE MINUTOS",_F_CINZA_H,True,"000000",8,False)
        _ec(ws_out,2,16,minA_t,_F_VERDE,True,"000000",8); _ec(ws_out,2,17,minB_t,_F_AMAR,True,"000000",8); _ec(ws_out,2,18,minC_t,_F_AZUL,True,"000000",8)
        ws_out.row_dimensions[2].height=13
        ws_out.merge_cells("A3:O3"); _ec(ws_out,3,1,"TOTAL DE HORAS",_F_CINZA_H,True,"000000",8,False)
        _ec(ws_out,3,16,minA_t/60,_F_VERDE,True,"000000",8); _ec(ws_out,3,17,minB_t/60,_F_AMAR,True,"000000",8); _ec(ws_out,3,18,minC_t/60,_F_AZUL,True,"000000",8)
        ws_out.row_dimensions[3].height=13
        ws_out.merge_cells("A4:O4"); _ec(ws_out,4,1,"Nº DIAS TRABALHADOS",_F_CINZA_H,True,"000000",8,False)
        _ec(ws_out,4,16,d_t,_F_VERDE,True,"FF0000",9); _ec(ws_out,4,17,d_t,_F_AMAR,True,"FF0000",9); _ec(ws_out,4,18,d_t,_F_AZUL,True,"FF0000",9)
        ws_out.row_dimensions[4].height=13
        n_mod=len(modelos_lista)
        ws_out.merge_cells(f"A5:{get_column_letter(19+n_mod)}5")
        _ec(ws_out,5,1,f"RESUMO DA CARGA — {mes_t.upper()} ({d_t} dias)",_F_VERDE_JD,True,"FFFFFF",10,True)
        ws_out.row_dimensions[5].height=18
        hdrs_f=[("Máquina",_F_CINZA2,"000000"),("PEÇA",_F_CINZA2,"000000"),("DESCRIÇÃO",_F_CINZA2,"000000"),("PÇ/TRAT",_F_CINZA2,"000000"),("UM",_F_CINZA2,"000000"),("Tempo Ciclo (min)",_F_PRETO,"FFFFFF"),("Tempo Labor (min)",_F_PRETO,"FFFFFF"),("Div. Carga",_PF("solid",fgColor="FF0000"),"FFFF00"),("Vol. Interna",_F_CINZA2,"000000"),("Div. Volume",_PF("solid",fgColor="FF0000"),"FFFF00"),("Disponib.",_F_CINZA2,"000000"),("Perf. Op.",_F_CINZA2,"000000"),("Indice Ciclo",_F_CINZA2,"000000"),("JA.A",_F_VERDE,"000000"),("JA.B",_F_AMAR,"000000"),("JA.C",_F_AZUL,"000000"),("TOTAL CICLOS (MIN)",_F_CINZA,"000000"),("TOTAL LABOR (MIN)",_F_CINZA,"000000"),("TOTAL PECAS",_F_CINZA,"000000")]
        largs_t=[9,8,16,6,5,9,9,8,8,8,8,8,9,8,8,8,12,12,8]
        for ci_t,(h_t,f_t,cor_t) in enumerate(hdrs_f,1):
            _ec(ws_out,6,ci_t,h_t,f_t,True,cor_t,8,True,True); ws_out.column_dimensions[get_column_letter(ci_t)].width=largs_t[ci_t-1]
        for mi_t,mod_t in enumerate(modelos_lista):
            ci_t=20+mi_t; _ec(ws_out,6,ci_t,mod_t,_F_CINZA,True,"000000",7,True,True); ws_out.column_dimensions[get_column_letter(ci_t)].width=7
        ws_out.row_dimensions[6].height=42
        _qtd_mes = {mod: int(_pmp_pivot.get("qtd", {}).get((mod, mes_t), 0)) for mod in modelos_lista}
        ri_t=7
        for cen_t,peca_t in pares_cp:
            _dk = (cen_t, peca_t); _dr = _dist_idx.get(_dk); _tr = _tempo_idx.get(_dk)
            if _dr is None or _tr is None: continue
            tc=float(_tr.t_ciclo); tl=float(_tr.t_labor)
            dc=float(_dr.div_carga); vi=float(_dr.vol_int)
            dv=float(_dr.div_volume); di=float(_dr.disponib)
            po=float(_dr.perf_op) if hasattr(_dr, "perf_op") else 1.0
            idx_c=(tc*dc*dv*vi)/(di*po) if (di>0 and po>0) else 0
            app_mod_v={}; tot_pecas=0
            for mod_t2 in modelos_lista:
                qtd_t = _qtd_mes.get(mod_t2, 0)
                flag_t = 1 if (cen_t, peca_t, mod_t2) in _aplic_set else 0
                app_mod_v[mod_t2]=qtd_t*flag_t; tot_pecas+=qtd_t*flag_t
            pc_t = _aplic_pc_idx.get((cen_t, peca_t), 1.0)
            mc_t = idx_c * tot_pecas
            ml_t = tl * dc * pc_t * tot_pecas
            pA_t=mc_t/minA_t if minA_t>0 else 0; pB_t=mc_t/minB_t if minB_t>0 else 0; pC_t=mc_t/minC_t if minC_t>0 else 0
            _ec(ws_out,ri_t,1,cen_t,_F_BRANCO,False,"000000",8,False); _ec(ws_out,ri_t,2,peca_t,_F_BRANCO,False,"000000",8,False)
            _ec(ws_out,ri_t,3,"",_F_BRANCO,False,"000000",8,False); _ec(ws_out,ri_t,4,int(pc_t),_F_BRANCO,False,"000000",8); _ec(ws_out,ri_t,5,"PC",_F_BRANCO,False,"000000",8)
            _ec(ws_out,ri_t,6,tc,_F_PRETO,False,"FFFFFF",8); _ec(ws_out,ri_t,7,tl,_F_PRETO,False,"FFFFFF",8)
            _fill_dc_p = _F_VERM if abs(dc-1.0)>0.001 else _F_BRANCO
            _fill_dv_p = _F_VERM if abs(dv-1.0)>0.001 else _F_BRANCO
            _ec(ws_out,ri_t,8,dc,_fill_dc_p,False,"000000",8); _ec(ws_out,ri_t,9,vi,_F_BRANCO,False,"000000",8)
            _ec(ws_out,ri_t,10,dv,_fill_dv_p,False,"000000",8); _ec(ws_out,ri_t,11,di,_F_BRANCO,False,"000000",8)
            _fill_po_p = _F_VERM if abs(po-1.0)>0.001 else _F_BRANCO
            _ec(ws_out,ri_t,12,po,_fill_po_p,False,"000000",8)
            _ec(ws_out,ri_t,13,idx_c,_F_BRANCO,False,"000000",8)
            _ec_pct(ws_out,ri_t,14,pA_t,_cor_pct(pA_t)); _ec_pct(ws_out,ri_t,15,pB_t,_cor_pct(pB_t)); _ec_pct(ws_out,ri_t,16,pC_t,_cor_pct(pC_t))
            _ec(ws_out,ri_t,17,mc_t,_F_BRANCO,False,"000000",8); _ec(ws_out,ri_t,18,ml_t,_F_BRANCO,False,"000000",8); _ec(ws_out,ri_t,19,tot_pecas,_F_BRANCO,False,"000000",8)
            for mi_t2,mod_t2 in enumerate(modelos_lista):
                ci_t2=20+mi_t2; v_app_t=app_mod_v.get(mod_t2,0)
                _ec(ws_out,ri_t,ci_t2,v_app_t if v_app_t else None,_F_CINZA if v_app_t else _F_BRANCO,False,"000000",7)
            ws_out.row_dimensions[ri_t].height=13; ri_t+=1
        _COL_F=6; _ROW_START=66; _F_CINZA_D=_PF("solid",fgColor="D9D9D9")
        ws_out.merge_cells(start_row=_ROW_START,start_column=_COL_F,end_row=_ROW_START,end_column=_COL_F+9)
        _ec(ws_out,_ROW_START,_COL_F,"DADOS AUTOMÁTICOS",_F_CINZA_D,True,"000000",9,True); ws_out.row_dimensions[_ROW_START].height=14
        _r67=_ROW_START+1
        _ec(ws_out,_r67,_COL_F,"PERÍODO:",_F_CINZA_D,True,"000000",8,False); _ec(ws_out,_r67,_COL_F+1,mes_t,_F_BRANCO,True,"FF0000",9,False)
        _ec(ws_out,_r67,_COL_F+3,"DATA DE REVISÃO:",_F_CINZA_D,True,"000000",8,False); _ec(ws_out,_r67,_COL_F+4,datetime.now().strftime("%d/%m/%Y"),_F_BRANCO,True,"FF0000",9)
        _ec(ws_out,_r67,_COL_F+6,"HORAS POR TURNO DE TRABALHO",_F_CINZA_D,True,"000000",8,True); ws_out.row_dimensions[_r67].height=14
        _r68=_ROW_START+2
        _ec(ws_out,_r68,_COL_F+6,heA_t,_F_CINZA_D,True,"000000",8); _ec(ws_out,_r68,_COL_F+7,heB_t,_F_CINZA_D,True,"000000",8); _ec(ws_out,_r68,_COL_F+8,heC_t,_F_CINZA_D,True,"000000",8); ws_out.row_dimensions[_r68].height=14
        _r69=_ROW_START+3
        for _ci_g,_txt_g,_fg in [(_COL_F+1,"TURNO A",_F_VERDE),(_COL_F+2,"TURNO B",_F_AMAR),(_COL_F+3,"TURNO C",_F_AZUL),(_COL_F+4,"TURNO A",_F_VERDE),(_COL_F+5,"TURNO B",_F_AMAR),(_COL_F+6,"TURNO C",_F_AZUL),(_COL_F+7,"TURNO A",_F_VERDE),(_COL_F+8,"TURNO B",_F_AMAR),(_COL_F+9,"TURNO C",_F_AZUL)]:
            _ec(ws_out,_r69,_ci_g,_txt_g,_fg,True,"000000",8,True)
        ws_out.row_dimensions[_r69].height=14
        _r70=_ROW_START+4; _ec(ws_out,_r70,_COL_F,"Centro",_F_PRETO,True,"FFFFFF",8)
        for _ci_s,_txt_s in [(_COL_F+1,"% Ocup"),(_COL_F+2,"% Ocup"),(_COL_F+3,"% Ocup"),(_COL_F+4,"Nº Func"),(_COL_F+5,"Nº Func"),(_COL_F+6,"Nº Func"),(_COL_F+7,"Horas"),(_COL_F+8,"Horas"),(_COL_F+9,"Horas")]:
            _ec(ws_out,_r70,_ci_s,_txt_s,_F_PRETO,True,"FFFFFF",8)
        ws_out.row_dimensions[_r70].height=14
        _ri_c=_ROW_START+5
        if r_auto:
            for _,_crow in r_auto["centros"].iterrows():
                def _cbg_t(v):
                    if v>=1.06: return _PF("solid",fgColor="FF0000")
                    if v>=1.00: return _PF("solid",fgColor="FFFF00")
                    if v>=0.40: return _PF("solid",fgColor="92D050")
                    return _F_BRANCO
                _ec(ws_out,_ri_c,_COL_F,_crow.centro,_F_BRANCO,False,"000000",8,False)
                _ec_pct(ws_out,_ri_c,_COL_F+1,_crow.ocup_A,_cbg_t(_crow.ocup_A)); _ec_pct(ws_out,_ri_c,_COL_F+2,_crow.ocup_B,_cbg_t(_crow.ocup_B)); _ec_pct(ws_out,_ri_c,_COL_F+3,_crow.ocup_C,_cbg_t(_crow.ocup_C))
                _nA_t=int(_crow.num_A) if hasattr(_crow,"num_A") else int(_crow.ativo_A); _nB_t=int(_crow.num_B) if hasattr(_crow,"num_B") else int(_crow.ativo_B); _nC_t=int(_crow.num_C) if hasattr(_crow,"num_C") else int(_crow.ativo_C)
                _ec(ws_out,_ri_c,_COL_F+4,_nA_t,_F_VERDE if _crow.ativo_A else _F_AMAR,True,"000000",8); _ec(ws_out,_ri_c,_COL_F+5,_nB_t,_F_VERDE if _crow.ativo_B else _F_AMAR,True,"000000",8); _ec(ws_out,_ri_c,_COL_F+6,_nC_t,_F_AZUL if _crow.ativo_C else _F_CINZA,True,"000000",8)
                _ec(ws_out,_ri_c,_COL_F+7,_crow.horas_disp_A if _crow.ativo_A else 0,_F_VERDE if _crow.ativo_A else _F_BRANCO,True,"000000",8)
                _ec(ws_out,_ri_c,_COL_F+8,_crow.horas_disp_B if _crow.ativo_B else 0,_F_AMAR if _crow.ativo_B else _F_BRANCO,True,"000000",8)
                _ec(ws_out,_ri_c,_COL_F+9,_crow.horas_disp_C if _crow.ativo_C else 0,_F_AZUL if _crow.ativo_C else _F_BRANCO,True,"000000",8)
                ws_out.row_dimensions[_ri_c].height=13; _ri_c+=1
            _sup_a=r_auto["suporte"]
            for _snm,_skey in [("TOTAL DE OPERADORES",None),("LAVADORA E INSPEÇÃO","lavadora"),("GRAVAÇÃO E ESTANQUEIDADE","gravacao"),("PRESET","preset"),("CORINGA","coringa"),("FACILITADOR","facilitador"),("TOTAL POR TURNO",None),("TOTAL FUNCIONÁRIOS",None)]:
                _bold_s="TOTAL" in _snm; _bg_s=_F_AMAR_JD if _bold_s else _F_BRANCO; _fg_s="1F4D19" if _bold_s else "000000"
                _ec(ws_out,_ri_c,_COL_F,_snm,_bg_s,_bold_s,_fg_s,8,False)
                if _skey:
                    _sv=_sup_a[_skey]
                    for _ci_sv,_tk in [(_COL_F+4,"A"),(_COL_F+5,"B"),(_COL_F+6,"C")]: _ec(ws_out,_ri_c,_ci_sv,_sv[_tk],_F_VERDE if _tk=="A" else (_F_AMAR if _tk=="B" else _F_AZUL),True,"000000",8)
                    for _ci_hv,_tk,_hef in [(_COL_F+7,"A",heA_t),(_COL_F+8,"B",heB_t),(_COL_F+9,"C",heC_t)]:
                        _hv=_sv[_tk]*_hef*d_t; _ec(ws_out,_ri_c,_ci_hv,_hv if _hv else 0,_F_VERDE if _tk=="A" else (_F_AMAR if _tk=="B" else _F_AZUL),True,"000000",8)
                elif "TOTAL DE OPERADORES" in _snm:
                    for _ci_sv,_vv in [(_COL_F+4,r_auto["op_A"]),(_COL_F+5,r_auto["op_B"]),(_COL_F+6,r_auto["op_C"])]: _ec(ws_out,_ri_c,_ci_sv,_vv,_F_AMAR_JD,True,"1F4D19",8)
                    for _ci_hv,_vv,_hef in [(_COL_F+7,r_auto["op_A"],heA_t),(_COL_F+8,r_auto["op_B"],heB_t),(_COL_F+9,r_auto["op_C"],heC_t)]: _ec(ws_out,_ri_c,_ci_hv,_vv*_hef*d_t,_F_AMAR_JD,True,"1F4D19",8)
                elif "TOTAL POR TURNO" in _snm:
                    for _ci_sv,_vv in [(_COL_F+4,r_auto["tot_A"]),(_COL_F+5,r_auto["tot_B"]),(_COL_F+6,r_auto["tot_C"])]: _ec(ws_out,_ri_c,_ci_sv,_vv,_F_AMAR_JD,True,"1F4D19",8)
                    for _ci_hv,_vv,_hef in [(_COL_F+7,r_auto["tot_A"],heA_t),(_COL_F+8,r_auto["tot_B"],heB_t),(_COL_F+9,r_auto["tot_C"],heC_t)]: _ec(ws_out,_ri_c,_ci_hv,_vv*_hef*d_t,_F_AMAR_JD,True,"1F4D19",8)
                elif "FUNCIONÁRIOS" in _snm:
                    _ec(ws_out,_ri_c,_COL_F+4,r_auto["total"],_F_AMAR_JD,True,"1F4D19",9)
                    _th=r_auto["tot_A"]*heA_t*d_t+r_auto["tot_B"]*heB_t*d_t+r_auto["tot_C"]*heC_t*d_t
                    _ec(ws_out,_ri_c,_COL_F+7,_th,_F_AMAR_JD,True,"1F4D19",9)
                ws_out.row_dimensions[_ri_c].height=13; _ri_c+=1
            _ri_c+=1
            for _pnm,_pv,_dest in [("PROD. CICLO OPERACIONAL",r_auto["prod_ciclo_op"],False),("PROD. CICLO TOTAL",r_auto["prod_ciclo_tot"],False),("PROD. LABOR OPERACIONAL",r_auto["prod_labor_op"],False),("PROD. LABOR TOTAL ★",r_auto["prod_labor_tot"],True)]:
                ws_out.merge_cells(start_row=_ri_c,start_column=_COL_F,end_row=_ri_c,end_column=_COL_F+8)
                _ec(ws_out,_ri_c,_COL_F,_pnm,_F_AMAR_JD if _dest else _F_BRANCO,_dest,"1F4D19" if _dest else "000000",8,False)
                _ec_pct(ws_out,_ri_c,_COL_F+9,_pv,_F_AMAR_JD if _dest else _F_BRANCO)
                ws_out.row_dimensions[_ri_c].height=14; _ri_c+=1
        for _ci_w,_ww in [(_COL_F,14),(_COL_F+1,8),(_COL_F+2,8),(_COL_F+3,8),(_COL_F+4,8),(_COL_F+5,8),(_COL_F+6,8),(_COL_F+7,10),(_COL_F+8,10),(_COL_F+9,10)]:
            ws_out.column_dimensions[get_column_letter(_ci_w)].width=_ww
    _fb_ano = st.session_state.get("_fb_anual")
    _cp_data_ano = build_cp_data_anual(resultados, tempo, dist, aplic, pmp, file_bytes=_fb_ano)
    _horas_ano = read_horas_anual(_fb_ano)
    _pmp_qtd_dict = _pmp_pivot.get("qtd", {})
    _pmp_ano_mod = {mod: sum(int(_pmp_qtd_dict.get((mod, mes_t), 0) or 0) for mes_t in MESES) for mod in modelos_lista}
    _app_mod_ano_t = {
        (_cen_a, _peca_a): {mod: _pmp_ano_mod[mod] if (_cen_a, _peca_a, mod) in _aplic_set else 0 for mod in modelos_lista}
        for _cen_a, _peca_a in pares_cp
    }
    gerar_aba_anual(wb_out, resultados, label="ANO", cp_data=_cp_data_ano, horas_anual=_horas_ano, modelos_lista=modelos_lista, app_mod_ano=_app_mod_ano_t)
    # ── Abas de Input — cópia fiel do arquivo original (formatação idêntica) ──
    _fb_src = st.session_state.get("_fb_anual")
    if _fb_src:
        try:
            from copy import copy as _copy
            _wb_src = _opx.load_workbook(BytesIO(_fb_src), read_only=False, data_only=True)
            _abas_map_inp = st.session_state.get("_abas_map", {})
            for _chave_inp, _nome_dst_inp in [
                ("INPUTPMP",          "INPUT_PMP"),
                ("INPUTTURNOS",       "INPUT_TURNOS"),
                ("INPUTTEMPO",        "INPUT_TEMPO"),
                ("INPUTDISTRIBUIÇÃO", "INPUT_DIST"),
                ("INPUTAPLICAÇÃO",    "INPUT_APLIC"),
            ]:
                _nome_src_inp = _abas_map_inp.get(_chave_inp) or _chave_inp
                if _nome_src_inp not in _wb_src.sheetnames:
                    continue
                _ws_s = _wb_src[_nome_src_inp]
                _ws_d = wb_out.create_sheet(_nome_dst_inp)
                for _col_inp, _cd_inp in _ws_s.column_dimensions.items():
                    _ws_d.column_dimensions[_col_inp].width = _cd_inp.width
                for _row_inp, _rd_inp in _ws_s.row_dimensions.items():
                    _ws_d.row_dimensions[_row_inp].height = _rd_inp.height
                if _ws_s.freeze_panes:
                    _ws_d.freeze_panes = _ws_s.freeze_panes
                for _row_inp in _ws_s.iter_rows():
                    for _cell_inp in _row_inp:
                        _nc = _ws_d.cell(row=_cell_inp.row, column=_cell_inp.column, value=_cell_inp.value)
                        if _cell_inp.has_style:
                            try: _nc.font      = _copy(_cell_inp.font)
                            except: pass
                            try: _nc.fill      = _copy(_cell_inp.fill)
                            except: pass
                            try: _nc.border    = _copy(_cell_inp.border)
                            except: pass
                            try: _nc.alignment = _copy(_cell_inp.alignment)
                            except: pass
                            try: _nc.number_format = _cell_inp.number_format
                            except: pass
                for _merged_inp in list(_ws_s.merged_cells.ranges):
                    _ws_d.merge_cells(str(_merged_inp))
            _wb_src.close()
        except Exception:
            pass
    buf_out = BytesIO(); wb_out.save(buf_out); buf_out.seek(0)
    return buf_out


def build_cp_data_anual(resultados, tempo, dist, aplic, pmp, file_bytes=None):
    if file_bytes is not None:
        try:
            wb_ref = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
            if "AnoFY26" in wb_ref.sheetnames:
                ws_ref = wb_ref["AnoFY26"]
                rows_ref = list(ws_ref.rows)
                result = []
                for row in rows_ref[6:]:
                    vals = [c.value for c in row[:18]]
                    if not vals[0] or not str(vals[0]).startswith("CEN"): continue
                    try:
                        cen  = str(vals[0]); peca = str(vals[1])
                        pc_t = float(vals[3] or 1)
                        tc   = float(vals[5] or 0); tl = float(vals[6] or 0)
                        dc   = float(vals[7] or 0); vi = float(vals[8] or 1)
                        dv   = float(vals[9] or 0); di = float(vals[10] or 1)
                        idx  = float(vals[11] or 0)
                        mc   = float(vals[15] or 0); ml = float(vals[16] or 0)
                        qt   = int(vals[17] or 0)
                        result.append((cen, peca, pc_t, tc, tl, dc, vi, dv, di, idx, mc, ml, qt))
                    except: pass
                wb_ref.close()
                if result: return result
            wb_ref.close()
        except: pass
    if pmp is None or tempo is None or dist is None or aplic is None: return None
    meses_com_resultado = [m for m in MESES if resultados.get(m)]
    meses_com_pmp = [m for m in MESES if not pmp[pmp.mes==m].empty and pmp[pmp.mes==m]["qtd"].sum()>0]
    meses_c = list(dict.fromkeys([m for m in MESES if m in meses_com_resultado or m in meses_com_pmp]))
    if not meses_c: return None
    try:
        df = (aplic.merge(pmp, on="modelo").merge(tempo, on=["centro","peca"]).merge(dist, on=["centro","peca"]))
        if "pc_trat_x" in df.columns and "pc_trat_y" in df.columns:
            df["pc_trat"]=df["pc_trat_y"].fillna(df["pc_trat_x"]).fillna(1.0); df.drop(columns=["pc_trat_x","pc_trat_y"],inplace=True)
        elif "pc_trat_y" in df.columns: df.rename(columns={"pc_trat_y":"pc_trat"},inplace=True)
        elif "pc_trat_x" in df.columns: df.rename(columns={"pc_trat_x":"pc_trat"},inplace=True)
        df["pc_trat"] = pd.to_numeric(df.get("pc_trat",1.0), errors="coerce").fillna(1.0).clip(lower=1.0)
        if "vol_int" not in df.columns: df["vol_int"] = 1.0
        df["vol_int"]    = pd.to_numeric(df["vol_int"],    errors="coerce").fillna(1.0)
        df["div_carga"]  = pd.to_numeric(df["div_carga"],  errors="coerce").fillna(0.0)
        df["div_volume"] = pd.to_numeric(df["div_volume"], errors="coerce").fillna(0.0)
        df["disponib"]   = pd.to_numeric(df["disponib"],   errors="coerce").fillna(1.0)
        df["perf_op"]    = pd.to_numeric(df["perf_op"],    errors="coerce").fillna(1.0) if "perf_op" in df.columns else 1.0
        df["indice_ciclo"] = (df.t_ciclo*df.div_carga*df.div_volume*df.vol_int)/(df.disponib*df.perf_op)
        df["min_ciclo"] = df.indice_ciclo * df.qtd
        df["min_labor"] = df.t_labor * df.div_carga * df.qtd * df.pc_trat
        df_ano = df[df.mes.isin(meses_c)]
        agg = df_ano.groupby(["centro","peca"])[["min_ciclo","min_labor","qtd"]].sum().reset_index()
        attrs = df_ano.drop_duplicates(["centro","peca"])[
            ["centro","peca","t_ciclo","t_labor","div_carga","vol_int","div_volume","disponib","perf_op","indice_ciclo","pc_trat"]
        ].set_index(["centro","peca"])
        result = []
        for _, row in agg.iterrows():
            cen, peca = row.centro, row.peca
            try:
                at = attrs.loc[(cen, peca)]
                tc=float(at.t_ciclo); tl=float(at.t_labor)
                dc=float(at.div_carga); vi=float(at.vol_int)
                dv=float(at.div_volume); di=float(at.disponib)
                po=float(at.perf_op) if hasattr(at,"perf_op") else 1.0
                idx=float(at.indice_ciclo)
                pct=float(at.pc_trat) if hasattr(at,"pc_trat") else 1.0
            except: tc=tl=dc=vi=dv=di=po=0.0; idx=0.0; pct=1.0
            result.append((cen, peca, pct, tc, tl, dc, vi, dv, di, po, idx,
                           float(row.min_ciclo), float(row.min_labor), int(row.qtd)))
        return result
    except: return None

def read_horas_anual(file_bytes):
    if file_bytes is None: return None
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        if "AnoFY26" not in wb.sheetnames:
            wb.close(); return None
        ws = wb["AnoFY26"]
        rows = list(ws.rows)
        l4 = [c.value for c in rows[3]]
        h_ciclo = float(l4[15]) if len(l4) > 15 and l4[15] else 0
        h_labor = float(l4[16]) if len(l4) > 16 and l4[16] else 0
        h_ativos = 0; h_todos = 0
        for row in rows:
            vals = [c.value for c in row]
            txt = str(vals[11]) if len(vals) > 11 and vals[11] else ""
            if "TOTAL DE OPERADORES" in txt:
                hA = float(vals[29]) if len(vals) > 29 and vals[29] else 0
                hB = float(vals[30]) if len(vals) > 30 and vals[30] else 0
                hC = float(vals[31]) if len(vals) > 31 and vals[31] else 0
                h_ativos = hA + hB + hC
            if "TOTAL FUNCION" in txt:
                h_todos = float(vals[29]) if len(vals) > 29 and vals[29] else 0
        wb.close()
        if h_ciclo > 0 and h_todos > 0:
            return {"h_ciclo": h_ciclo, "h_labor": h_labor,
                    "h_ativos": h_ativos, "h_todos": h_todos}
        return None
    except: return None

def calcular_ano_fy26(file_bytes, overrides_ano, horas_efetivas, suporte_cfg, horas_turno):
    if file_bytes is None: return None
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        if "AnoFY26" not in wb.sheetnames:
            wb.close(); return None
        ws = wb["AnoFY26"]
        rows = list(ws.rows)
        l2  = [c.value for c in rows[1]]
        l5  = [c.value for c in rows[4]]
        l4d = [c.value for c in rows[3]]
        minA_ano = float(l2[12]) if l2[12] else 96300.0
        minB_ano = float(l2[13]) if l2[13] else 182970.0
        minC_ano = float(l2[14]) if l2[14] else 250380.0
        hA = float(l5[12]) if l5[12] else 7.5
        hB = float(l5[13]) if l5[13] else 14.25
        hC = float(l5[14]) if l5[14] else 19.5
        dias_ano = int(l4d[11]) if l4d[11] else 214
        heA = horas_efetivas.get("A", hA)
        heB = horas_efetivas.get("B", hB)
        heC = horas_efetivas.get("C", hC)
        thr_A = 0.0; thr_B = 0.0; thr_C = 0.0
        from collections import defaultdict
        cen_mc = defaultdict(float); cen_ml = defaultdict(float)
        for row in rows[6:]:
            vals = [c.value for c in row[:18]]
            if not vals[0] or not str(vals[0]).startswith("CEN"): continue
            cen = str(vals[0])
            cen_mc[cen] += float(vals[15] or 0)
            cen_ml[cen] += float(vals[16] or 0)
        cen_base = {}
        wb.close()
        centros = []
        for cen in sorted(cen_mc.keys()):
            mc = cen_mc[cen]; ml = cen_ml[cen]
            base = cen_base.get(cen, {})
            oA = mc / minA_ano if minA_ano > 0 else 0
            oB = mc / minB_ano if minB_ano > 0 else 0
            oC = mc / minC_ano if minC_ano > 0 else 0
            aA = base.get("aA", 1 if oA > 0.40 else 0)
            aB = base.get("aB", 1 if oA > 0.75 else 0)
            aC = base.get("aC", 1 if oB > 0.75 else 0)
            nA = aA; nB = aB; nC = aC
            if overrides_ano and cen in overrides_ano:
                ov = overrides_ano[cen]
                if "A" in ov: nA = int(ov["A"]); aA = 1 if nA > 0 else 0
                if "B" in ov: nB = int(ov["B"]); aB = 1 if nB > 0 else 0
                if "C" in ov: nC = int(ov["C"]); aC = 1 if nC > 0 else 0
            centros.append({
                "centro": cen, "ocup_A": oA, "ocup_B": oB, "ocup_C": oC,
                "ativo_A": aA, "ativo_B": aB, "ativo_C": aC,
                "num_A": nA, "num_B": nB, "num_C": nC,
                "min_ciclo_total": mc, "min_labor_total": ml,
                "min_disp_A": minA_ano, "min_disp_B": minB_ano, "min_disp_C": minC_ano,
                "horas_ciclo": mc / 60, "horas_labor": ml / 60,
                "horas_disp_A": dias_ano * heA * nA,
                "horas_disp_B": dias_ano * heB * nB,
                "horas_disp_C": dias_ano * heC * nC,
            })
        df_c = pd.DataFrame(centros)
        op_A = int(df_c.num_A.sum()); op_B = int(df_c.num_B.sum()); op_C = int(df_c.num_C.sum())
        def _sup(key, t, op_count):
            cfg = suporte_cfg[key]
            if op_count == 0: return 0
            if cfg["modo"] == "auto":
                defaults = {"lavadora":{"A":1,"B":1,"C":0},"gravacao":{"A":1,"B":1,"C":0},
                            "preset":{"A":2,"B":1,"C":1},"coringa":{"A":1,"B":0,"C":0},
                            "facilitador":{"A":1,"B":1,"C":0}}
                return defaults[key][t]
            return cfg[t] if op_count > 0 else 0
        lav={t:_sup("lavadora",t,[op_A,op_B,op_C]["ABC".index(t)]) for t in "ABC"}
        gra={t:_sup("gravacao",t,[op_A,op_B,op_C]["ABC".index(t)]) for t in "ABC"}
        pre={t:_sup("preset",t,[op_A,op_B,op_C]["ABC".index(t)]) for t in "ABC"}
        cor={t:_sup("coringa",t,[op_A,op_B,op_C]["ABC".index(t)]) for t in "ABC"}
        fac={t:_sup("facilitador",t,[op_A,op_B,op_C]["ABC".index(t)]) for t in "ABC"}
        tot_A = op_A+lav["A"]+gra["A"]+pre["A"]+cor["A"]+fac["A"]
        tot_B = op_B+lav["B"]+gra["B"]+pre["B"]+cor["B"]+fac["B"]
        tot_C = op_C+lav["C"]+gra["C"]+pre["C"]+cor["C"]+fac["C"]
        h_ciclo  = float(df_c.horas_ciclo.sum())
        h_labor  = float(df_c.horas_labor.sum())
        h_ativos = float((df_c.horas_disp_A + df_c.horas_disp_B + df_c.horas_disp_C).sum())
        h_todos  = tot_A * dias_ano * heA + tot_B * dias_ano * heB + tot_C * dias_ano * heC
        return {
            "centros": df_c,
            "op_A": op_A, "op_B": op_B, "op_C": op_C,
            "tot_A": tot_A, "tot_B": tot_B, "tot_C": tot_C, "total": tot_A + tot_B + tot_C,
            "suporte": {"lavadora":lav,"gravacao":gra,"preset":pre,"coringa":cor,"facilitador":fac},
            "h_ciclo": h_ciclo, "h_labor": h_labor, "h_ativos": h_ativos, "h_todos": h_todos,
            "prod_ciclo_op":  h_ciclo / h_ativos if h_ativos > 0 else 0,
            "prod_ciclo_tot": h_ciclo / h_todos  if h_todos  > 0 else 0,
            "prod_labor_op":  h_labor / h_ativos if h_ativos > 0 else 0,
            "prod_labor_tot": h_labor / h_todos  if h_todos  > 0 else 0,
            "dias": dias_ano, "hA": hA, "hB": hB, "hC": hC,
            "heA": heA, "heB": heB, "heC": heC,
            "thr_A": thr_A, "thr_B": thr_B, "thr_C": thr_C,
            "minA": minA_ano, "minB": minB_ano, "minC": minC_ano,
        }
    except Exception as _e:
        return None

def build_cp_data_from_meses(resultados, tempo, dist, aplic, pmp, dias_por_mes, horas_turno, horas_efetivas, overrides_ano=None, suporte_cfg=None):
    from collections import defaultdict
    meses_com_dados = [m for m in MESES if resultados.get(m)]
    if not meses_com_dados: return None, None
    try:
        df = (aplic.merge(pmp, on="modelo").merge(tempo, on=["centro","peca"]).merge(dist, on=["centro","peca"]))
        # Resolver conflitos de colunas _x/_y gerados por overlaps entre tempo e dist
        for _col in ["pc_trat","div_carga","vol_int","div_volume","disponib","perf_op","t_ciclo","t_labor"]:
            if f"{_col}_y" in df.columns:
                df[_col] = df[f"{_col}_y"]
                df.drop(columns=[c for c in [f"{_col}_x", f"{_col}_y"] if c in df.columns], inplace=True)
            elif f"{_col}_x" in df.columns:
                df.rename(columns={f"{_col}_x": _col}, inplace=True)
        df["pc_trat"] = pd.to_numeric(df.get("pc_trat",1.0), errors="coerce").fillna(1.0).clip(lower=1.0)
        df["vol_int"]    = pd.to_numeric(df.get("vol_int",1),    errors="coerce").fillna(1.0)
        df["div_carga"]  = pd.to_numeric(df["div_carga"],  errors="coerce").fillna(0.0)
        df["div_volume"] = pd.to_numeric(df["div_volume"], errors="coerce").fillna(0.0)
        df["disponib"]   = pd.to_numeric(df["disponib"],   errors="coerce").fillna(1.0)
        df["perf_op"]    = pd.to_numeric(df["perf_op"],    errors="coerce").fillna(1.0) if "perf_op" in df.columns else 1.0
        df["indice_ciclo"] = (df.t_ciclo*df.div_carga*df.div_volume*df.vol_int)/(df.disponib*df.perf_op)
        df_ano = df[df.mes.isin(meses_com_dados)]
        agg = df_ano.groupby(["centro","peca"])[["qtd"]].sum().reset_index()
        attrs = df_ano.drop_duplicates(["centro","peca"])[
            ["centro","peca","t_ciclo","t_labor","div_carga","vol_int","div_volume","disponib","perf_op","indice_ciclo","pc_trat"]
        ].set_index(["centro","peca"])
        cp_data = []
        for _, row in agg.iterrows():
            cen, peca = row.centro, row.peca
            try:
                at = attrs.loc[(cen,peca)]
                tc=float(at.t_ciclo); tl=float(at.t_labor)
                dc=float(at.div_carga); vi=float(at.vol_int)
                dv=float(at.div_volume); di=float(at.disponib)
                po=float(at.perf_op) if hasattr(at,"perf_op") else 1.0
                idx=float(at.indice_ciclo)
                pct=float(at.pc_trat) if hasattr(at,'pc_trat') else 1.0
                qt=int(row.qtd)
                mc=idx*qt; ml=tl*dc*qt*pct
            except: continue
            cp_data.append((cen, peca, pct, tc, tl, dc, vi, dv, di, po, idx, mc, ml, qt))
        if not cp_data: return None, None
        dias_total = sum(dias_por_mes.get(m, 0) for m in meses_com_dados)
        hA = horas_turno.get("A", 7.5); hB = horas_turno.get("B", 14.25); hC = horas_turno.get("C", 19.5)
        heA = horas_efetivas.get("A", hA); heB = horas_efetivas.get("B", hB); heC = horas_efetivas.get("C", hC)
        minA = dias_total * hA * 60; minB = dias_total * hB * 60; minC = dias_total * hC * 60
        cen_mc = defaultdict(float); cen_ml = defaultdict(float)
        for (_cen,_peca,_pct,_tc,_tl,_dc,_vi,_dv,_di,_po,_idx,_mc,_ml,_qt) in cp_data:
            cen_mc[_cen]+=_mc; cen_ml[_cen]+=_ml
        cen_ativos = defaultdict(lambda: {"aA":0,"aB":0,"aC":0})
        for m in meses_com_dados:
            r = resultados[m]
            for _, row in r["centros"].iterrows():
                cen = row.centro
                cen_ativos[cen]["aA"] = max(cen_ativos[cen]["aA"], 1 if int(row.ativo_A) > 0 else 0)
                cen_ativos[cen]["aB"] = max(cen_ativos[cen]["aB"], 1 if int(row.ativo_B) > 0 else 0)
                cen_ativos[cen]["aC"] = max(cen_ativos[cen]["aC"], 1 if int(row.ativo_C) > 0 else 0)
        centros = []
        for cen in sorted(cen_mc.keys()):
            mc=cen_mc[cen]; ml=cen_ml[cen]
            oA=mc/minA if minA>0 else 0; oB=mc/minB if minB>0 else 0; oC=mc/minC if minC>0 else 0
            ca = cen_ativos.get(cen,{"aA":0,"aB":0,"aC":0})
            aA=ca["aA"]; aB=ca["aB"]; aC=ca["aC"]
            nA = aA; nB = aB; nC = aC
            if overrides_ano and cen in overrides_ano:
                ov = overrides_ano[cen]
                if "A" in ov: nA = int(ov["A"]); aA = 1 if nA > 0 else 0
                if "B" in ov: nB = int(ov["B"]); aB = 1 if nB > 0 else 0
                if "C" in ov: nC = int(ov["C"]); aC = 1 if nC > 0 else 0
            centros.append({"centro":cen,"ocup_A":oA,"ocup_B":oB,"ocup_C":oC,
                            "ativo_A":aA,"ativo_B":aB,"ativo_C":aC,
                            "num_A":nA,"num_B":nB,"num_C":nC,
                            "min_ciclo_total":mc,"min_labor_total":ml,
                            "min_disp_A":minA,"min_disp_B":minB,"min_disp_C":minC,
                            "horas_ciclo":mc/60,"horas_labor":ml/60,
                            "horas_disp_A":dias_total*heA*nA,
                            "horas_disp_B":dias_total*heB*nB,
                            "horas_disp_C":dias_total*heC*nC})
        df_c = pd.DataFrame(centros)
        op_A=int(df_c.num_A.sum()); op_B=int(df_c.num_B.sum()); op_C=int(df_c.num_C.sum())
        h_ciclo=float(df_c.horas_ciclo.sum()); h_labor=float(df_c.horas_labor.sum())
        h_ativos=float((df_c.horas_disp_A+df_c.horas_disp_B+df_c.horas_disp_C).sum())
        if suporte_cfg:
            def _sup2(key,t,op_count):
                cfg=suporte_cfg[key]
                if op_count==0: return 0
                if cfg["modo"]=="auto":
                    defaults={"lavadora":{"A":1,"B":1,"C":0},"gravacao":{"A":1,"B":1,"C":0},
                              "preset":{"A":2,"B":1,"C":1},"coringa":{"A":1,"B":0,"C":0},
                              "facilitador":{"A":1,"B":1,"C":0}}
                    return defaults[key][t]
                return cfg[t]
            sup_d={k:{t:_sup2(k,t,[op_A,op_B,op_C]["ABC".index(t)]) for t in "ABC"}
                   for k in ["lavadora","gravacao","preset","coringa","facilitador"]}
            sup_tot_A=sum(sup_d[k]["A"] for k in sup_d); sup_tot_B=sum(sup_d[k]["B"] for k in sup_d); sup_tot_C=sum(sup_d[k]["C"] for k in sup_d)
        else:
            _defs={"lavadora":{"A":1,"B":1,"C":0},"gravacao":{"A":1,"B":1,"C":0},
                   "preset":{"A":2,"B":1,"C":1},"coringa":{"A":1,"B":0,"C":0},"facilitador":{"A":1,"B":1,"C":0}}
            _ops={"A":op_A,"B":op_B,"C":op_C}
            sup_d={k:{t:(_defs[k][t] if _ops[t]>0 else 0) for t in "ABC"} for k in _defs}
            sup_tot_A=sum(sup_d[k]["A"] for k in sup_d); sup_tot_B=sum(sup_d[k]["B"] for k in sup_d); sup_tot_C=sum(sup_d[k]["C"] for k in sup_d)
        tot_A=op_A+sup_tot_A; tot_B=op_B+sup_tot_B; tot_C=op_C+sup_tot_C
        h_todos=tot_A*dias_total*heA+tot_B*dias_total*heB+tot_C*dias_total*heC
        res_ano = {"centros":df_c,"op_A":op_A,"op_B":op_B,"op_C":op_C,
                   "tot_A":tot_A,"tot_B":tot_B,"tot_C":tot_C,"total":tot_A+tot_B+tot_C,
                   "suporte":sup_d,
                   "h_ciclo":h_ciclo,"h_labor":h_labor,"h_ativos":h_ativos,"h_todos":h_todos,
                   "prod_ciclo_op":h_ciclo/h_ativos if h_ativos>0 else 0,
                   "prod_ciclo_tot":h_ciclo/h_todos if h_todos>0 else 0,
                   "prod_labor_op":h_labor/h_ativos if h_ativos>0 else 0,
                   "prod_labor_tot":h_labor/h_todos if h_todos>0 else 0,
                   "dias":dias_total,"hA":hA,"hB":hB,"hC":hC,
                   "heA":heA,"heB":heB,"heC":heC,
                   "thr_A":0.0,"thr_B":0.0,"thr_C":0.0,
                   "minA":minA,"minB":minB,"minC":minC}
        return cp_data, res_ano
    except: return None, None


def show_memoria(r, mes, df_intermediario, agg, horas_turno, thresholds):
    st.markdown(f'<div class="jd-section">Memória de cálculo — {mes}</div>', unsafe_allow_html=True)
    sup=r["suporte"]; d=r["dias"]; hA,hB,hC=r["hA"],r["hB"],r["hC"]
    heA,heB,heC=r.get("heA",hA),r.get("heB",hB),r.get("heC",hC)
    st.markdown('<div class="mem-step"><span class="step-num">1</span> <b>Inputs utilizados</b></div>', unsafe_allow_html=True)
    c1,c2,c3=st.columns(3)
    c1.metric("Turno A",f"{r['minA']:.0f} min",f"{d}×{hA}×60"); c2.metric("Turno B",f"{r['minB']:.0f} min",f"{d}×{hB}×60"); c3.metric("Turno C",f"{r['minC']:.0f} min")
    df_inp=df_intermediario[df_intermediario.mes==mes][["centro","peca","modelo","t_ciclo","t_labor","div_carga","div_volume","vol_int","disponib","perf_op","qtd"]].head(8).copy()
    df_inp.columns=["Centro","Peça","Modelo","T.Ciclo","T.Labor","Div.Carga","Div.Volume","Vol.Int","Disponib","Perf.Op","Qtd"]
    st.dataframe(df_inp.reset_index(drop=True),use_container_width=True,hide_index=True)
    st.markdown('<div class="mem-step"><span class="step-num">2</span> <b>Índice de ciclo</b></div>', unsafe_allow_html=True)
    st.markdown('<div class="formula-box">indice_ciclo = (t_ciclo × div_carga × div_volume × vol_interna) ÷ (disponibilidade × performance_operador)</div>', unsafe_allow_html=True)
    st.markdown('<div class="mem-step"><span class="step-num">3</span> <b>Minutos por linha</b></div>', unsafe_allow_html=True)
    st.markdown('<div class="formula-box">min_ciclo = indice_ciclo × qtd_pecas<br>min_labor = t_labor × div_carga × pç_trat × qtd_pecas</div>', unsafe_allow_html=True)
    st.markdown('<div class="mem-step"><span class="step-num">4</span> <b>Agrupamento e ocupação por centro</b></div>', unsafe_allow_html=True)
    df_p4=r["centros"][["centro","min_ciclo_total","ocup_A","ocup_B","ocup_C"]].copy()
    df_p4.columns=["Centro","Σ Min.Ciclo","Ocup. A","Ocup. B","Ocup. C"]
    st.dataframe(df_p4.reset_index(drop=True).style.format({"Ocup. A":"{:.1%}","Ocup. B":"{:.1%}","Ocup. C":"{:.1%}","Σ Min.Ciclo":"{:.1f}"}),use_container_width=True,hide_index=True)
    st.markdown('<div class="mem-step"><span class="step-num">5</span> <b>Ativação de turno</b></div>', unsafe_allow_html=True)
    st.markdown(f"- Turno A abre se ocup_A > **{thresholds['A']}%**\n- Turno B abre se ocup_A > **{thresholds['B']}%**\n- Turno C abre se ocup_B > **{thresholds['C']}%**")
    st.markdown('<div class="mem-step"><span class="step-num">6</span> <b>Total por turno</b></div>', unsafe_allow_html=True)
    tot_data={"Função":["Operadores CEN","Lavadora","Gravação","Preset","Coringa","Facilitador","TOTAL ✓"],"Turno A":[r["op_A"],sup["lavadora"]["A"],sup["gravacao"]["A"],sup["preset"]["A"],sup["coringa"]["A"],sup["facilitador"]["A"],r["tot_A"]],"Turno B":[r["op_B"],sup["lavadora"]["B"],sup["gravacao"]["B"],sup["preset"]["B"],sup["coringa"]["B"],sup["facilitador"]["B"],r["tot_B"]],"Turno C":[r["op_C"],sup["lavadora"]["C"],sup["gravacao"]["C"],sup["preset"]["C"],sup["coringa"]["C"],sup["facilitador"]["C"],r["tot_C"]]}
    st.dataframe(pd.DataFrame(tot_data),use_container_width=True,hide_index=True)
    st.markdown('<div class="mem-step"><span class="step-num">7</span> <b>Produtividades</b></div>', unsafe_allow_html=True)
    st.markdown('<div class="formula-box">Labor Total ★ = horas_labor ÷ horas_todos</div>', unsafe_allow_html=True)
    prod_data={"Indicador":["Ciclo Operacional","Ciclo Total","Labor Operacional","⭐ Labor Total"],"Resultado":[f"{r['prod_ciclo_op']:.1%}",f"{r['prod_ciclo_tot']:.1%}",f"{r['prod_labor_op']:.1%}",f"{r['prod_labor_tot']:.1%}"]}
    st.dataframe(pd.DataFrame(prod_data),use_container_width=True,hide_index=True)
    _buf_mem_key = f"mem_base_{mes}"
    if st.session_state.get(_buf_mem_key) is None:
        _buf_mem=BytesIO(); df_intermediario[df_intermediario.mes==mes].to_excel(_buf_mem,index=False); _buf_mem.seek(0)
        st.session_state[_buf_mem_key]=_buf_mem.read()
    st.download_button("📥 Baixar base completa pós-JOIN",data=st.session_state[_buf_mem_key],file_name=f"base_{mes}.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",key=f"dl_mem_{mes}")

def show_memoria_ano(res_base, df_intermediario, agg, horas_turno, thresholds):
    meses_ativos = [(m, res_base[m]) for m in MESES if res_base.get(m)]
    if not meses_ativos:
        st.warning("Nenhum mês calculado."); return
    st.markdown('<div class="jd-section">Memória de cálculo — 📅 ANO COMPLETO</div>', unsafe_allow_html=True)
    st.markdown('<div class="mem-step"><span class="step-num">1</span> <b>Inputs consolidados do ano</b></div>', unsafe_allow_html=True)
    n_meses = len(meses_ativos)
    dias_ano = sum(r["dias"] for _, r in meses_ativos)
    hA = meses_ativos[0][1]["hA"]; hB = meses_ativos[0][1]["hB"]; hC = meses_ativos[0][1]["hC"]
    heA = meses_ativos[0][1].get("heA", hA); heB = meses_ativos[0][1].get("heB", hB); heC = meses_ativos[0][1].get("heC", hC)
    minA_ano = dias_ano * hA * 60; minB_ano = dias_ano * hB * 60; minC_ano = dias_ano * hC * 60
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Meses calculados", n_meses)
    c2.metric("Dias trabalhados (total)", dias_ano)
    c3.metric("Min. disponíveis Turno A", f"{minA_ano:,.0f}")
    c4.metric("Min. disponíveis Turno B/C", f"{minB_ano:,.0f}")
    st.markdown('<div class="mem-step"><span class="step-num">2</span> <b>Índice de ciclo (igual ao mensal)</b></div>', unsafe_allow_html=True)
    st.markdown('<div class="formula-box">indice_ciclo = (t_ciclo × div_carga × div_volume × vol_interna) ÷ (disponibilidade × performance_operador)<br>min_ciclo_ano = Σ (indice_ciclo × pç_trat × qtd) em todos os meses</div>', unsafe_allow_html=True)
    st.markdown('<div class="mem-step"><span class="step-num">3</span> <b>Minutos acumulados por centro (todos os meses)</b></div>', unsafe_allow_html=True)
    from collections import defaultdict
    cen_mc = defaultdict(float); cen_ml = defaultdict(float)
    for _, r in meses_ativos:
        for _, row in r["centros"].iterrows():
            cen_mc[row.centro] += row.min_ciclo_total
            cen_ml[row.centro] += row.min_labor_total
    rows_cen = []
    for cen in sorted(cen_mc.keys()):
        mc = cen_mc[cen]; ml = cen_ml[cen]
        pA = mc / minA_ano if minA_ano > 0 else 0
        pB = mc / minB_ano if minB_ano > 0 else 0
        pC = mc / minC_ano if minC_ano > 0 else 0
        thr_A = thresholds["A"] / 100; thr_B = thresholds["B"] / 100; thr_C = thresholds["C"] / 100
        aA = 1 if pA > thr_A else 0; aB = 1 if pA > thr_B else 0; aC = 1 if pB > thr_C else 0
        rows_cen.append({"Centro": cen, "Σ Min.Ciclo (ano)": round(mc, 1), "Σ Min.Labor (ano)": round(ml, 1),
                         "Ocup. A (ano)": pA, "Ocup. B (ano)": pB, "Ocup. C (ano)": pC,
                         "Ativo A": aA, "Ativo B": aB, "Ativo C": aC})
    df_ano_cen = pd.DataFrame(rows_cen)
    st.dataframe(df_ano_cen.style.format({"Ocup. A (ano)": "{:.1%}", "Ocup. B (ano)": "{:.1%}", "Ocup. C (ano)": "{:.1%}", "Σ Min.Ciclo (ano)": "{:,.1f}", "Σ Min.Labor (ano)": "{:,.1f}"}),
                 use_container_width=True, hide_index=True)
    st.markdown('<div class="mem-step"><span class="step-num">4</span> <b>Ativação de turno — lógica anual</b></div>', unsafe_allow_html=True)
    st.markdown(f"- Turno A abre se ocup_A (ano) > **{thresholds['A']}%**\n- Turno B abre se ocup_A (ano) > **{thresholds['B']}%**\n- Turno C abre se ocup_B (ano) > **{thresholds['C']}%**")
    st.markdown('<div class="mem-step"><span class="step-num">5</span> <b>Evolução mês a mês</b></div>', unsafe_allow_html=True)
    rows_mes = []
    for m, r in meses_ativos:
        rows_mes.append({"Mês": m, "Dias": r["dias"],
                         "Op. A": r["op_A"], "Op. B": r["op_B"], "Op. C": r["op_C"],
                         "Total func.": r["total"],
                         "Labor Total": r["prod_labor_tot"], "Ciclo Total": r["prod_ciclo_tot"]})
    df_mes = pd.DataFrame(rows_mes)
    def _sty_mes(row):
        styles = [""] * len(row)
        try:
            lv = float(row["Labor Total"])
            cor = "#003D10" if lv >= 0.45 else ("#3D2D00" if lv >= 0.30 else "#3D0000")
            txt = "#B9F6CA" if lv >= 0.45 else ("#FFE57F" if lv >= 0.30 else "#FF8A80")
            for i, col in enumerate(df_mes.columns):
                if col == "Labor Total": styles[i] = f"background-color:{cor};color:{txt};font-weight:600"
        except: pass
        return styles
    st.dataframe(df_mes.style.apply(_sty_mes, axis=1).format({"Labor Total": "{:.1%}", "Ciclo Total": "{:.1%}"}),
                 use_container_width=True, hide_index=True)
    st.markdown('<div class="mem-step"><span class="step-num">6</span> <b>Totais médios por turno (ano)</b></div>', unsafe_allow_html=True)
    sh_ciclo = sum(r["h_ciclo"] for _, r in meses_ativos)
    sh_labor = sum(r["h_labor"] for _, r in meses_ativos)
    sh_ativos = sum(r["h_ativos"] for _, r in meses_ativos)
    sh_todos = sum(r["h_todos"] for _, r in meses_ativos)
    prod_lt = sh_labor / sh_todos if sh_todos > 0 else 0
    prod_ct = sh_ciclo / sh_todos if sh_todos > 0 else 0
    prod_lo = sh_labor / sh_ativos if sh_ativos > 0 else 0
    prod_co = sh_ciclo / sh_ativos if sh_ativos > 0 else 0
    media_A = sum(r["tot_A"] for _, r in meses_ativos) / n_meses
    media_B = sum(r["tot_B"] for _, r in meses_ativos) / n_meses
    media_C = sum(r["tot_C"] for _, r in meses_ativos) / n_meses
    media_tot = sum(r["total"] for _, r in meses_ativos) / n_meses
    prod_data = {"Indicador": ["Ciclo Operacional", "Ciclo Total", "Labor Operacional", "⭐ Labor Total"],
                 "Resultado": [f"{prod_co:.1%}", f"{prod_ct:.1%}", f"{prod_lo:.1%}", f"{prod_lt:.1%}"]}
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Média Turno A", f"{media_A:.1f}"); c2.metric("Média Turno B", f"{media_B:.1f}")
    c3.metric("Média Turno C", f"{media_C:.1f}"); c4.metric("Média Total", f"{media_tot:.1f}")
    st.markdown('<div class="formula-box">Labor Total ★ (ano) = Σ horas_labor ÷ Σ horas_todos (todos os meses)</div>', unsafe_allow_html=True)
    st.dataframe(pd.DataFrame(prod_data), use_container_width=True, hide_index=True)
    if st.session_state.get("mem_base_ano") is None:
        _buf_ano=BytesIO(); df_intermediario.to_excel(_buf_ano,index=False); _buf_ano.seek(0)
        st.session_state["mem_base_ano"]=_buf_ano.read()
    st.download_button("📥 Baixar base completa pós-JOIN (ano todo)", data=st.session_state["mem_base_ano"],
                       file_name="base_ano_completo.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       key="dl_mem_ano")


def gerar_aba_anual(wb, resultados, label="ANO", cp_data=None, horas_anual=None, eh_cenario=False, ws_existente=None, res_ano_override=None, modelos_lista=None, app_mod_ano=None):
    meses_com_dados = [(m, resultados[m]) for m in MESES if resultados.get(m)]
    if not meses_com_dados: return
    brd = Border(left=Side(style='thin',color='CCCCCC'),right=Side(style='thin',color='CCCCCC'),
                 top=Side(style='thin',color='CCCCCC'),bottom=Side(style='thin',color='CCCCCC'))
    F_VERDE_a  = PatternFill("solid",fgColor="92D050"); F_AMAR_a  = PatternFill("solid",fgColor="FFDE00")
    F_AZUL_a   = PatternFill("solid",fgColor="00B0F0"); F_PRETO_a = PatternFill("solid",fgColor="000000")
    F_CINZA_a  = PatternFill("solid",fgColor="D9D9D9"); F_CINZA2_a= PatternFill("solid",fgColor="BFBFBF")
    F_BRANCO_a = PatternFill("solid",fgColor="FFFFFF"); F_VD_JD_a = PatternFill("solid",fgColor="1F4D19")
    F_AM_JD_a  = PatternFill("solid",fgColor="FFDE00"); F_VERM_a  = PatternFill("solid",fgColor="FF0000")
    F_CINZA_H_a= PatternFill("solid",fgColor="D9D9D9")
    def _e(ws,r,c,val,fill=None,bold=False,color="000000",size=9,center=True,wrap=False):
        cell=ws.cell(row=r,column=c,value=val)
        cell.font=Font(name="Arial",bold=bold,color=color,size=size)
        cell.fill=fill or F_BRANCO_a
        cell.alignment=Alignment(horizontal="center" if center else "left",vertical="center",wrap_text=wrap)
        cell.border=brd; return cell
    def _cor_pct_a(v):
        try:
            f=float(str(v).strip('%'))/100 if '%' in str(v) else float(v)
            if f>=1.06: return F_VERM_a
            if f>=1.00: return PatternFill("solid",fgColor="FFFF00")
            if f>=0.40: return F_VERDE_a
            return F_BRANCO_a
        except: return F_BRANCO_a
    n_meses=len(meses_com_dados); dias_ano=sum(r["dias"] for _,r in meses_com_dados)
    hA=meses_com_dados[0][1]["hA"]; hB=meses_com_dados[0][1]["hB"]; hC=meses_com_dados[0][1]["hC"]
    heA=meses_com_dados[0][1].get("heA",hA); heB=meses_com_dados[0][1].get("heB",hB); heC=meses_com_dados[0][1].get("heC",hC)
    minA_ano=dias_ano*hA*60; minB_ano=dias_ano*hB*60; minC_ano=dias_ano*hC*60
    from collections import defaultdict
    cen_mc=defaultdict(float); cen_ml=defaultdict(float)
    sup_somas={k:{"A":0,"B":0,"C":0} for k in ["lavadora","gravacao","preset","coringa","facilitador"]}
    tot_A_ano=0; tot_B_ano=0; tot_C_ano=0; tot_func_ano=0
    sum_hciclo=0; sum_hlabor=0; sum_hativos=0; sum_htodos=0
    for mes,r in meses_com_dados:
        for _,row in r["centros"].iterrows():
            cen_mc[row.centro]+=row.min_ciclo_total; cen_ml[row.centro]+=row.min_labor_total
        sum_hciclo+=r["h_ciclo"]; sum_hlabor+=r["h_labor"]
        sum_hativos+=r["h_ativos"]; sum_htodos+=r["h_todos"]
        tot_A_ano+=r["tot_A"]; tot_B_ano+=r["tot_B"]; tot_C_ano+=r["tot_C"]; tot_func_ano+=r["total"]
        for k in sup_somas:
            sup_somas[k]["A"]+=r["suporte"][k]["A"]; sup_somas[k]["B"]+=r["suporte"][k]["B"]; sup_somas[k]["C"]+=r["suporte"][k]["C"]
    if cp_data:
        _cmc=defaultdict(float); _cml=defaultdict(float)
        for _r in cp_data:
            _cmc[_r[0]]+=_r[11]; _cml[_r[0]]+=_r[12]
        cen_mc=_cmc; cen_ml=_cml
    if res_ano_override:
        _ro = res_ano_override
        tot_A_ano=_ro.get("tot_A",tot_A_ano); tot_B_ano=_ro.get("tot_B",tot_B_ano)
        tot_C_ano=_ro.get("tot_C",tot_C_ano); tot_func_ano=_ro.get("total",tot_func_ano)
        sum_hciclo=_ro.get("h_ciclo",sum_hciclo); sum_hlabor=_ro.get("h_labor",sum_hlabor)
        sum_hativos=_ro.get("h_ativos",sum_hativos); sum_htodos=_ro.get("h_todos",sum_htodos)
        sup_somas={k:{"A":_ro["suporte"][k]["A"]*n_meses,"B":_ro["suporte"][k]["B"]*n_meses,"C":_ro["suporte"][k]["C"]*n_meses}
                   for k in sup_somas} if _ro.get("suporte") else sup_somas
    if horas_anual and not eh_cenario:
        _hc=horas_anual["h_ciclo"]; _hl=horas_anual["h_labor"]
        _ha=horas_anual["h_ativos"]; _ht=horas_anual["h_todos"]
        prod_co=_hc/_ha if _ha>0 else 0; prod_ct=_hc/_ht if _ht>0 else 0
        prod_lo=_hl/_ha if _ha>0 else 0; prod_lt=_hl/_ht if _ht>0 else 0
    else:
        prod_lt=sum_hlabor/sum_htodos if sum_htodos>0 else 0
        prod_ct=sum_hciclo/sum_htodos if sum_htodos>0 else 0
        prod_lo=sum_hlabor/sum_hativos if sum_hativos>0 else 0
        prod_co=sum_hciclo/sum_hativos if sum_hativos>0 else 0
    centros_ord=list(cen_mc.keys())
    thr_A=meses_com_dados[0][1]["thr_A"]; thr_B=meses_com_dados[0][1]["thr_B"]; thr_C=meses_com_dados[0][1]["thr_C"]
    def ocup_ano(cen,t):
        mc=cen_mc[cen]
        if t=="A": return mc/minA_ano if minA_ano>0 else 0
        if t=="B": return mc/minB_ano if minB_ano>0 else 0
        return mc/minC_ano if minC_ano>0 else 0
    def ativo_ano_A(cen): return 1 if ocup_ano(cen,"A")>thr_A else 0
    def ativo_ano_B(cen): return 1 if ocup_ano(cen,"A")>thr_B else 0
    def ativo_ano_C(cen): return 1 if ocup_ano(cen,"B")>thr_C else 0
    def hdA(cen): return dias_ano*heA*ativo_ano_A(cen)
    def hdB(cen): return dias_ano*heB*ativo_ano_B(cen)
    def hdC(cen): return dias_ano*heC*ativo_ano_C(cen)
    if res_ano_override and "centros" in res_ano_override and not res_ano_override["centros"].empty:
        _df_ro = res_ano_override["centros"]
        op_A_ano = int(_df_ro.num_A.sum()) if "num_A" in _df_ro.columns else int(_df_ro.ativo_A.sum())
        op_B_ano = int(_df_ro.num_B.sum()) if "num_B" in _df_ro.columns else int(_df_ro.ativo_B.sum())
        op_C_ano = int(_df_ro.num_C.sum()) if "num_C" in _df_ro.columns else int(_df_ro.ativo_C.sum())
        if res_ano_override.get("suporte"):
            sup_somas = {k: {"A": res_ano_override["suporte"][k]["A"]*n_meses,
                             "B": res_ano_override["suporte"][k]["B"]*n_meses,
                             "C": res_ano_override["suporte"][k]["C"]*n_meses}
                         for k in sup_somas}
    else:
        op_A_ano=sum(ativo_ano_A(c) for c in centros_ord)
        op_B_ano=sum(ativo_ano_B(c) for c in centros_ord)
        op_C_ano=sum(ativo_ano_C(c) for c in centros_ord)
    def sup_ano(key,t): return round(sup_somas[key][t]/n_meses) if n_meses else 0
    tot_suporte_A=sum(sup_ano(k,"A") for k in sup_somas)
    tot_suporte_B=sum(sup_ano(k,"B") for k in sup_somas)
    tot_suporte_C=sum(sup_ano(k,"C") for k in sup_somas)
    tot_A_calc=op_A_ano+tot_suporte_A; tot_B_calc=op_B_ano+tot_suporte_B; tot_C_calc=op_C_ano+tot_suporte_C
    tot_func_calc=tot_A_calc+tot_B_calc+tot_C_calc
    if cp_data and not (horas_anual and not eh_cenario) and not eh_cenario:
        _hc_cp = sum(r[11] for r in cp_data) / 60
        _hl_cp = sum(r[12] for r in cp_data) / 60
        _ha_cp = sum(hdA(c)+hdB(c)+hdC(c) for c in centros_ord)
        _ht_cp = tot_A_calc*dias_ano*heA + tot_B_calc*dias_ano*heB + tot_C_calc*dias_ano*heC
        if _ha_cp > 0: prod_co=_hc_cp/_ha_cp; prod_lo=_hl_cp/_ha_cp
        if _ht_cp > 0: prod_ct=_hc_cp/_ht_cp; prod_lt=_hl_cp/_ht_cp
    if ws_existente is not None:
        ws=ws_existente; ws.title=label; ws.freeze_panes="F7"
    else:
        ws=wb.create_sheet(label); ws.freeze_panes="F7"
    ws.merge_cells("A1:O1"); _e(ws,1,1,"TOTAIS — ANO",F_VD_JD_a,True,"FFFFFF",9,True)
    for ci,txt,f in [(16,"TURNO A",F_VERDE_a),(17,"TURNO B",F_AMAR_a),(18,"TURNO C",F_AZUL_a)]:
        _e(ws,1,ci,txt,f,True,"000000",8)
    ws.row_dimensions[1].height=14
    ws.merge_cells("A2:O2"); _e(ws,2,1,"TOTAL DE MINUTOS",F_CINZA_H_a,True,"000000",8,False)
    _e(ws,2,16,round(minA_ano,1),F_VERDE_a,True,"000000",8); _e(ws,2,17,round(minB_ano,1),F_AMAR_a,True,"000000",8); _e(ws,2,18,round(minC_ano,1),F_AZUL_a,True,"000000",8)
    ws.row_dimensions[2].height=13
    ws.merge_cells("A3:O3"); _e(ws,3,1,"TOTAL DE HORAS",F_CINZA_H_a,True,"000000",8,False)
    _e(ws,3,16,round(minA_ano/60,2),F_VERDE_a,True,"000000",8); _e(ws,3,17,round(minB_ano/60,2),F_AMAR_a,True,"000000",8); _e(ws,3,18,round(minC_ano/60,2),F_AZUL_a,True,"000000",8)
    ws.row_dimensions[3].height=13
    ws.merge_cells("A4:O4"); _e(ws,4,1,"Nº DIAS TRABALHADOS",F_CINZA_H_a,True,"000000",8,False)
    _e(ws,4,16,dias_ano,F_VERDE_a,True,"FF0000",9); _e(ws,4,17,dias_ano,F_AMAR_a,True,"FF0000",9); _e(ws,4,18,dias_ano,F_AZUL_a,True,"FF0000",9)
    ws.row_dimensions[4].height=13
    # LINHA 5 — 19 colunas fixas + colunas de modelo se disponíveis
    _n_mod_a = len(modelos_lista) if modelos_lista else 0
    _end_col_5a = get_column_letter(19 + _n_mod_a) if _n_mod_a else "S"
    ws.merge_cells(f"A5:{_end_col_5a}5"); _e(ws,5,1,f"RESUMO DA CARGA — ANO ({dias_ano} dias / {n_meses} meses)",F_VD_JD_a,True,"FFFFFF",10,True)
    ws.row_dimensions[5].height=18
    # CABEÇALHO — Perf. Op. entre Disponib. e Indice Ciclo
    hdrs_f=[("Máquina",F_CINZA2_a,"000000"),("PEÇA",F_CINZA2_a,"000000"),("DESCRIÇÃO",F_CINZA2_a,"000000"),("PÇ/TRAT",F_CINZA2_a,"000000"),("UM",F_CINZA2_a,"000000"),
            ("Tempo Ciclo (min)",F_PRETO_a,"FFFFFF"),("Tempo Labor (min)",F_PRETO_a,"FFFFFF"),
            ("Div. Carga",PatternFill("solid",fgColor="FF0000"),"FFFF00"),("Vol. Interna",F_CINZA2_a,"000000"),
            ("Div. Volume",PatternFill("solid",fgColor="FF0000"),"FFFF00"),("Disponib.",F_CINZA2_a,"000000"),("Perf. Op.",F_CINZA2_a,"000000"),("Indice Ciclo",F_CINZA2_a,"000000"),
            ("JA.A",F_VERDE_a,"000000"),("JA.B",F_AMAR_a,"000000"),("JA.C",F_AZUL_a,"000000"),
            ("TOTAL CICLOS (MIN)",F_CINZA_a,"000000"),("TOTAL LABOR (MIN)",F_CINZA_a,"000000"),("TOTAL PECAS",F_CINZA_a,"000000")]
    largs=[9,8,16,6,5,9,9,8,8,8,8,8,9,8,8,8,12,12,8]
    for ci,(h,f,cor) in enumerate(hdrs_f,1):
        _e(ws,6,ci,h,f,True,cor,8,True,True); ws.column_dimensions[get_column_letter(ci)].width=largs[ci-1]
    if modelos_lista:
        for _mi_a,_mod_a in enumerate(modelos_lista):
            _ci_a=20+_mi_a; _e(ws,6,_ci_a,_mod_a,F_CINZA_a,True,"000000",7,True,True); ws.column_dimensions[get_column_letter(_ci_a)].width=7
    ws.row_dimensions[6].height=42
    ri=7
    if cp_data:
        # Normaliza para 14 elementos (compatibilidade com tuplas de 13 da leitura AnoFY26)
        _cp14 = []
        for _r in cp_data:
            if len(_r) == 13:
                _c,_p,_pc,_tc,_tl,_dc,_vi,_dv,_di,_idx,_mc,_ml,_qt = _r
                _cp14.append((_c,_p,_pc,_tc,_tl,_dc,_vi,_dv,_di,1.0,_idx,_mc,_ml,_qt))
            else:
                _cp14.append(_r)
        for (cen,peca,_pc_t,tc,tl,dc,vi,dv,di,po,idx_c,mc,ml,_qtd) in _cp14:
            pA=mc/minA_ano if minA_ano>0 else 0; pB=mc/minB_ano if minB_ano>0 else 0; pC=mc/minC_ano if minC_ano>0 else 0
            _e(ws,ri,1,cen,F_BRANCO_a,False,"000000",8,False); _e(ws,ri,2,peca,F_BRANCO_a,False,"000000",8,False)
            _e(ws,ri,3,"ANO",F_BRANCO_a,False,"000000",8,False); _e(ws,ri,4,int(_pc_t),F_BRANCO_a,False,"000000",8); _e(ws,ri,5,"PC",F_BRANCO_a,False,"000000",8)
            _e(ws,ri,6,round(tc,4) if tc else "",F_PRETO_a,False,"FFFFFF",8); _e(ws,ri,7,round(tl,4) if tl else "",F_PRETO_a,False,"FFFFFF",8)
            _e(ws,ri,8,round(dc,4) if dc else "",F_VERM_a if abs(float(dc or 0)-1.0)>0.001 else F_BRANCO_a,False,"000000",8)
            _e(ws,ri,9,round(vi,4) if vi else "",F_BRANCO_a,False,"000000",8)
            _e(ws,ri,10,round(dv,4) if dv else "",F_VERM_a if abs(float(dv or 0)-1.0)>0.001 else F_BRANCO_a,False,"000000",8)
            _e(ws,ri,11,round(di,4) if di else "",F_CINZA2_a,False,"000000",8)
            # col 12 = Perf. Op. (vermelho quando ≠ 1)
            _e(ws,ri,12,round(po,4) if po else "",F_VERM_a if abs(float(po or 0)-1.0)>0.001 else F_CINZA2_a,False,"000000",8)
            # col 13 = Indice Ciclo
            _e(ws,ri,13,round(idx_c,4),F_BRANCO_a,False,"000000",8)
            # cols 14/15/16 = JA.A / JA.B / JA.C
            _e(ws,ri,14,f"{pA:.1%}",_cor_pct_a(pA),False,"000000",8)
            _e(ws,ri,15,f"{pB:.1%}",_cor_pct_a(pB),False,"000000",8)
            _e(ws,ri,16,f"{pC:.1%}",_cor_pct_a(pC),False,"000000",8)
            # cols 17/18/19 = TOTAL CICLOS / LABOR / PECAS
            _e(ws,ri,17,round(mc,4),F_BRANCO_a,False,"000000",8)
            _e(ws,ri,18,round(ml,4),F_BRANCO_a,False,"000000",8)
            _e(ws,ri,19,_qtd,F_BRANCO_a,False,"000000",8)
            if modelos_lista and app_mod_ano:
                _mods_v_a=app_mod_ano.get((cen,peca),{})
                for _mi_a,_mod_a in enumerate(modelos_lista):
                    _ci_a=20+_mi_a; _v_a=_mods_v_a.get(_mod_a,0)
                    _e(ws,ri,_ci_a,_v_a if _v_a else None,F_CINZA_a if _v_a else F_BRANCO_a,False,"000000",7)
            ws.row_dimensions[ri].height=13; ri+=1
    else:
        for cen in centros_ord:
            mc=cen_mc[cen]; ml=cen_ml[cen]
            pA=mc/minA_ano if minA_ano>0 else 0; pB=mc/minB_ano if minB_ano>0 else 0; pC=mc/minC_ano if minC_ano>0 else 0
            _e(ws,ri,1,cen,F_BRANCO_a,False,"000000",8,False); _e(ws,ri,2,"—",F_BRANCO_a,False,"000000",8)
            for ci_z in range(3,14): _e(ws,ri,ci_z,"",F_BRANCO_a,False,"000000",8)
            _e(ws,ri,14,f"{pA:.1%}",_cor_pct_a(pA),False,"000000",8)
            _e(ws,ri,15,f"{pB:.1%}",_cor_pct_a(pB),False,"000000",8)
            _e(ws,ri,16,f"{pC:.1%}",_cor_pct_a(pC),False,"000000",8)
            _e(ws,ri,17,round(mc,4),F_BRANCO_a,False,"000000",8); _e(ws,ri,18,round(ml,4),F_BRANCO_a,False,"000000",8)
            ws.row_dimensions[ri].height=13; ri+=1
    _CF=6; _RS=66
    ws.merge_cells(start_row=_RS-1,start_column=_CF,end_row=_RS-1,end_column=_CF+9)
    _e(ws,_RS-1,_CF,"RESUMO DA CARGA MÁQUINA X QUADRO DE LOTAÇÃO",F_CINZA_H_a,True,"000000",9,True); ws.row_dimensions[_RS-1].height=14
    _e(ws,_RS,_CF,"PERÍODO:",F_CINZA_H_a,True,"000000",8,False); _e(ws,_RS,_CF+1,"ANO",F_BRANCO_a,True,"FF0000",9,False)
    _e(ws,_RS,_CF+3,"DATA DE REVISÃO:",F_CINZA_H_a,True,"000000",8,False)
    _e(ws,_RS,_CF+4,datetime.now().strftime("%d/%m/%Y"),F_BRANCO_a,True,"FF0000",9)
    _e(ws,_RS,_CF+6,"HORAS POR TURNO DE TRABALHO",F_CINZA_H_a,True,"000000",8,True); ws.row_dimensions[_RS].height=14
    _e(ws,_RS+1,_CF+6,heA,F_CINZA_H_a,True,"000000",8); _e(ws,_RS+1,_CF+7,heB,F_CINZA_H_a,True,"000000",8); _e(ws,_RS+1,_CF+8,heC,F_CINZA_H_a,True,"000000",8); ws.row_dimensions[_RS+1].height=14
    for _ci,_txt,_f in [(_CF+1,"TURNO A",F_VERDE_a),(_CF+2,"TURNO B",F_AMAR_a),(_CF+3,"TURNO C",F_AZUL_a),(_CF+4,"TURNO A",F_VERDE_a),(_CF+5,"TURNO B",F_AMAR_a),(_CF+6,"TURNO C",F_AZUL_a),(_CF+7,"TURNO A",F_VERDE_a),(_CF+8,"TURNO B",F_AMAR_a),(_CF+9,"TURNO C",F_AZUL_a)]:
        _e(ws,_RS+2,_ci,_txt,_f,True,"000000",8); ws.row_dimensions[_RS+2].height=14
    _e(ws,_RS+3,_CF,"Centro",F_PRETO_a,True,"FFFFFF",8)
    for _ci,_txt in [(_CF+1,"% Ocup"),(_CF+2,"% Ocup"),(_CF+3,"% Ocup"),(_CF+4,"Nº Func"),(_CF+5,"Nº Func"),(_CF+6,"Nº Func"),(_CF+7,"Horas"),(_CF+8,"Horas"),(_CF+9,"Horas")]:
        _e(ws,_RS+3,_ci,_txt,F_PRETO_a,True,"FFFFFF",8); ws.row_dimensions[_RS+3].height=14
    _ri=_RS+4
    def _cbg_a(v):
        if v>=1.06: return F_VERM_a
        if v>=1.00: return PatternFill("solid",fgColor="FFFF00")
        if v>=0.40: return F_VERDE_a
        return F_BRANCO_a
    _ro_df = res_ano_override["centros"] if (res_ano_override and "centros" in res_ano_override and not res_ano_override["centros"].empty) else None
    _ro_map = {row.centro: row for _, row in _ro_df.iterrows()} if _ro_df is not None else {}
    for cen in centros_ord:
        oA=ocup_ano(cen,"A"); oB=ocup_ano(cen,"B"); oC=ocup_ano(cen,"C")
        if cen in _ro_map:
            _row_ro = _ro_map[cen]
            nA=int(_row_ro.num_A) if hasattr(_row_ro,"num_A") else int(_row_ro.ativo_A)
            nB=int(_row_ro.num_B) if hasattr(_row_ro,"num_B") else int(_row_ro.ativo_B)
            nC=int(_row_ro.num_C) if hasattr(_row_ro,"num_C") else int(_row_ro.ativo_C)
            aA=1 if nA>0 else 0; aB=1 if nB>0 else 0; aC=1 if nC>0 else 0
        else:
            aA=ativo_ano_A(cen); aB=ativo_ano_B(cen); aC=ativo_ano_C(cen)
            nA=aA; nB=aB; nC=aC
        _e(ws,_ri,_CF,cen,F_BRANCO_a,False,"000000",8,False)
        for _pci,_pv,_pbg in [(_CF+1,oA,_cbg_a(oA)),(_CF+2,oB,_cbg_a(oB)),(_CF+3,oC,_cbg_a(oC))]:
            _pc=_e(ws,_ri,_pci,_pv,_pbg,False,"000000",8)
            _pc.number_format="0.0000000000%"
        _e(ws,_ri,_CF+4,nA,F_VERDE_a if aA else F_AMAR_a,True,"000000",8); _e(ws,_ri,_CF+5,nB,F_VERDE_a if aB else F_AMAR_a,True,"000000",8); _e(ws,_ri,_CF+6,nC,F_AZUL_a if aC else F_CINZA_a,True,"000000",8)
        _e(ws,_ri,_CF+7,round(dias_ano*heA*nA,2) if aA else 0,F_VERDE_a if aA else F_BRANCO_a,True,"000000",8)
        _e(ws,_ri,_CF+8,round(dias_ano*heB*nB,2) if aB else 0,F_AMAR_a if aB else F_BRANCO_a,True,"000000",8)
        _e(ws,_ri,_CF+9,round(dias_ano*heC*nC,2) if aC else 0,F_AZUL_a if aC else F_BRANCO_a,True,"000000",8)
        ws.row_dimensions[_ri].height=13; _ri+=1
    for _snm,_sk in [("TOTAL DE OPERADORES",None),("LAVADORA E INSPEÇÃO","lavadora"),("GRAVAÇÃO E ESTANQUEIDADE","gravacao"),("PRESET","preset"),("CORINGA","coringa"),("FACILITADOR","facilitador"),("TOTAL POR TURNO",None),("TOTAL FUNCIONÁRIOS",None)]:
        _bold="TOTAL" in _snm; _bg=F_AM_JD_a if _bold else F_BRANCO_a; _fg="1F4D19" if _bold else "000000"
        _e(ws,_ri,_CF,_snm,_bg,_bold,_fg,8,False)
        if _sk:
            sA=sup_ano(_sk,"A"); sB=sup_ano(_sk,"B"); sC=sup_ano(_sk,"C")
            for _ci,_v in [(_CF+4,sA),(_CF+5,sB),(_CF+6,sC)]: _e(ws,_ri,_ci,_v,F_VERDE_a if _ci==_CF+4 else (F_AMAR_a if _ci==_CF+5 else F_AZUL_a),True,"000000",8)
            for _ci,_v,_he in [(_CF+7,sA,heA),(_CF+8,sB,heB),(_CF+9,sC,heC)]: _e(ws,_ri,_ci,round(_v*_he*dias_ano,2) if _v else 0,F_VERDE_a if _ci==_CF+7 else (F_AMAR_a if _ci==_CF+8 else F_AZUL_a),True,"000000",8)
        elif "TOTAL DE OPERADORES" in _snm:
            for _ci,_v in [(_CF+4,op_A_ano),(_CF+5,op_B_ano),(_CF+6,op_C_ano)]: _e(ws,_ri,_ci,_v,F_AM_JD_a,True,"1F4D19",8)
            for _ci,_v,_he in [(_CF+7,op_A_ano,heA),(_CF+8,op_B_ano,heB),(_CF+9,op_C_ano,heC)]: _e(ws,_ri,_ci,round(_v*_he*dias_ano,2),F_AM_JD_a,True,"1F4D19",8)
        elif "TOTAL POR TURNO" in _snm:
            for _ci,_v in [(_CF+4,tot_A_calc),(_CF+5,tot_B_calc),(_CF+6,tot_C_calc)]: _e(ws,_ri,_ci,_v,F_AM_JD_a,True,"1F4D19",8)
            for _ci,_v,_he in [(_CF+7,tot_A_calc,heA),(_CF+8,tot_B_calc,heB),(_CF+9,tot_C_calc,heC)]: _e(ws,_ri,_ci,round(_v*_he*dias_ano,2),F_AM_JD_a,True,"1F4D19",8)
        elif "FUNCIONÁRIOS" in _snm:
            _e(ws,_ri,_CF+4,tot_func_calc,F_AM_JD_a,True,"1F4D19",9)
            _e(ws,_ri,_CF+7,round(tot_A_calc*heA*dias_ano+tot_B_calc*heB*dias_ano+tot_C_calc*heC*dias_ano,2),F_AM_JD_a,True,"1F4D19",9)
        ws.row_dimensions[_ri].height=13; _ri+=1
    _ri+=1
    for _pnm,_pv,_dest in [("PRODUTIVIDADE POR TEMPO DE CICLO OPERACIONAL",prod_co,False),("PRODUTIVIDADE POR TEMPO DE CICLO TOTAL",prod_ct,False),("PRODUTIVIDADE POR TEMPO DE LABOR OPERACIONAL",prod_lo,False),("PRODUTIVIDADE POR TEMPO DE LABOR TOTAL ★",prod_lt,True)]:
        ws.merge_cells(start_row=_ri,start_column=_CF,end_row=_ri,end_column=_CF+8)
        _e(ws,_ri,_CF,_pnm,F_AM_JD_a if _dest else F_BRANCO_a,_dest,"1F4D19" if _dest else "000000",8,False)
        _pv_cell=_e(ws,_ri,_CF+9,_pv,F_AM_JD_a if _dest else F_BRANCO_a,_dest,"1F4D19" if _dest else "000000",8)
        _pv_cell.number_format="0.0000000000%"
        ws.row_dimensions[_ri].height=14; _ri+=1
    for ci,w in enumerate([9,8,16,6,5,9,9,8,8,8,8,8,9,8,8,8,12,12,8],1): ws.column_dimensions[get_column_letter(ci)].width=w
    for _ci,_ww in [(_CF,14),(_CF+1,8),(_CF+2,8),(_CF+3,8),(_CF+4,8),(_CF+5,8),(_CF+6,8),(_CF+7,10),(_CF+8,10),(_CF+9,10)]:
        ws.column_dimensions[get_column_letter(_ci)].width=_ww
    _nota_row = _ri + 2
    F_NOTA = PatternFill("solid", fgColor="FFF2CC")
    F_NOTA_H = PatternFill("solid", fgColor="FFD966")
    _brd_n = Border(left=Side(style="thin",color="C9A700"),right=Side(style="thin",color="C9A700"),
                    top=Side(style="thin",color="C9A700"),bottom=Side(style="thin",color="C9A700"))
    ws.merge_cells(start_row=_nota_row, start_column=1, end_row=_nota_row, end_column=19)
    _nh = ws.cell(row=_nota_row, column=1, value="⚠️  NOTA SOBRE O TOTAL LABOR E TOTAL CICLOS — ANO")
    _nh.font = Font(name="Arial", bold=True, size=9, color="7D4E00")
    _nh.fill = F_NOTA_H
    _nh.border = _brd_n
    _nh.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    ws.row_dimensions[_nota_row].height = 16
    _linhas_nota = [
        ("Por que alguns valores do TOTAL LABOR (MIN) ou TOTAL CICLOS (MIN) podem diferir do Excel de referência (AnoFY26)?", True),
        ("O App calcula os totais anuais a partir do INPUTPMP × INPUTAPLICAÇÃO × INPUTTEMPO × INPUTDISTRIBUIÇÃO.", False),
        ("O Excel de referência calcula via fórmulas próprias em cada aba mensal, que podem incorporar ajustes manuais,", False),
        ("   células editadas diretamente ou lógicas de arredondamento diferentes das fórmulas do App.", False),
        ("", False),
        ("✅  Os valores calculados pelo App são fiéis ao INPUTPMP e às regras definidas nos inputs.", True),
        ("   Se desejar que o anual bata com o Excel de referência, alinhe o INPUTPMP com as quantidades reais de cada aba mensal.", False),
    ]
    for i, (txt, bold) in enumerate(_linhas_nota):
        r = _nota_row + 1 + i
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=19)
        cell = ws.cell(row=r, column=1, value=txt)
        cell.font = Font(name="Arial", bold=bold, size=8, color="7D4E00")
        cell.fill = F_NOTA
        cell.border = _brd_n
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 14 if txt else 6


@st.cache_data(show_spinner=False)
def exportar_cenario_vs_base_cached(hash_key, _res_base, _res_cenario, meses_lista, nome_cenario, _res_ano_fy26_b=None, _res_ano_fy26_c=None, _file_bytes_ano=None, _cp_data_fallback=None):
    return exportar_cenario_vs_base(_res_base, _res_cenario, meses_lista, nome_cenario, _res_ano_fy26_b, _res_ano_fy26_c, _file_bytes_ano, _cp_data_fallback)

@st.cache_data(show_spinner=False)
def exportar_cached(res_hash, _resultados, _tempo=None, _dist=None, _aplic=None, _pmp=None, _file_bytes=None, _eh_cenario=False):
    return exportar(_resultados, _tempo, _dist, _aplic, _pmp, _file_bytes, _eh_cenario)

def exportar(resultados, _tempo=None, _dist=None, _aplic=None, _pmp=None, _file_bytes=None, _eh_cenario=False):
    out=BytesIO(); wb=openpyxl.Workbook()
    brd=Border(left=Side(style='thin',color='CCCCCC'),right=Side(style='thin',color='CCCCCC'),top=Side(style='thin',color='CCCCCC'),bottom=Side(style='thin',color='CCCCCC'))
    def ec_l(c,bg="FFFFFF",fg="000000",bold=False,fmt=None,center=True):
        c.font=Font(name="Arial",bold=bold,color=fg,size=9); c.fill=PatternFill("solid",fgColor=bg)
        c.alignment=Alignment(horizontal="center" if center else "left",vertical="center"); c.border=brd
        if fmt:
            try: c.number_format=fmt
            except: pass
    ws=wb.active; ws.title="RESUMO MO"
    JD_V=JD_VERDE_ESC.replace("#",""); JD_Y=JD_AMARELO.replace("#","")
    for i,h in enumerate(["Mês","Dias","Turno A","Turno B","Turno C","Total","Ciclo Op.","Ciclo Total","Labor Op.","Labor Total ★"],1):
        ec_l(ws.cell(1,i,h),JD_V,"FFFFFF",True)
    for ri,(m,abr) in enumerate(zip(MESES,MESES_ABREV),2):
        r=resultados.get(m); bg="EAF3FB" if ri%2==0 else "FFFFFF"
        vals=[abr,0,"-","-","-","-","-","-","-","-"] if not r else [abr,r["dias"],r["tot_A"],r["tot_B"],r["tot_C"],r["total"],r["prod_ciclo_op"],r["prod_ciclo_tot"],r["prod_labor_op"],r["prod_labor_tot"]]
        for ci,v in enumerate(vals,1):
            v_cell=f"{v:.1%}" if ci>=7 and isinstance(v,float) else v
            c_obj=ws.cell(ri,ci,v_cell)
            ec_l(c_obj,JD_Y if ci==10 and isinstance(v,float) else bg,JD_V if ci==10 and isinstance(v,float) else "000000",ci==10 and isinstance(v,float))
        ws.row_dimensions[ri].height=15
    for mes in MESES:
        r=resultados.get(mes)
        if not r: continue
        wsm=wb.create_sheet(mes[:10]); hA,hB,hC,dias=r["hA"],r["hB"],r["hC"],r["dias"]
        heA,heB,heC=r.get("heA",hA),r.get("heB",hB),r.get("heC",hC)
        for ci,txt in [(1,""),(2,"TURNO A"),(3,"TURNO B"),(4,"TURNO C"),(5,"TURNO A"),(6,"TURNO B"),(7,"TURNO C"),(8,"TURNO A"),(9,"TURNO B"),(10,"TURNO C")]:
            ec_l(wsm.cell(1,ci,txt),JD_V,"FFFFFF",True)
        wsm.cell(1,1,mes.upper()); ec_l(wsm.cell(1,1),JD_V,"FFFFFF",True)
        JD_V2=JD_VERDE_ESC.replace("#",""); JD_Y2=JD_AMARELO.replace("#","")
        wsm.merge_cells("B2:D2")
        c2=wsm.cell(2,1,"CENTRO"); c2.font=Font(name="Arial",bold=True,color="FFFFFF",size=9); c2.fill=PatternFill("solid",fgColor=JD_V2); c2.alignment=Alignment(horizontal="center",vertical="center"); c2.border=brd
        c2=wsm.cell(2,2,"% OCUPAÇÃO"); c2.font=Font(name="Arial",bold=True,color="FFFFFF",size=9); c2.fill=PatternFill("solid",fgColor=JD_V2); c2.alignment=Alignment(horizontal="center",vertical="center"); c2.border=brd
        wsm.merge_cells("E2:G2")
        c2=wsm.cell(2,5,"TURNO ATIVO (0=inativo 1=ativo)"); c2.font=Font(name="Arial",bold=True,color=JD_V2,size=9); c2.fill=PatternFill("solid",fgColor=JD_Y2); c2.alignment=Alignment(horizontal="center",vertical="center"); c2.border=brd
        wsm.merge_cells("H2:J2")
        c2=wsm.cell(2,8,"HORAS DISPONIVEIS NO MES"); c2.font=Font(name="Arial",bold=True,color="FFFFFF",size=9); c2.fill=PatternFill("solid",fgColor="1565C0"); c2.alignment=Alignment(horizontal="center",vertical="center"); c2.border=brd
        wsm.row_dimensions[2].height=16
        def cbg(v):
            if v>1.0: return "FFCDD2"
            if v>=0.85: return "FFFDE7"
            return "E8F5E9"
        ri2=3
        for _,row in r["centros"].iterrows():
            for ci,(val,bg,ctr) in enumerate([(row.centro,"FFFFFF",False),(f"{row.ocup_A:.1%}",cbg(row.ocup_A),True),(f"{row.ocup_B:.1%}",cbg(row.ocup_B),True),(f"{row.ocup_C:.1%}",cbg(row.ocup_C),True),(row.ativo_A,"B3E5FC" if row.ativo_A else "FFFDE7",True),(row.ativo_B,"B3E5FC" if row.ativo_B else "FFFDE7",True),(row.ativo_C,"B3E5FC" if row.ativo_C else "FFFDE7",True),(row.horas_disp_A if row.ativo_A else "0","B3E5FC" if row.ativo_A else "F5F5F5",True),(row.horas_disp_B if row.ativo_B else "0","B3E5FC" if row.ativo_B else "F5F5F5",True),(row.horas_disp_C if row.ativo_C else "0","B3E5FC" if row.ativo_C else "F5F5F5",True)],1):
                ec_l(wsm.cell(ri2,ci,val),bg,center=ctr)
            ri2+=1
        sup=r["suporte"]
        for nome,key in [("TOTAL DE OPERADORES",None),("LAVADORA E INSPEÇÃO","lavadora"),("GRAVAÇÃO E ESTANQUEIDADE","gravacao"),("PRESET","preset"),("CORINGA","coringa"),("FACILITADOR","facilitador"),("TOTAL POR TURNO",None),("TOTAL FUNCIONÁRIOS",None)]:
            bold="TOTAL" in nome; bg_r=JD_Y if bold else "FFFFFF"; fg_r=JD_V if bold else "000000"
            ec_l(wsm.cell(ri2,1,nome),bg_r,fg_r,bold,center=False)
            if key:
                s=sup[key]
                for ci,t in [(5,"A"),(6,"B"),(7,"C")]: ec_l(wsm.cell(ri2,ci,s[t]),"B3E5FC" if s[t] else "FFFDE7",bold=bold)
                for ci,t,h in [(8,"A",heA),(9,"B",heB),(10,"C",heC)]:
                    v=s[t]*h*dias; ec_l(wsm.cell(ri2,ci,v if v else 0),"B3E5FC" if v else "F5F5F5",bold=bold)
            elif "TOTAL DE OPERADORES" in nome:
                for ci,v in [(5,r["op_A"]),(6,r["op_B"]),(7,r["op_C"])]: ec_l(wsm.cell(ri2,ci,v),JD_Y,JD_V,True)
                for ci,v,h in [(8,r["op_A"],heA),(9,r["op_B"],heB),(10,r["op_C"],heC)]: ec_l(wsm.cell(ri2,ci,v*h*dias),JD_Y,JD_V,True)
            elif "TOTAL POR TURNO" in nome:
                for ci,v in [(5,r["tot_A"]),(6,r["tot_B"]),(7,r["tot_C"])]: ec_l(wsm.cell(ri2,ci,v),JD_Y,JD_V,True)
                for ci,v,h in [(8,r["tot_A"],heA),(9,r["tot_B"],heB),(10,r["tot_C"],heC)]: ec_l(wsm.cell(ri2,ci,v*h*dias),JD_Y,JD_V,True)
            elif "FUNCIONÁRIOS" in nome:
                ec_l(wsm.cell(ri2,4,r["total"]),JD_Y,JD_V,True)
                ec_l(wsm.cell(ri2,8,r['tot_A']*heA*dias+r['tot_B']*heB*dias+r['tot_C']*heC*dias),JD_Y,JD_V,True)
            ri2+=1
        ri2+=1
        for nm,v,dest in [("PROD. CICLO OPERACIONAL",r["prod_ciclo_op"],False),("PROD. CICLO TOTAL",r["prod_ciclo_tot"],False),("PROD. LABOR OPERACIONAL",r["prod_labor_op"],False),("PROD. LABOR TOTAL ★",r["prod_labor_tot"],True)]:
            wsm.merge_cells(f"H{ri2}:I{ri2}")
            ec_l(wsm.cell(ri2,8,nm),JD_Y if dest else "FFFFFF",JD_V if dest else "000000",dest,center=False)
            ec_l(wsm.cell(ri2,10,f"{v:.1%}" if isinstance(v,float) else v),JD_Y if dest else "FFFFFF",JD_V if dest else "000000",dest)
            ri2+=1
        for ci,w in enumerate([14,8,8,8,8,8,8,24,10,10],1): wsm.column_dimensions[get_column_letter(ci)].width=w
    _fb_ano = _file_bytes or st.session_state.get("_fb_anual")
    if _tempo is not None and _dist is not None and _aplic is not None and _pmp is not None:
        _cp_ano=build_cp_data_anual(resultados,_tempo,_dist,_aplic,_pmp,file_bytes=_fb_ano)
    else:
        _cp_ano=None
    _horas_ano=read_horas_anual(_fb_ano)
    gerar_aba_anual(wb,resultados,cp_data=_cp_ano,horas_anual=_horas_ano,eh_cenario=_eh_cenario)
    # ── Aba de metadados ──────────────────────────────────────────────────────
    ws_meta = wb.create_sheet("METADADOS")
    ws_meta.column_dimensions["A"].width = 32
    ws_meta.column_dimensions["B"].width = 48
    _FMH = PatternFill("solid", fgColor=JD_VERDE_ESC.replace("#",""))
    _FML = PatternFill("solid", fgColor="F0F4F0")
    def _mc(r, c, v, header=False):
        cell = ws_meta.cell(r, c, v)
        cell.font = Font(name="Arial", bold=header, color="FFFFFF" if header else "000000", size=9)
        cell.fill = _FMH if header else _FML
        cell.alignment = Alignment(horizontal="left", vertical="center")
    _mc(1,1,"METADADOS DO CÁLCULO", header=True); _mc(1,2,"", header=True)
    ws_meta.row_dimensions[1].height = 16
    _meta_rows = [
        ("Data/hora do cálculo", datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
        ("Arquivo de input (hash)", str(abs(hash(_fb_ano))) if _fb_ano else "não disponível"),
        ("Meses calculados", ", ".join(m for m in MESES if resultados.get(m))),
        ("Total combinações ativas (aplic)", str(len(_aplic)) if _aplic is not None else "—"),
        ("Total linhas INPUTTEMPO", str(len(_tempo)) if _tempo is not None else "—"),
        ("Total linhas INPUTDISTRIBUIÇÃO", str(len(_dist)) if _dist is not None else "—"),
        ("Modelos únicos (PMP)", str(_pmp.modelo.nunique()) if _pmp is not None else "—"),
        ("Gerado por", "Calculadora de Recursos — Usinagem · John Deere"),
    ]
    for i,(k,v) in enumerate(_meta_rows, 2):
        _mc(i, 1, k); _mc(i, 2, v)
        ws_meta.row_dimensions[i].height = 14
    wb.save(out); out.seek(0); return out

def exportar_cenario_vs_base(res_base, res_cenario, meses_lista, nome_cenario, res_ano_fy26_b=None, res_ano_fy26_c=None, _file_bytes_ano=None, _cp_data_fallback=None):
    """Exporta base vs cenário para todos os meses em meses_lista.
    Gera: uma aba BASE por mês, uma aba CENÁRIO por mês, e uma aba Comparação consolidada.
    """
    # meses_lista pode ser str (compatibilidade retroativa) ou list
    if isinstance(meses_lista, str):
        meses_lista = [meses_lista]

    out=BytesIO(); wb=openpyxl.Workbook()
    _wb_default_sheet=wb.active  # aba padrão "Sheet" criada pelo openpyxl
    _usar_ano_fy26 = res_ano_fy26_b is not None and res_ano_fy26_c is not None
    brd=Border(left=Side(style='thin',color='CCCCCC'),right=Side(style='thin',color='CCCCCC'),top=Side(style='thin',color='CCCCCC'),bottom=Side(style='thin',color='CCCCCC'))
    JD_V=JD_VERDE_ESC.replace("#",""); JD_Y=JD_AMARELO.replace("#","")
    def ec_c(c,bg="FFFFFF",fg="000000",bold=False,center=True):
        c.font=Font(name="Arial",bold=bold,color=fg,size=9); c.fill=PatternFill("solid",fgColor=bg)
        c.alignment=Alignment(horizontal="center" if center else "left",vertical="center"); c.border=brd
    def escrever_mes(ws,r,titulo):
        hA,hB,hC,dias=r["hA"],r["hB"],r["hC"],r["dias"]
        heA,heB,heC=r.get("heA",hA),r.get("heB",hB),r.get("heC",hC)
        for ci,txt in [(1,titulo),(2,"TURNO A"),(3,"TURNO B"),(4,"TURNO C"),(5,"TURNO A"),(6,"TURNO B"),(7,"TURNO C"),(8,"TURNO A"),(9,"TURNO B"),(10,"TURNO C")]:
            ec_c(ws.cell(1,ci,txt),JD_V,"FFFFFF",True)
        ws.row_dimensions[1].height=16
        ec_c(ws.cell(2,1,"CENTRO"),JD_V,"FFFFFF",True); ec_c(ws.cell(2,2,"% OCUPAÇÃO"),JD_V,"FFFFFF",True)
        ws.merge_cells("E2:G2"); ec_c(ws.cell(2,5,"TURNO ATIVO (0=inativo 1=ativo)"),JD_Y,JD_V,True)
        ws.merge_cells("H2:J2"); ec_c(ws.cell(2,8,"HORAS DISPONÍVEIS NO MÊS"),"1565C0","FFFFFF",True)
        ws.row_dimensions[2].height=16
        def cbg(v):
            if v>1.0: return "FFCDD2"
            if v>=0.85: return "FFFDE7"
            return "E8F5E9"
        ri=3
        for _,row in r["centros"].iterrows():
            # Usa num_A/num_B/num_C se disponível (número real de funcionários), senão ativo binário
            _nA = int(row.num_A) if hasattr(row,"num_A") and row.num_A is not None else int(row.ativo_A)
            _nB = int(row.num_B) if hasattr(row,"num_B") and row.num_B is not None else int(row.ativo_B)
            _nC = int(row.num_C) if hasattr(row,"num_C") and row.num_C is not None else int(row.ativo_C)
            for ci,(val,bg,ctr) in enumerate([(row.centro,"FFFFFF",False),(f"{row.ocup_A:.1%}",cbg(row.ocup_A),True),(f"{row.ocup_B:.1%}",cbg(row.ocup_B),True),(f"{row.ocup_C:.1%}",cbg(row.ocup_C),True),(_nA,"B3E5FC" if row.ativo_A else "FFFDE7",True),(_nB,"B3E5FC" if row.ativo_B else "FFFDE7",True),(_nC,"B3E5FC" if row.ativo_C else "FFFDE7",True),(row.horas_disp_A if row.ativo_A else "0","B3E5FC" if row.ativo_A else "F5F5F5",True),(row.horas_disp_B if row.ativo_B else "0","B3E5FC" if row.ativo_B else "F5F5F5",True),(row.horas_disp_C if row.ativo_C else "0","B3E5FC" if row.ativo_C else "F5F5F5",True)],1):
                ec_c(ws.cell(ri,ci,val),bg,center=ctr)
            ri+=1
        sup=r["suporte"]
        for nome,key in [("TOTAL DE OPERADORES",None),("LAVADORA E INSPEÇÃO","lavadora"),("GRAVAÇÃO E ESTANQUEIDADE","gravacao"),("PRESET","preset"),("CORINGA","coringa"),("FACILITADOR","facilitador"),("TOTAL POR TURNO",None),("TOTAL FUNCIONÁRIOS",None)]:
            bold="TOTAL" in nome; bg_r=JD_Y if bold else "FFFFFF"; fg_r=JD_V if bold else "000000"
            ec_c(ws.cell(ri,1,nome),bg_r,fg_r,bold,False)
            if key:
                s=sup[key]
                for ci,t in [(5,"A"),(6,"B"),(7,"C")]: ec_c(ws.cell(ri,ci,s[t]),"B3E5FC" if s[t] else "FFFDE7",bold=bold)
                for ci,t,h in [(8,"A",heA),(9,"B",heB),(10,"C",heC)]:
                    v2=s[t]*h*dias; ec_c(ws.cell(ri,ci,v2 if v2 else 0),"B3E5FC" if v2 else "F5F5F5",bold=bold)
            elif "TOTAL DE OPERADORES" in nome:
                for ci,v2 in [(5,r["op_A"]),(6,r["op_B"]),(7,r["op_C"])]: ec_c(ws.cell(ri,ci,v2),JD_Y,JD_V,True)
                for ci,v2,h in [(8,r["op_A"],heA),(9,r["op_B"],heB),(10,r["op_C"],heC)]: ec_c(ws.cell(ri,ci,v2*h*dias),JD_Y,JD_V,True)
            elif "TOTAL POR TURNO" in nome:
                for ci,v2 in [(5,r["tot_A"]),(6,r["tot_B"]),(7,r["tot_C"])]: ec_c(ws.cell(ri,ci,v2),JD_Y,JD_V,True)
                for ci,v2,h in [(8,r["tot_A"],heA),(9,r["tot_B"],heB),(10,r["tot_C"],heC)]: ec_c(ws.cell(ri,ci,v2*h*dias),JD_Y,JD_V,True)
            elif "FUNCIONÁRIOS" in nome:
                ec_c(ws.cell(ri,4,r["total"]),JD_Y,JD_V,True)
                ec_c(ws.cell(ri,8,r['tot_A']*heA*dias+r['tot_B']*heB*dias+r['tot_C']*heC*dias),JD_Y,JD_V,True)
            ri+=1
        ri+=1
        for nm2,v2,dest in [("PROD. CICLO OPERACIONAL",r["prod_ciclo_op"],False),("PROD. CICLO TOTAL",r["prod_ciclo_tot"],False),("PROD. LABOR OPERACIONAL",r["prod_labor_op"],False),("PROD. LABOR TOTAL ★",r["prod_labor_tot"],True)]:
            ws.merge_cells(f"H{ri}:I{ri}")
            ec_c(ws.cell(ri,8,nm2),JD_Y if dest else "FFFFFF",JD_V if dest else "000000",dest,False)
            ec_c(ws.cell(ri,10,f"{v2:.1%}"),JD_Y if dest else "FFFFFF",JD_V if dest else "000000",dest)
            ri+=1
        for ci,w in enumerate([14,8,8,8,8,8,8,24,10,10],1): ws.column_dimensions[get_column_letter(ci)].width=w

    # ── gera uma aba BASE e uma aba CENÁRIO para cada mês ─────────────
    # Modo ANO FY26: não gera abas mensais, só a aba ANO Consolidado
    first_sheet_used = False
    _used_titles = set()
    def _unique_title(base_title):
        t = base_title[:31]
        if t not in _used_titles:
            _used_titles.add(t); return t
        for i in range(2, 99):
            tt = f"{base_title[:28]}_{i}"[:31]
            if tt not in _used_titles:
                _used_titles.add(tt); return tt
        return base_title[:31]

    if not _usar_ano_fy26:
        for mes in meses_lista:
            r_b = res_base.get(mes); r_c = res_cenario.get(mes)
            abrev = mes[:3].upper()
            # Aba BASE
            title_base = _unique_title(f"{abrev} Base")
            if not first_sheet_used:
                ws1 = wb.active; ws1.title = title_base; first_sheet_used = True
            else:
                ws1 = wb.create_sheet(title_base)
            if r_b: escrever_mes(ws1, r_b, f"{mes.upper()} — BASE")
            else: ws1.cell(1,1,"Sem dados para este mês no cenário base")
            # Aba CENÁRIO
            title_cen = _unique_title(f"{abrev} {nome_cenario[:10]}")
            ws2 = wb.create_sheet(title_cen)
            if r_c: escrever_mes(ws2, r_c, f"{mes.upper()} — {nome_cenario.upper()}")
            else: ws2.cell(1,1,"Sem dados para este mês no cenário")

    # ── aba Comparação consolidada (apenas no modo meses, não ANO FY26) ─
    if not _usar_ano_fy26:
        ws3 = wb.create_sheet("Comparação")
        header_title = f"COMPARAÇÃO | Base vs {nome_cenario} | {' / '.join(m[:3].upper() for m in meses_lista)}"
        ws3.merge_cells("A1:N1")
        ct = ws3.cell(1,1,header_title)
        ct.font=Font(name="Arial",bold=True,color="FFFFFF",size=11); ct.fill=PatternFill("solid",fgColor=JD_V)
        ct.alignment=Alignment(horizontal="center",vertical="center"); ct.border=brd
        ws3.row_dimensions[1].height=24
        for i,h in enumerate(["Mês","","Base","Cenário","Δ"],1): ec_c(ws3.cell(2,i,h),JD_V,"FFFFFF",True)
        ws3.row_dimensions[2].height=15
        ri3 = 3
        for mes in meses_lista:
            r_b = res_base.get(mes); r_c = res_cenario.get(mes)
            if not r_b or not r_c: continue
            metricas=[("Turno A (total)",r_b["tot_A"],r_c["tot_A"]),("Turno B (total)",r_b["tot_B"],r_c["tot_B"]),("Turno C (total)",r_b["tot_C"],r_c["tot_C"]),("TOTAL FUNCIONÁRIOS",r_b["total"],r_c["total"]),("Operadores CEN A",r_b["op_A"],r_c["op_A"]),("Operadores CEN B",r_b["op_B"],r_c["op_B"]),("Operadores CEN C",r_b["op_C"],r_c["op_C"])]
            # Cabeçalho do mês
            ws3.merge_cells(start_row=ri3,start_column=1,end_row=ri3,end_column=5)
            mc = ws3.cell(ri3,1,mes.upper())
            mc.font=Font(name="Arial",bold=True,color="FFFFFF",size=9); mc.fill=PatternFill("solid",fgColor="1F4D19")
            mc.alignment=Alignment(horizontal="left",vertical="center"); mc.border=brd
            ws3.row_dimensions[ri3].height=14; ri3+=1
            for nome3,vb3,vc3 in metricas:
                is_total="TOTAL" in nome3; delta3=vc3-vb3
                ec_c(ws3.cell(ri3,1,mes[:3].upper()),"F0F0F0","888888",False,True)
                ec_c(ws3.cell(ri3,2,nome3),JD_Y if is_total else "F5F5F5",JD_V if is_total else "000000",is_total,False)
                ec_c(ws3.cell(ri3,3,vb3),"EAF3FB","000000",is_total); ec_c(ws3.cell(ri3,4,vc3),"EAF3FB","000000",is_total)
                cor_d="003D10" if delta3<0 else ("3D0000" if delta3>0 else "555555")
                bg_d="B9F6CA" if delta3<0 else ("FFCDD2" if delta3>0 else "F5F5F5")
                ec_c(ws3.cell(ri3,5,f"{delta3:+d}"),bg_d,cor_d,is_total); ws3.row_dimensions[ri3].height=14; ri3+=1
            ri3+=1  # linha em branco entre meses
        for ci,w in enumerate([8,18,9,9,9],1): ws3.column_dimensions[get_column_letter(ci)].width=w

    # ── aba ANO FY26: gerar aba ANO base + ANO cenário (igual ao exportar normal) ──
    if _usar_ano_fy26:
        # cp_data: preferir AnoFY26; fallback = recalcular dos inputs do res_ano_fy26_b
        _cp_ano = build_cp_data_anual({}, None, None, None, None, file_bytes=_file_bytes_ano)
        if _cp_ano is None:
            # Sem AnoFY26: usar cp_data calculado externamente por (centro, peça)
            _cp_ano = _cp_data_fallback
        _ha = read_horas_anual(_file_bytes_ano)
        # Aba BASE: usa res_base (todos os meses) com cp_data
        _ws_base = wb.active; _ws_base.title = "ANO Base"
        gerar_aba_anual(wb, res_base, label="ANO Base", cp_data=_cp_ano,
                        horas_anual=_ha, eh_cenario=False, ws_existente=_ws_base)
        # Aba CENÁRIO: usa res_cenario (todos os meses com overrides) com cp_data
        _ws_cen = wb.create_sheet("ANO Cenário")
        gerar_aba_anual(wb, res_cenario, label="ANO Cenário", cp_data=_cp_ano,
                        horas_anual=None, eh_cenario=True, ws_existente=_ws_cen,
                        res_ano_override=res_ano_fy26_c)

    # ── aba ANO CONSOLIDADO (soma de meses, quando há múltiplos meses) ──────────
    meses_com_dados = [(m, res_base.get(m), res_cenario.get(m)) for m in meses_lista if res_base.get(m) and res_cenario.get(m)]
    if not _usar_ano_fy26 and len(meses_com_dados) > 1:
        ws_ano = wb.create_sheet("ANO Consolidado")
        # Cabeçalho
        if _usar_ano_fy26:
            _titulo_ano = f"ANO FY26 | Base vs {nome_cenario} (aba AnoFY26)"
        else:
            _titulo_ano = f"ANO CONSOLIDADO | Base vs {nome_cenario} | {len(meses_com_dados)} meses"
        ws_ano.merge_cells("A1:J1")
        _ca = ws_ano.cell(1,1,_titulo_ano)
        _ca.font=Font(name="Arial",bold=True,color="FFFFFF",size=11); _ca.fill=PatternFill("solid",fgColor=JD_V)
        _ca.alignment=Alignment(horizontal="center",vertical="center"); _ca.border=brd
        ws_ano.row_dimensions[1].height=24

        # Sub-cabeçalho
        ws_ano.merge_cells("A2:J2")
        if _usar_ano_fy26:
            _sub = "Fonte: aba AnoFY26 — valores anuais diretos (não soma dos meses)"
        else:
            _sub = f"Meses: {', '.join(m[:3].upper() for m,_,__ in meses_com_dados)}"
        _cm = ws_ano.cell(2,1,_sub)
        _cm.font=Font(name="Arial",bold=False,color="1F4D19",size=9); _cm.fill=PatternFill("solid",fgColor="E8F5E9")
        _cm.alignment=Alignment(horizontal="left",vertical="center"); _cm.border=brd
        ws_ano.row_dimensions[2].height=14

        # Header colunas
        for _ci,_txt,_bg,_fg in [(1,"Métrica",JD_V,"FFFFFF"),(2,"Base",JD_V,"FFFFFF"),(3,"Cenário",JD_V,"FFFFFF"),
                                   (4,"Δ absoluto",JD_V,"FFFFFF"),(5,"Δ %",JD_V,"FFFFFF")]:
            _ch = ws_ano.cell(3,_ci,_txt)
            _ch.font=Font(name="Arial",bold=True,color=_fg,size=9); _ch.fill=PatternFill("solid",fgColor=_bg)
            _ch.alignment=Alignment(horizontal="center",vertical="center"); _ch.border=brd
        ws_ano.row_dimensions[3].height=16

        if _usar_ano_fy26:
            # Modo AnoFY26: usar res_ano_fy26_b e res_ano_fy26_c diretamente (período único)
            _rb = res_ano_fy26_b; _rc = res_ano_fy26_c
            _heA = _rb.get("heA", _rb.get("hA", 8.8))
            _heB = _rb.get("heB", _rb.get("hB", 8.23))
            _heC = _rb.get("heC", _rb.get("hC", 7.68))
            _dias = _rb.get("dias", 214)
            _sup_keys = ["lavadora","gravacao","preset","coringa","facilitador"]
            metricas_ano = [
                ("─── OPERADORES POR TURNO ───", None, None, True, "section"),
                ("Turno A — Operadores",   _rb.get("op_A",0), _rc.get("op_A",0), False, "int"),
                ("Turno B — Operadores",   _rb.get("op_B",0), _rc.get("op_B",0), False, "int"),
                ("Turno C — Operadores",   _rb.get("op_C",0), _rc.get("op_C",0), False, "int"),
                ("─── TOTAL FUNCIONÁRIOS ───", None, None, True, "section"),
                ("Total Turno A (c/ suporte)", _rb.get("tot_A",0), _rc.get("tot_A",0), True, "int"),
                ("Total Turno B (c/ suporte)", _rb.get("tot_B",0), _rc.get("tot_B",0), True, "int"),
                ("Total Turno C (c/ suporte)", _rb.get("tot_C",0), _rc.get("tot_C",0), True, "int"),
                ("TOTAL FUNCIONÁRIOS (ANO)",    _rb.get("total",0), _rc.get("total",0), True, "int"),
                ("─── HORAS DISPONÍVEIS ───",  None, None, True, "section"),
                ("Horas totais Turno A", _rb.get("tot_A",0)*_heA*_dias, _rc.get("tot_A",0)*_heA*_dias, False, "float"),
                ("Horas totais Turno B", _rb.get("tot_B",0)*_heB*_dias, _rc.get("tot_B",0)*_heB*_dias, False, "float"),
                ("Horas totais Turno C", _rb.get("tot_C",0)*_heC*_dias, _rc.get("tot_C",0)*_heC*_dias, False, "float"),
                ("─── PRODUTIVIDADE ───", None, None, True, "section"),
                ("Prod. Ciclo Operacional", _rb.get("prod_ciclo_op",0), _rc.get("prod_ciclo_op",0), False, "pct"),
                ("Prod. Ciclo Total",       _rb.get("prod_ciclo_tot",0), _rc.get("prod_ciclo_tot",0), False, "pct"),
                ("Prod. Labor Operacional", _rb.get("prod_labor_op",0), _rc.get("prod_labor_op",0), False, "pct"),
                ("Prod. Labor Total ★",     _rb.get("prod_labor_tot",0), _rc.get("prod_labor_tot",0), True, "pct"),
            ]
        else:
            # Modo soma de meses (comportamento original)
            _sum = lambda key: (
                sum(rb.get(key,0) for _,rb,__ in meses_com_dados),
                sum(rc.get(key,0) for _,__,rc in meses_com_dados)
            )
            _avg = lambda key: (
                sum(rb.get(key,0) for _,rb,__ in meses_com_dados) / len(meses_com_dados),
                sum(rc.get(key,0) for _,__,rc in meses_com_dados) / len(meses_com_dados)
            )
            _heA = meses_com_dados[0][1].get("heA", meses_com_dados[0][1].get("hA",0))
            _heB = meses_com_dados[0][1].get("heB", meses_com_dados[0][1].get("hB",0))
            _heC = meses_com_dados[0][1].get("heC", meses_com_dados[0][1].get("hC",0))
            metricas_ano = [
                ("─── OPERADORES POR TURNO ───", None, None, True, "section"),
                ("Turno A — Operadores",         *_sum("op_A"),    False, "int"),
                ("Turno B — Operadores",         *_sum("op_B"),    False, "int"),
                ("Turno C — Operadores",         *_sum("op_C"),    False, "int"),
                ("─── TOTAL FUNCIONÁRIOS ───",   None, None, True, "section"),
                ("Total Turno A (c/ suporte)",   *_sum("tot_A"),   True, "int"),
                ("Total Turno B (c/ suporte)",   *_sum("tot_B"),   True, "int"),
                ("Total Turno C (c/ suporte)",   *_sum("tot_C"),   True, "int"),
                ("TOTAL FUNCIONÁRIOS (ano)",      *_sum("total"),   True, "int"),
                ("─── HORAS DISPONÍVEIS ───",    None, None, True, "section"),
                ("Horas totais Turno A",
                    sum(rb.get("tot_A",0)*_heA*rb.get("dias",0) for _,rb,__ in meses_com_dados),
                    sum(rc.get("tot_A",0)*_heA*rc.get("dias",0) for _,__,rc in meses_com_dados),
                    False, "float"),
                ("Horas totais Turno B",
                    sum(rb.get("tot_B",0)*_heB*rb.get("dias",0) for _,rb,__ in meses_com_dados),
                    sum(rc.get("tot_B",0)*_heB*rc.get("dias",0) for _,__,rc in meses_com_dados),
                    False, "float"),
                ("Horas totais Turno C",
                    sum(rb.get("tot_C",0)*_heC*rb.get("dias",0) for _,rb,__ in meses_com_dados),
                    sum(rc.get("tot_C",0)*_heC*rc.get("dias",0) for _,__,rc in meses_com_dados),
                    False, "float"),
                ("─── PRODUTIVIDADE (média) ───", None, None, True, "section"),
                ("Prod. Ciclo Operacional (méd.)", *_avg("prod_ciclo_op"), False, "pct"),
                ("Prod. Ciclo Total (méd.)",       *_avg("prod_ciclo_tot"), False, "pct"),
                ("Prod. Labor Operacional (méd.)", *_avg("prod_labor_op"), False, "pct"),
                ("Prod. Labor Total ★ (méd.)",     *_avg("prod_labor_tot"), True, "pct"),
            ]

        _ri_ano = 4
        for item in metricas_ano:
            nome_m, vb_m, vc_m, bold_m, tipo_m = item
            if tipo_m == "section":
                ws_ano.merge_cells(start_row=_ri_ano,start_column=1,end_row=_ri_ano,end_column=5)
                _cs = ws_ano.cell(_ri_ano,1,nome_m)
                _cs.font=Font(name="Arial",bold=True,color="FFFFFF",size=8)
                _cs.fill=PatternFill("solid",fgColor="1F4D19")
                _cs.alignment=Alignment(horizontal="left",vertical="center"); _cs.border=brd
                ws_ano.row_dimensions[_ri_ano].height=13; _ri_ano+=1; continue

            # Formata valores
            if tipo_m == "pct":
                vb_str = f"{vb_m:.2%}"; vc_str = f"{vc_m:.2%}"
                delta_str = f"{vc_m-vb_m:+.2%}" if vb_m is not None else "—"
                delta_pct_str = "—"
            elif tipo_m == "float":
                vb_str = f"{vb_m:.1f}"; vc_str = f"{vc_m:.1f}"
                delta_str = f"{vc_m-vb_m:+.1f}" if vb_m is not None else "—"
                delta_pct_str = f"{(vc_m-vb_m)/vb_m*100:+.1f}%" if vb_m and vb_m!=0 else "—"
            else:
                vb_str = str(int(vb_m)) if vb_m is not None else "—"
                vc_str = str(int(vc_m)) if vc_m is not None else "—"
                delta_v = int(vc_m-vb_m) if vb_m is not None else 0
                delta_str = f"{delta_v:+d}"
                delta_pct_str = f"{delta_v/vb_m*100:+.1f}%" if vb_m and vb_m!=0 else "—"

            delta_num = (vc_m - vb_m) if vb_m is not None else 0
            _bg_nome = JD_Y if bold_m else "F9F9F9"
            _fg_nome = JD_V if bold_m else "000000"
            _bg_delta = "B9F6CA" if delta_num < 0 else ("FFCDD2" if delta_num > 0 else "F5F5F5")
            _fg_delta = "003D10" if delta_num < 0 else ("3D0000" if delta_num > 0 else "555555")

            for _ci,_val,_bg,_fg,_ctr in [
                (1, nome_m, _bg_nome, _fg_nome, False),
                (2, vb_str, "EAF3FB", "000000", True),
                (3, vc_str, "EAF3FB", "000000", True),
                (4, delta_str, _bg_delta, _fg_delta, True),
                (5, delta_pct_str, _bg_delta, _fg_delta, True),
            ]:
                _cell = ws_ano.cell(_ri_ano, _ci, _val)
                _cell.font=Font(name="Arial",bold=bold_m,color=_fg,size=9)
                _cell.fill=PatternFill("solid",fgColor=_bg)
                _cell.alignment=Alignment(horizontal="center" if _ctr else "left",vertical="center")
                _cell.border=brd
            ws_ano.row_dimensions[_ri_ano].height=14; _ri_ano+=1

        for _ci,_ww in enumerate([28,10,10,12,10],1):
            ws_ano.column_dimensions[get_column_letter(_ci)].width=_ww

    # Remover aba "Sheet" padrão vazia (criada pelo openpyxl) quando ANO FY26
    # Sheet padrão já foi renomeada para "ANO Base" — nada a remover
    wb.save(out); out.seek(0); return out

def _localizar_dados_xl(ws_r):
    """
    Extrai dados de uma aba do Excel de referência buscando por labels,
    sem depender de posições fixas de linha/coluna.
    Retorna: {"op_A", "op_B", "op_C", "total", "labor", "centros": {cen: {...}}}
    """
    MAX_R, MAX_C = 160, 55
    def _nv(v):
        return re.sub(r'\s+', ' ', str(v).upper().strip()) if v is not None else ""

    # Carregar planilha em memória via iter_rows (bulk — muito mais rápido que cell-by-cell)
    grid = {}
    for r_idx, row in enumerate(
        ws_r.iter_rows(min_row=1, max_row=MAX_R, max_col=MAX_C, values_only=True), 1
    ):
        for c_idx, v in enumerate(row, 1):
            if v is not None:
                grid[(r_idx, c_idx)] = v

    result = {"op_A": 0, "op_B": 0, "op_C": 0, "total": 0, "labor": None, "centros": {}}

    # ── 1. Encontrar colunas de % ocupação e turno ativo via sub-cabeçalhos ──
    col_ocup_A = col_ativo_A = None
    for (r, c), v in sorted(grid.items()):
        n = _nv(v)
        if "OCUP" in n and col_ocup_A is None:
            # Confirma que há "TURNO A/B/C" na linha anterior ou mesma linha
            for dc in range(0, 4):
                if _nv(grid.get((r - 1, c + dc))) == "TURNO A" or _nv(grid.get((r, c + dc))) == "TURNO A":
                    col_ocup_A = c + dc
                    break
        if ("ATIVO" in n or "TURNO ATIVO" in n) and col_ativo_A is None:
            for dc in range(0, 4):
                if _nv(grid.get((r - 1, c + dc))) == "TURNO A" or _nv(grid.get((r, c + dc))) == "TURNO A":
                    col_ativo_A = c + dc
                    break
        if col_ocup_A and col_ativo_A:
            break

    # Fallback: inferir pelas linhas de dados (ocup = decimal 0-3, ativo = 0 ou 1)
    if col_ocup_A is None or col_ativo_A is None:
        for (r, c), v in sorted(grid.items()):
            if c > 5 or not (v and re.match(r'^CEN\d+', str(v).strip().upper())):
                continue
            row_nums = {cc: grid.get((r, cc)) for cc in range(c + 1, c + 15) if grid.get((r, cc)) is not None}
            pct_cols = [cc for cc, vv in row_nums.items() if _is_pct(vv)]
            bin_cols = [cc for cc, vv in row_nums.items() if _is_bin(vv) and cc not in pct_cols]
            if pct_cols and bin_cols:
                col_ocup_A = col_ocup_A or min(pct_cols)
                col_ativo_A = col_ativo_A or min(bin_cols)
                break

    # ── 2. Ler dados dos centros ──────────────────────────────────────────────
    for (r, c), v in sorted(grid.items()):
        if c > 5 or not (v and re.match(r'^CEN\d+', str(v).strip().upper())):
            continue
        cen = str(v).strip()
        if col_ocup_A and col_ativo_A:
            result["centros"][cen] = {
                "ocup_A":  safe_float(grid.get((r, col_ocup_A))),
                "ocup_B":  safe_float(grid.get((r, col_ocup_A + 1))),
                "ativo_A": safe_int(grid.get((r, col_ativo_A))),
                "ativo_B": safe_int(grid.get((r, col_ativo_A + 1))),
                "ativo_C": safe_int(grid.get((r, col_ativo_A + 2))),
            }
        else:
            # Último recurso: varrer colunas à direita separando por tipo
            pcts, bins = [], []
            for dc in range(1, 20):
                vv = grid.get((r, c + dc))
                if vv is None: continue
                if _is_pct(vv): pcts.append(safe_float(vv))
                elif _is_bin(vv) and len(pcts) >= 2: bins.append(safe_int(vv))
            result["centros"][cen] = {
                "ocup_A":  pcts[0] if pcts else 0.0,
                "ocup_B":  pcts[1] if len(pcts) > 1 else 0.0,
                "ativo_A": bins[0] if bins else 0,
                "ativo_B": bins[1] if len(bins) > 1 else 0,
                "ativo_C": bins[2] if len(bins) > 2 else 0,
            }

    # ── 3. TOTAL DE OPERADORES ────────────────────────────────────────────────
    for (r, c), v in sorted(grid.items()):
        if "TOTAL DE OPERADORES" not in _nv(v) and "TOTAL OPERADORES" not in _nv(v):
            continue
        if col_ativo_A:
            result["op_A"] = safe_int(grid.get((r, col_ativo_A)))
            result["op_B"] = safe_int(grid.get((r, col_ativo_A + 1)))
            result["op_C"] = safe_int(grid.get((r, col_ativo_A + 2)))
        else:
            nums = []
            for dc in range(1, 25):
                vv = grid.get((r, c + dc))
                if vv is not None:
                    try: nums.append(int(float(vv)))
                    except: pass
                if len(nums) >= 3: break
            if len(nums) >= 3:
                result["op_A"], result["op_B"], result["op_C"] = nums[0], nums[1], nums[2]
        break

    # ── 4. TOTAL FUNCIONÁRIOS ─────────────────────────────────────────────────
    for (r, c), v in sorted(grid.items()):
        n = _nv(v)
        if "TOTAL FUNCIONÁRIOS" in n or "TOTAL FUNCIONARIOS" in n:
            for dc in range(1, 30):
                vv = grid.get((r, c + dc))
                if vv is not None:
                    try: result["total"] = int(float(vv)); break
                    except: pass
            break

    # ── 5. PROD. LABOR TOTAL ──────────────────────────────────────────────────
    for (r, c), v in sorted(grid.items()):
        n = _nv(v)
        if "LABOR TOTAL" in n or "PROD. LABOR TOTAL" in n or "PROD LABOR TOTAL" in n:
            for dc in range(1, 15):
                vv = grid.get((r, c + dc))
                if vv is not None:
                    try:
                        f = float(vv)
                        result["labor"] = f if f <= 1.5 else f / 100.0
                        break
                    except: pass
            break

    return result

def _is_pct(v):
    """Retorna True se o valor parece ser % de ocupação (decimal 0–3, não inteiro 0/1)."""
    try:
        f = float(v)
        return 0.0 <= f <= 3.0 and f not in (0.0, 1.0)
    except: return False

def _is_bin(v):
    """Retorna True se o valor é 0 ou 1 (flag ativo/inativo)."""
    try: return int(float(v)) in (0, 1) and float(v) == int(float(v))
    except: return False

@st.cache_data(show_spinner=False)
def comparar_com_excel_cached(res_hash, _res_app, file_hash, _file_bytes, _tempo, _dist, _aplic, _pmp, _dias, _horas_turno, _thresholds, _suporte_cfg):
    return comparar_com_excel(_res_app, _file_bytes, _tempo, _dist, _aplic, _pmp, _dias, _horas_turno, _thresholds, _suporte_cfg)

@st.cache_data(show_spinner=False)
def df_to_xlsx_cached(df_hash, _df):
    b = BytesIO(); _df.to_excel(b, index=False); b.seek(0); return b.read()

def comparar_com_excel(res_app, file_bytes, tempo, dist, aplic, pmp, dias, horas_turno, thresholds, suporte_cfg):
    _CANDS={"Novembro":["NovFY26","NOV","Nov","NOVEMBRO"],"Dezembro":["DezFY26","DEZ","Dez","DEZEMBRO"],
            "Janeiro":["JanFY26","JAN","Jan","JANEIRO"],"Fevereiro":["FevFY26","FEV","Fev","FEVEREIRO"],
            "Março":["MarFY26","MAR","Mar","MARÇO"],"Abril":["AbrFY26","ABR","Abr","ABRIL"],
            "Maio":["MaiFY26","MAI","Mai","MAIO"],"Junho":["JunFY26","JUN","Jun","JUNHO"],
            "Julho":["JulFY26","JUL","Jul","JULHO"],"Agosto":["AgoFY26","AGO","Ago","AGOSTO"],
            "Setembro":["SetFY26","SET","Set","SETEMBRO"],"Outubro":["OutFY26","OUT","Out","OUTUBRO"]}
    try:
        wb=openpyxl.load_workbook(BytesIO(file_bytes),read_only=True,data_only=True); abas=wb.sheetnames
    except Exception as e:
        return None,None,f"Erro ao abrir: {e}"
    MAPA={m: find_aba(abas, cands) for m, cands in _CANDS.items()}
    MAPA={m: a for m, a in MAPA.items() if a}
    thr_A=thresholds["A"]/100; thr_B=thresholds["B"]/100; thr_C=thresholds["C"]/100
    hA=horas_turno["A"]; hB=horas_turno["B"]
    try:
        df_all=(aplic.merge(pmp,on="modelo").merge(tempo,on=["centro","peca"]).merge(dist,on=["centro","peca"]))
        if "pc_trat_x" in df_all.columns and "pc_trat_y" in df_all.columns:
            df_all["pc_trat"]=df_all["pc_trat_y"].fillna(df_all["pc_trat_x"]).fillna(1.0); df_all.drop(columns=["pc_trat_x","pc_trat_y"],inplace=True)
        elif "pc_trat_y" in df_all.columns: df_all.rename(columns={"pc_trat_y":"pc_trat"},inplace=True)
        elif "pc_trat_x" in df_all.columns: df_all.rename(columns={"pc_trat_x":"pc_trat"},inplace=True)
        df_all["pc_trat"]=pd.to_numeric(df_all.get("pc_trat",1.0),errors="coerce").fillna(1.0).clip(lower=1.0)
        if "vol_int" not in df_all.columns: df_all["vol_int"]=1.0
        df_all["vol_int"]    = pd.to_numeric(df_all["vol_int"],    errors="coerce").fillna(1.0)
        df_all["div_carga"]  = pd.to_numeric(df_all["div_carga"],  errors="coerce").fillna(0.0)
        df_all["div_volume"] = pd.to_numeric(df_all["div_volume"], errors="coerce").fillna(0.0)
        df_all["disponib"]   = pd.to_numeric(df_all["disponib"],   errors="coerce").fillna(1.0)
        if "perf_op" not in df_all.columns: df_all["perf_op"]=1.0
        df_all["perf_op"]=pd.to_numeric(df_all["perf_op"],errors="coerce").fillna(1.0)
        df_all["indice_ciclo"]=(df_all.t_ciclo*df_all.div_carga*df_all.div_volume*df_all.vol_int)/(df_all.disponib*df_all["perf_op"])
        df_all["min_ciclo"]=df_all.indice_ciclo*df_all.qtd
        agg_all=df_all.groupby(["centro","mes"]).agg(min_ciclo=("min_ciclo","sum"),qtd_total=("qtd","sum"),indice_medio=("indice_ciclo","mean")).reset_index()
    except Exception as e:
        wb.close(); return None,None,f"Erro no cálculo: {e}"
    resumo_rows=[]; detalhe_rows=[]
    for mes,aba in MAPA.items():
        r_app=res_app.get(mes)
        if not r_app: continue
        if aba not in abas:
            resumo_rows.append({"Mês":mes,"Status":"⚠️ aba ausente","Observação":f"Aba {aba} não encontrada"})
            continue
        ws_r=wb[aba]; d=dias.get(mes,0)
        if d==0: continue
        minA=d*hA*60; minB=d*hB*60
        _xl = _localizar_dados_xl(ws_r)
        xl_opA=_xl["op_A"]; xl_opB=_xl["op_B"]; xl_opC=_xl["op_C"]
        xl_tot=_xl["total"]; xl_labor=_xl["labor"]
        dA=r_app["op_A"]-xl_opA; dB=r_app["op_B"]-xl_opB; dC=r_app["op_C"]-xl_opC; dT=r_app["total"]-xl_tot
        if dT==0 and dA==0 and dB==0 and dC==0: status="✅ Igual"
        elif abs(dT)<=2: status="🟡 Pequena diferença"
        else: status="🔴 Divergência"
        resumo_rows.append({"Mês":mes,"Status":status,"CEN A App":r_app["op_A"],"CEN A Excel":xl_opA,"Δ A":f"{dA:+d}","CEN B App":r_app["op_B"],"CEN B Excel":xl_opB,"Δ B":f"{dB:+d}","CEN C App":r_app["op_C"],"CEN C Excel":xl_opC,"Δ C":f"{dC:+d}","Total App":r_app["total"],"Total Excel":xl_tot,"Δ Total":f"{dT:+d}","Labor App":f"{r_app['prod_labor_tot']:.1%}","Labor Excel":f"{xl_labor:.1%}" if xl_labor else "—"})
        if status=="✅ Igual": continue
        agg_mes=agg_all[agg_all.mes==mes].copy()
        centros_xl=_xl["centros"]
        for _,row in agg_mes.iterrows():
            try:
                cen=row.centro; mc=row.min_ciclo; qtd_app=row.qtd_total; idx_medio=row.indice_medio
                oA_app=mc/minA if minA>0 else 0; oB_app=mc/minB if minB>0 else 0
                aA_app=1 if oA_app>thr_A else 0; aB_app=1 if oA_app>thr_B else 0; aC_app=1 if oB_app>thr_C else 0
                xl=centros_xl.get(cen,{}); aA_xl=xl.get("ativo_A",0); aB_xl=xl.get("ativo_B",0); aC_xl=xl.get("ativo_C",0)
                oA_xl=xl.get("ocup_A",0.0); oB_xl=xl.get("ocup_B",0.0)
                for turno,a_app,a_xl,ocup_app,ocup_xl in [("A",aA_app,aA_xl,oA_app,oA_xl),("B",aB_app,aB_xl,oA_app,oA_xl),("C",aC_app,aC_xl,oB_app,oB_xl)]:
                    if a_app==a_xl: continue
                    delta_ocup=ocup_app-float(ocup_xl); abs_delta=abs(delta_ocup)
                    mc_xl_esp=float(ocup_xl)*(minA if turno in ("A","B") else minB)
                    vol_xl_estim=mc_xl_esp/idx_medio if idx_medio>0 else 0
                    idx_esperado=mc_xl_esp/qtd_app if qtd_app>0 else 0
                    if abs_delta>0.15:
                        if qtd_app<vol_xl_estim*0.7: causa="Volume menor que esperado"; origem=f"INPUTAPLICAÇÃO — verifique modelos do {cen}"; expl=f"App: {qtd_app:.0f} peças vs Excel: ~{vol_xl_estim:.0f}"
                        elif qtd_app>vol_xl_estim*1.3: causa="Volume maior que esperado"; origem=f"INPUTAPLICAÇÃO — verifique modelos extras do {cen}"; expl=f"App: {qtd_app:.0f} peças vs Excel: ~{vol_xl_estim:.0f}"
                        else: causa="Índice de ciclo diferente"; origem=f"INPUTDISTRIBUIÇÃO — div_carga/div_volume/disponib do {cen}"; expl=f"Índice app={idx_medio:.2f} vs esperado={idx_esperado:.2f}"
                    else:
                        thr_u=thr_A if turno=="A" else (thr_B if turno=="B" else thr_C)
                        causa=f"Ocupação próxima do threshold ({thr_u:.0%})"; origem=f"INPUTPMP — volumes do {cen}"; expl=f"Ocup app={ocup_app:.1%} vs Excel={ocup_xl:.1%}"
                    detalhe_rows.append({"Mês":mes,"Centro":cen,"Turno":turno,"App Ativo":"✅ Sim" if a_app else "❌ Não","Excel Ativo":"✅ Sim" if a_xl else "❌ Não","Ocup. App":f"{ocup_app:.1%}","Ocup. Excel":f"{float(ocup_xl):.1%}","Δ Ocupação":f"{delta_ocup:+.1%}","Causa":causa,"Onde investigar":origem,"Explicação":expl})
            except: continue
    wb.close()
    return pd.DataFrame(resumo_rows),pd.DataFrame(detalhe_rows) if detalhe_rows else pd.DataFrame(),None

@st.cache_data
def gerar_template_input():
    import openpyxl as _ox
    from openpyxl.styles import PatternFill as _PF, Font as _Ft, Alignment as _Al
    wb = _ox.Workbook()
    _FH = _PF("solid", fgColor="1F4D19"); _FD = _PF("solid", fgColor="D9EAD3")
    _FA = _PF("solid", fgColor="FFF2CC"); _FB = _PF("solid", fgColor="FFFFFF")
    _fh = lambda: _Ft(name="Arial", bold=True, color="FFFFFF", size=8)
    _fd = lambda: _Ft(name="Arial", size=8)
    _fa = lambda: _Ft(name="Arial", italic=True, color="666666", size=7)
    _ac = _Al(horizontal="center", vertical="center", wrap_text=True)
    _al = _Al(horizontal="left",   vertical="center", wrap_text=True)

    def _h(ws, r, c, v, fill=None, bold=False, italic=False, center=True, color="000000", size=8):
        cell = ws.cell(r, c, v)
        cell.font = _Ft(name="Arial", bold=bold, italic=italic, color=color, size=size)
        cell.fill = fill or _FB
        cell.alignment = _ac if center else _al
        return cell

    MESES_A = ["NOV","DEZ","JAN","FEV","MAR","ABR","MAI","JUN","JUL","AGO","SET","OUT"]
    MESES_EX = ["Novembro","Dezembro","Janeiro","Fevereiro","Março","Abril","Maio","Junho","Julho","Agosto","Setembro","Outubro"]

    # Paleta de cores das linhas de dados
    _FY  = _PF("solid", fgColor="FFFF00")  # amarelo — DIAS TRABALHADOS
    _FG1 = _PF("solid", fgColor="C6EFCE")  # verde claro — linhas ímpares
    _FG2 = _PF("solid", fgColor="FFC7CE")  # rosa claro — linhas pares
    _FW  = _PF("solid", fgColor="FFFFFF")  # branco — linhas neutras

    # ── INPUTPMP ──────────────────────────────────────────
    ws = wb.active; ws.title = "INPUTPMP"
    ws.column_dimensions["A"].width = 16
    for i in range(2, 15): ws.column_dimensions[get_column_letter(i)].width = 8
    # Linha 1: dias trabalhados — amarelo (lida pelo app)
    _h(ws,1,1,"DIAS TRABALHADOS",_FY,True,center=False,color="000000")
    for i,d in enumerate([17,20,19,20,22,20,21,22,21,22,20,21],2):
        _h(ws,1,i,d,_FY,True,color="000000")
    ws.row_dimensions[1].height = 16
    # Linha 2: labels dos meses + Total
    _h(ws,2,1,"MODELO",_FH,True,center=False,color="FFFFFF")
    for i,m in enumerate(MESES_EX,2):
        _h(ws,2,i,m,_FH,True,color="FFFFFF")
    _h(ws,2,14,"Total",_FY,True,color="000000")
    ws.row_dimensions[2].height = 14
    # Linhas de exemplo — alternando verde/rosa (col N = soma)
    for row_idx, (mod, vals) in enumerate([
        ("MODELO 1", [42,45,0,38,45,55,41,42,58,46,43,55]),
        ("MODELO 2", [7,8,0,7,7,10,8,7,10,8,8,10]),
    ], 3):
        fill = _FG1 if (row_idx % 2 == 1) else _FG2
        _h(ws,row_idx,1,mod,fill,center=False)
        for ci,v in enumerate(vals,2): _h(ws,row_idx,ci,v,fill)
        _h(ws,row_idx,14,sum(vals),_FY,True,color="000000")
        ws.row_dimensions[row_idx].height = 14

    # ── INPUTTURNOS ───────────────────────────────────────
    ws2 = wb.create_sheet("INPUTTURNOS")
    ws2.column_dimensions["A"].width = 22
    for c in ["B","C","D"]: ws2.column_dimensions[c].width = 14
    # Linha 1: valores lidos pelo app (B1=hA, C1=hB, D1=hC)
    _h(ws2,1,1,"HORAS ACUMULADAS",_FH,True,center=False,color="FFFFFF")
    _h(ws2,1,2,7.5,  _FG1, center=True)
    _h(ws2,1,3,14.25,_FG1, center=True)
    _h(ws2,1,4,19.5, _FG1, center=True)
    ws2.row_dimensions[1].height = 16
    # Linha 2: labels descritivos (ignorados pelo app)
    _h(ws2,2,1,"",_FA)
    for i,t in enumerate(["Turno A","Turno B","Turno C"],2):
        _h(ws2,2,i,t,_FA,italic=True,color="666666")
    ws2.row_dimensions[2].height = 14

    # ── INPUTTEMPO ────────────────────────────────────────
    ws3 = wb.create_sheet("INPUTTEMPO")
    for c,w in zip(["A","B","C","D","E","F","G"],[10,14,22,14,6,18,18]):
        ws3.column_dimensions[c].width = w
    for i,h in enumerate(["Máquina","PEÇA","Descrição","PEÇA/TRATOR","UM","Tempo Ciclo (min)","Tempo Labor (min)"],1):
        _h(ws3,1,i,h,_FH,True,color="FFFFFF")
    ws3.row_dimensions[1].height = 16
    for row_idx, row_vals in enumerate([
        ["CEN005","R182470","REDUÇÃO FINAL",1,"PC",25.0,12.5],
        ["CEN005","R182471","PINHÃO",2,"PC",18.0,9.0],
        ["CEN006","P123456","ENGRENAGEM",1,"PC",22.5,11.0],
    ], 2):
        fill = _FG1 if (row_idx % 2 == 0) else _FG2
        for i,v in enumerate(row_vals,1): _h(ws3,row_idx,i,v,fill)
        ws3.row_dimensions[row_idx].height = 14

    # ── INPUTDISTRIBUIÇÃO ─────────────────────────────────
    ws4 = wb.create_sheet("INPUTDISTRIBUIÇÃO")
    for c,w in zip(["A","B","C","D","E","F","G","H","I","J","K","M","N"],[10,14,22,10,6,18,18,10,11,11,14,26,14]):
        ws4.column_dimensions[c].width = w
    _dist_hdrs = ["Máquina","PEÇA","Descrição","PEÇA/TRATOR","UM","Tempo Ciclo (min)","Tempo Labor (min)",
                  "Div Carga","Vol. Interna","Div Volume","Disponibilidade","Performance Operador X Máquina","Índice Ciclo"]
    for i,h in enumerate(_dist_hdrs,1):
        _h(ws4,1,i,h,_FH,True,color="FFFFFF")
    ws4.row_dimensions[1].height = 16
    for row_idx, row_vals in enumerate([
        ["CEN005","R182470","REDUÇÃO FINAL",1,"PC",25.0,12.5,1.0,1.0,1.0,0.9,1.0,"=F{r}*H{r}*J{r}*I{r}/(K{r}*L{r})"],
        ["CEN005","R182471","PINHÃO",2,"PC",18.0,9.0,0.5,1.0,1.0,0.95,1.0,"=F{r}*H{r}*J{r}*I{r}/(K{r}*L{r})"],
        ["CEN006","P123456","ENGRENAGEM",1,"PC",22.5,11.0,1.0,1.0,1.0,0.9,0.95,"=F{r}*H{r}*J{r}*I{r}/(K{r}*L{r})"],
    ], 2):
        fill = _FG1 if (row_idx % 2 == 0) else _FG2
        for i,v in enumerate(row_vals,1):
            _v = v.replace("{r}", str(row_idx)) if isinstance(v, str) and "{r}" in v else v
            _h(ws4,row_idx,i,_v,fill)
        ws4.row_dimensions[row_idx].height = 14

    # ── INPUTAPLICAÇÃO ────────────────────────────────────
    ws5 = wb.create_sheet("INPUTAPLICAÇÃO")
    for c,w in zip(["A","B","C","D","E","F","G","H"],[10,14,22,8,6,10,10,10]):
        ws5.column_dimensions[c].width = w
    for i,h in enumerate(["Máquina","PEÇA","Descrição","PÇ/TRAT","UM","MODELO 1","MODELO 2","MODELO 3"],1):
        _h(ws5,1,i,h,_FH,True,color="FFFFFF")
    ws5.row_dimensions[1].height = 16
    for row_idx, row_vals in enumerate([
        ["CEN005","R182470","REDUÇÃO FINAL",1,"PC",1,0,0],
        ["CEN005","R182471","PINHÃO",2,"PC",1,1,0],
        ["CEN006","P123456","ENGRENAGEM",1,"PC",0,1,1],
    ], 2):
        fill = _FG1 if (row_idx % 2 == 0) else _FG2
        for i,v in enumerate(row_vals,1): _h(ws5,row_idx,i,v,fill)
        ws5.row_dimensions[row_idx].height = 14

    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()

# ════════════════════════════════════════
# INTERFACE PRINCIPAL
# ════════════════════════════════════════
st.markdown("""
<div class="jd-header">
  <h1>🏭 Calculadora de Recursos — Usinagem</h1>
  <p>Ferramenta de planejamento de headcount por turno · John Deere Manufatura</p>
</div>
""", unsafe_allow_html=True)

with st.expander("👋 Primeira vez aqui? Veja como usar em 3 passos", expanded=False):
    col1,col2,col3=st.columns(3)
    with col1:
        st.markdown('<div class="mem-step"><span class="step-num">1</span> <b>Suba seu arquivo Excel</b><br><br>📁 O app lê automaticamente as abas:<br>• <b>INPUTPMP</b> — demanda por mês<br>• <b>INPUTTEMPO</b> — tempos de ciclo e labor<br>• <b>INPUTDISTRIBUIÇÃO</b> — divisão de carga<br>• <b>INPUTAPLICAÇÃO</b> — modelos por máquina<br>• <b>INPUTTURNOS</b> — horas por turno</div>', unsafe_allow_html=True)
    with col2:
        st.markdown('<div class="mem-step"><span class="step-num">2</span> <b>Confira os resultados</b><br><br>📊 <b>Resultado por Mês</b> — headcount por turno (A/B/C)<br>🔬 <b>Como foi Calculado</b> — passo a passo do cálculo, inclusive visão <b>anual</b><br>🔄 <b>Comparar com Excel</b> — valida se o app bate com seu Excel atual<br>📥 <b>Exportar</b> — baixa o resultado formatado</div>', unsafe_allow_html=True)
    with col3:
        st.markdown('<div class="mem-step"><span class="step-num">3</span> <b>Crie cenários</b><br><br>🎯 Na aba <b>Cenários</b>: simule alterações de turno por centro<br>• Por mês: ajuste um período específico<br>• <b>ANO FY26</b>: simula o período anual consolidado (AnoFY26) com overrides aplicados em todos os meses<br>• Compare múltiplos cenários no mesmo gráfico<br>• Baixe o cenário comparado com a base</div>', unsafe_allow_html=True)

with st.expander("📄 Baixar arquivo modelo de inputs (template)", expanded=False):
    st.caption(
        "Arquivo Excel com uma linha de exemplo em cada aba. "
        "Use como base para montar seu arquivo de inputs — basta substituir os valores de exemplo pelos reais."
    )
    st.download_button(
        "📥 Baixar template",
        data=gerar_template_input(),
        file_name="template_inputs_usinagem.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="dl_template"
    )

uploaded=st.file_uploader("Upload do arquivo de inputs (.xlsm ou .xlsx)",type=["xlsm","xlsx"])
if not uploaded:
    st.info("👆 Faça upload do arquivo para começar.")
    st.stop()

file_bytes=uploaded.read()
_file_id=hash(file_bytes)
st.session_state["_fb_anual"]=file_bytes
if st.session_state.get("_file_id")!=_file_id:
    for _k in ["diag_mensal_buf","diag_inp_buf","layout_buf","tabelona_buf","tab_pura_buf",
               "cmp_cache_key","cmp_cache_resumo","cmp_cache_detalhe","cmp_cache_err",
               "base_tratada_cache","_base_cache_key","mem_base_ano"]:
        st.session_state.pop(_k,None)
    for _k in list(st.session_state.keys()):
        if _k.startswith("mem_base_") or _k.startswith("_inputs_read_"):
            st.session_state.pop(_k, None)
    st.session_state["_file_id"]=_file_id

_abas_status=verificar_abas(file_bytes)
st.session_state["_abas_map"] = {k: v for k, v in _abas_status.items() if v}
_abas_falt=[a for a in ["INPUTPMP","INPUTTEMPO","INPUTDISTRIBUIÇÃO","INPUTAPLICAÇÃO"] if not _abas_status.get(a)]
if _abas_falt:
    st.error(f"🔴 Abas obrigatórias não encontradas: {', '.join(_abas_falt)}")
    st.stop()

# Cache de leitura por hash de arquivo — relê só quando o arquivo mudar
_read_cache_key = f"_inputs_read_{_file_id}"
if _read_cache_key not in st.session_state:
    with st.spinner("Lendo planilha..."):
        try:
            _rlog = []
            st.session_state.log_leitura = _rlog
            pmp,  dias       = read_pmp(file_bytes, _rlog)
            tempo            = read_tempo(file_bytes, _rlog)
            dist             = read_dist(file_bytes, _rlog)
            aplic            = read_aplic(file_bytes, _rlog)
            turnos_arq, _turnos_ok = read_turnos(file_bytes)
            _rlog.append(f"✅ Leitura concluída em {datetime.now().strftime('%H:%M:%S')}")
            st.session_state[_read_cache_key] = (pmp, dias, tempo, dist, aplic,
                                                  turnos_arq, _turnos_ok, list(_rlog))
            st.session_state["turnos_arq"] = turnos_arq
            st.session_state[f"_hashes_{_file_id}"] = {
                "pmp":   hash(pmp.to_json()),
                "tempo": hash(tempo.to_json()),
                "dist":  hash(dist.to_json()),
                "aplic": hash(aplic.to_json()),
                "dias":  hash(str(dias)),
            }
        except ValueError as e:
            st.error(f"🔴 Erro de formato: {e}"); st.stop()
        except Exception as e:
            st.error(f"🔴 Erro inesperado: {e}"); st.stop()
else:
    pmp, dias, tempo, dist, aplic, turnos_arq, _turnos_ok, _rlog = st.session_state[_read_cache_key]
    st.session_state.log_leitura = list(_rlog)
    st.session_state["turnos_arq"] = turnos_arq

if "log_leitura" not in st.session_state: st.session_state.log_leitura = []

# ── Card de resumo de leitura ─────────────────────────────────────────────
_norm_warns = [l for l in st.session_state.get("log_leitura",[]) if "via normalização" in l or "via norm" in l]
_blank_warns = [l for l in st.session_state.get("log_leitura",[]) if "linha(s) em branco" in l]
_pc_warn = next((l for l in st.session_state.get("log_leitura",[]) if "PÇ/TRAT lido" in l and "INPUTTEMPO" in l), None)

def _leitura_row(icon, aba, detalhe, aviso=""):
    av_html = f'<span style="color:#FFD600;font-size:10px;margin-left:6px">{aviso}</span>' if aviso else ""
    return f'''<div style="display:flex;align-items:center;gap:8px;padding:5px 0;border-bottom:1px solid #2A2A2A">
  <span style="font-size:14px">{icon}</span>
  <span style="font-size:11px;font-weight:700;color:#7BC67A;min-width:130px">{aba}</span>
  <span style="font-size:11px;color:#CCCCCC">{detalhe}</span>{av_html}
</div>'''

_meses_com_demanda = pmp.mes.nunique()
_dias_ok = sum(1 for d in dias.values() if d > 0)
_pc_info = ""
if _pc_warn:
    _m = re.search(r'(\d+) linha\(s\) com valor >1', _pc_warn)
    if _m: _pc_info = f"· {_m.group(1)} peça(s) com PÇ/TRAT>1"

_turnos_info = f"A={turnos_arq.get('A',7.5):.2f}h · B={turnos_arq.get('B',14.25):.2f}h · C={turnos_arq.get('C',19.5):.2f}h"
_turnos_aviso = "" if _turnos_ok else "⚠️ usando padrão"

_rows_html = "".join([
    _leitura_row("✅","INPUTPMP",       f"{pmp.modelo.nunique()} modelos · {_meses_com_demanda} meses com demanda · {_dias_ok} meses com dias >0"),
    _leitura_row("✅","INPUTTEMPO",     f"{len(tempo)} combinações centro+peça {_pc_info}"),
    _leitura_row("✅","INPUTDISTRIBUIÇÃO", f"{len(dist)} combinações"),
    _leitura_row("✅","INPUTAPLICAÇÃO", f"{len(aplic)} combinações ativas · {aplic.modelo.nunique()} modelos"),
    _leitura_row("✅" if _turnos_ok else "⚠️","INPUTTURNOS", _turnos_info, _turnos_aviso),
])

_extra_html = ""
if _norm_warns:
    _extra_html += f'<div style="margin-top:6px;font-size:10px;color:#FFD600">ℹ️ {len(_norm_warns)} coluna(s) encontrada(s) via nome aproximado (acento/espaço diferente) — verifique aba Dados de Input se quiser confirmar</div>'
if _blank_warns:
    _extra_html += f'<div style="font-size:10px;color:#FFD600">ℹ️ Linhas em branco ignoradas em: {", ".join(w.split("[")[1].split("]")[0] for w in _blank_warns if "[" in w)}</div>'

st.markdown(f'''<div style="background:#0D1F0D;border:1px solid #2A4A2A;border-radius:10px;padding:12px 16px;margin:8px 0">
  <div style="font-size:12px;font-weight:700;color:#FFDE00;margin-bottom:8px">📋 Dados carregados</div>
  {_rows_html}
  {_extra_html}
</div>''', unsafe_allow_html=True)
_validar_key = f"_validar_{_file_id}"
if _validar_key not in st.session_state:
    st.session_state[_validar_key] = validar(pmp, tempo, dist, aplic, dias)
erros,alertas,oks = st.session_state[_validar_key]
n_prob=len(erros)+len(alertas)

# Peças excluídas do cálculo — mostrar fora do expander para não passar despercebido
_chaves_tempo = set(zip(tempo.centro, tempo.peca))
_chaves_aplic = set(zip(aplic.centro, aplic.peca))
_chaves_dist  = set(zip(dist.centro,  dist.peca))
_excl_aplic = _chaves_tempo - _chaves_aplic
_excl_dist  = _chaves_tempo - _chaves_dist
_total_excl = len(_excl_aplic | _excl_dist)
if _total_excl > 0:
    _ex_aplic_str = f"{len(_excl_aplic)} sem INPUTAPLICAÇÃO" if _excl_aplic else ""
    _ex_dist_str  = f"{len(_excl_dist)} sem INPUTDISTRIBUIÇÃO" if _excl_dist else ""
    _detalhes = " · ".join(filter(None,[_ex_aplic_str,_ex_dist_str]))
    st.warning(f"⚠️ **{_total_excl} combinação(ões) centro+peça excluída(s) do cálculo** ({_detalhes}) — o headcount dessas peças não será computado. Expanda a validação abaixo para ver quais.")

label_exp=(f"🔴 {len(erros)} erro(s)  " if erros else "")+(f"⚠️ {len(alertas)} aviso(s)" if alertas else "")+("✅ Inputs validados sem problemas" if not n_prob else "")
with st.expander(label_exp,expanded=bool(erros or _total_excl > 0)):
    for e in erros: st.markdown(f'<div class="aviso-erro">🔴 <b>ERRO:</b> {e} — <i>o cálculo continuará com os dados disponíveis, mas o resultado pode ser parcial.</i></div>',unsafe_allow_html=True)
    for a in alertas: st.markdown(f'<div class="aviso-warn">⚠️ {a}</div>',unsafe_allow_html=True)
    for o in oks: st.markdown(f'<div class="aviso-ok">✅ {o}</div>',unsafe_allow_html=True)
    if _excl_aplic:
        _ex_list = sorted(_excl_aplic)[:10]
        st.markdown(f'<div class="aviso-warn">⚠️ <b>Sem INPUTAPLICAÇÃO ({len(_excl_aplic)}):</b> {", ".join(f"{c}/{p}" for c,p in _ex_list)}{"..." if len(_excl_aplic)>10 else ""}</div>',unsafe_allow_html=True)
    if _excl_dist:
        _ex_list2 = sorted(_excl_dist)[:10]
        st.markdown(f'<div class="aviso-warn">⚠️ <b>Sem INPUTDISTRIBUIÇÃO ({len(_excl_dist)}):</b> {", ".join(f"{c}/{p}" for c,p in _ex_list2)}{"..." if len(_excl_dist)>10 else ""}</div>',unsafe_allow_html=True)
# Não bloqueamos mais — apenas avisamos. O app calcula com o que tiver.

# ── SIDEBAR
with st.sidebar:
    st.markdown("## ⚙️ Configurações")
    _def=st.session_state.get("turnos_arq",{"A":7.5,"B":14.25,"C":19.5})
    st.markdown("**Horas acumuladas por turno (INPUTTURNOS)**")
    hA=st.number_input("Turno A",value=_def["A"],step=0.01,format="%.2f")
    hB=st.number_input("Turno B",value=_def["B"],step=0.01,format="%.2f")
    hC=st.number_input("Turno C",value=_def["C"],step=0.01,format="%.2f")
    horas_turno={"A":hA,"B":hB,"C":hC}
    st.markdown("---")
    st.markdown("**Horas efetivas por turno**")
    heA=st.number_input("A (efetivas)",value=8.80,step=0.01,format="%.2f",key="input_heA")
    heB=st.number_input("B (efetivas)",value=8.23,step=0.01,format="%.2f",key="input_heB")
    heC=st.number_input("C (efetivas)",value=7.68,step=0.01,format="%.2f",key="input_heC")
    horas_efetivas={"A":heA,"B":heB,"C":heC}
    st.markdown("---")
    st.markdown("**Thresholds de ativação (%)**")
    thr_A=st.number_input("A abre quando ocup.A >",value=40,min_value=0,max_value=200,step=1)
    thr_B=st.number_input("B abre quando ocup.A >",value=106,min_value=0,max_value=200,step=1)
    thr_C=st.number_input("C abre quando ocup.B >",value=100,min_value=0,max_value=200,step=1)
    thresholds={"A":thr_A,"B":thr_B,"C":thr_C}
    st.markdown("---")
    st.markdown("**Funções de suporte**")
    st.info("ℹ️ Turnos sem operadores ativos têm suporte zerado automaticamente.", icon=None)
    # Mostra operadores por turno do último cálculo como referência
    try:
        _rb = st.session_state.get("last_res_base")
        if _rb:
            _mes_ref = list(_rb.keys())[0]
            _op_ref = _rb[_mes_ref]
            st.caption(f"📊 Ref. ({_mes_ref}): A={_op_ref['op_A']} op · B={_op_ref['op_B']} op · C={_op_ref['op_C']} op")
    except: pass
    suporte_cfg={}
    for key,label,defs in [("lavadora","Lavadora e Inspeção",{"A":1,"B":1,"C":0}),("gravacao","Gravação e Estanqueidade",{"A":1,"B":1,"C":0}),("preset","Preset",{"A":2,"B":1,"C":1}),("coringa","Coringa",{"A":1,"B":0,"C":0}),("facilitador","Facilitador",{"A":1,"B":1,"C":0})]:
        _exp_open = st.session_state.get(f"m_{key}", "Automático") == "Manual"
        with st.expander(f"🔧 {label}", expanded=_exp_open):
            modo=st.radio("Modo",["Automático","Manual"],key=f"m_{key}",horizontal=True,label_visibility="collapsed")
            if modo=="Automático":
                st.caption(f"Padrão: A={defs['A']} · B={defs['B']} · C={defs['C']} · Zerado automaticamente em turnos sem operadores")
                suporte_cfg[key]={"modo":"auto",**defs}
            else:
                st.caption("Os valores abaixo são zerados automaticamente se o turno não tiver operadores.")
                c1,c2,c3=st.columns(3)
                vA=c1.number_input("A",0,10,defs["A"],key=f"s_{key}_A"); vB=c2.number_input("B",0,10,defs["B"],key=f"s_{key}_B"); vC=c3.number_input("C",0,10,defs["C"],key=f"s_{key}_C")
                suporte_cfg[key]={"modo":"manual","A":vA,"B":vB,"C":vC}

_ABA_NAMES = ["🏠 Visão Geral","📂 Dados de Input","🔬 Como foi Calculado","📊 Resultado por Mês","🎯 Cenários","🔄 Comparar com Excel","📥 Exportar"]
_aba = st.radio("", _ABA_NAMES, horizontal=True, key="aba_ativa", label_visibility="collapsed")

@st.cache_data(show_spinner=False)
def calcular_cached(pmp_hash,_pmp,_tempo,_dist,_aplic,aplic_hash,dias_hash,dias,hA,hB,hC,heA,heB,heC,tA,tB,tC,sup_hash,_sup):
    return calcular(_pmp,_tempo,_dist,_aplic,dias,{"A":hA,"B":hB,"C":hC},{"A":tA,"B":tB,"C":tC},_sup,horas_efetivas={"A":heA,"B":heB,"C":heC},retornar_intermediarios=True)

_h = st.session_state.get(f"_hashes_{_file_id}", {})
pmp_hash  = _h.get("pmp",  hash(pmp.to_json()))
dias_hash = _h.get("dias", hash(str(dias)))
aplic_hash= _h.get("aplic",hash(aplic.to_json()))
_sup_hash = hash(str(suporte_cfg))
res_base,df_interm,agg_interm=calcular_cached(pmp_hash,pmp,tempo,dist,aplic,aplic_hash,dias_hash,dias,horas_turno["A"],horas_turno["B"],horas_turno["C"],horas_efetivas["A"],horas_efetivas["B"],horas_efetivas["C"],thresholds["A"],thresholds["B"],thresholds["C"],_sup_hash,suporte_cfg)
st.session_state["last_res_base"]=res_base

# ── TAB 1 VISÃO GERAL
if _aba == "🏠 Visão Geral":
    st.plotly_chart(grafico_cenarios({"Base":res_base}),use_container_width=True)
    meses_ok=[m for m in MESES if res_base.get(m)]
    if meses_ok:
        media_labor=np.mean([res_base[m]["prod_labor_tot"] for m in meses_ok])
        max_total=max(res_base[m]["total"] for m in meses_ok)
        min_total=min(res_base[m]["total"] for m in meses_ok)
        mes_pico=max(meses_ok,key=lambda m:res_base[m]["total"]); mes_vale=min(meses_ok,key=lambda m:res_base[m]["total"])
        c1,c2,c3,c4=st.columns(4)
        c1.metric("Meses calculados",len(meses_ok)); c2.metric("⭐ Labor Total médio",f"{media_labor:.0%}")
        c3.metric("Pico de headcount",f"{max_total} func.",delta=f"em {mes_pico[:3].upper()}")
        c4.metric("Variação anual",f"{max_total-min_total} func.")

# ── TAB 2 INPUTS
if _aba == "📂 Dados de Input":
    st.markdown('<div class="jd-section">Dados carregados</div>',unsafe_allow_html=True)
    aba_inp=st.radio("Qual dado conferir?",["INPUTPMP","INPUTTEMPO","INPUTDISTRIBUIÇÃO","INPUTAPLICAÇÃO"],horizontal=True)
    if aba_inp=="INPUTPMP":
        st.dataframe(pmp.head(100),use_container_width=True,hide_index=True)
    elif aba_inp=="INPUTTEMPO":
        st.dataframe(tempo.head(100),use_container_width=True,hide_index=True)
    elif aba_inp=="INPUTDISTRIBUIÇÃO":
        st.dataframe(dist.head(100),use_container_width=True,hide_index=True)
    elif aba_inp=="INPUTAPLICAÇÃO":
        st.dataframe(aplic.head(200),use_container_width=True,hide_index=True)
    log_html="".join([f'<div class="log-line {"log-ok" if "✅" in l else "log-warn" if "⚠️" in l else ""}">{l}</div>' for l in st.session_state.get("log_leitura",[])])
    st.markdown(f'<div style="background:#1A1A1A;padding:12px;border-radius:8px;max-height:180px;overflow-y:auto">{log_html}</div>',unsafe_allow_html=True)
    def to_xlsx(df): b=BytesIO(); df.to_excel(b,index=False); b.seek(0); return b
    c1,c2,c3,c4=st.columns(4)
    _h = st.session_state.get(f"_hashes_{_file_id}", {})
    _ph  = _h.get("pmp",   hash(pmp.to_json()))
    _th  = _h.get("tempo", hash(tempo.to_json()))
    _dh  = _h.get("dist",  hash(dist.to_json()))
    _ah  = _h.get("aplic", hash(aplic.to_json()))
    c1.download_button("📥 INPUTPMP",data=df_to_xlsx_cached(_ph,pmp),file_name="pmp.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",key="dl_inp_pmp")
    c2.download_button("📥 INPUTTEMPO",data=df_to_xlsx_cached(_th,tempo),file_name="tempo.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",key="dl_inp_tempo")
    c3.download_button("📥 INPUTDIST.",data=df_to_xlsx_cached(_dh,dist),file_name="dist.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",key="dl_inp_dist")
    c4.download_button("📥 INPUTAPLIC.",data=df_to_xlsx_cached(_ah,aplic),file_name="aplic.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",key="dl_inp_aplic")

# ── GLOSSÁRIO (reutilizado em tab_mem e tab_res)
def _render_glossario():
    with st.expander("📖 O que significa cada métrica?", expanded=False):
        st.markdown("""
| Métrica | O que é | Fórmula resumida |
|---|---|---|
| **Índice de Ciclo** | Minutos que a máquina precisa rodar por peça produzida, considerando distribuição e disponibilidade | `(T.Ciclo × Div.Carga × Div.Volume × Vol.Interna) ÷ (Disponib. × Perf.Operador)` |
| **Min.Ciclo** | Total de minutos-máquina necessários no mês | `Índice de Ciclo × Qtd produzida` |
| **Min.Labor** | Total de minutos-operador necessários no mês | `T.Labor × Div.Carga × PÇ/TRAT × Qtd produzida` |
| **H.Ciclo / H.Labor** | Horas totais convertidas de minutos | `Min ÷ 60` |
| **Ocup. A/B/C** | Quanto % do turno está ocupado (>100% = sobrecarga) | `H.necessárias ÷ H.disponíveis no turno` |
| **Prod. Ciclo Operacional** | % das horas de ciclo vs horas dos turnos com operadores ativos | `H.Ciclo ÷ H.ativos` |
| **Prod. Labor Operacional** | % das horas de labor vs horas dos turnos com operadores ativos | `H.Labor ÷ H.ativos` |
| **Prod. Labor Total ★** | Métrica principal — labor sobre todos os turnos disponíveis | `H.Labor ÷ H.todos` |
| **Div.Carga** | Fração da demanda que passa por este centro (ex: 0.5 = 50% das peças) | Input INPUTDISTRIBUIÇÃO |
| **Vol.Interna** | Percentual produzido internamente (1.0 = 100% interno) | Input INPUTDISTRIBUIÇÃO |
| **PÇ/TRAT** | Quantas peças por veículo/tratamento (ex: 2 = uma ordem produz 2 peças) | Input INPUTTEMPO |
""")

# ── TAB 3 MEMÓRIA
if _aba == "🔬 Como foi Calculado":
    st.markdown('<div class="jd-section">Como foi calculado</div>', unsafe_allow_html=True)
    _render_glossario()
    st.markdown('<div class="aviso-ok">💡 Selecione <b>📅 ANO FY26</b> para ver a memória consolidada do período anual, ou escolha um mês específico para detalhar aquele período.</div>', unsafe_allow_html=True)
    _opcoes_mem = ["📅 ANO (visão consolidada)"] + [m for m in MESES if res_base.get(m)]
    mes_mem = st.selectbox("Período de análise", _opcoes_mem, key="mes_mem")
    if mes_mem == "📅 ANO (visão consolidada)":
        show_memoria_ano(res_base, df_interm, agg_interm, horas_turno, thresholds)
    elif mes_mem and res_base.get(mes_mem):
        show_memoria(res_base[mes_mem], mes_mem, df_interm, agg_interm, horas_turno, thresholds)

# ── TAB 4 RESULTADOS
if _aba == "📊 Resultado por Mês":
    st.markdown('<div class="jd-section">Resultado por mês</div>',unsafe_allow_html=True)
    _render_glossario()
    st.markdown('<div class="aviso-ok">💡 Selecione <b>📅 ANO</b> para ver o resumo consolidado anual, ou escolha um mês específico.</div>', unsafe_allow_html=True)
    _opcoes_res=["📅 ANO (resumo anual)"] + [m for m in MESES if res_base.get(m)]
    mes_r=st.selectbox("Período",_opcoes_res,key="mes_r")

    if mes_r=="📅 ANO (resumo anual)":
        _meses_ano=[(m,res_base[m]) for m in MESES if res_base.get(m)]
        if _meses_ano:
            _n=len(_meses_ano); _d=sum(r["dias"] for _,r in _meses_ano)
            _shc=sum(r["h_ciclo"] for _,r in _meses_ano); _shl=sum(r["h_labor"] for _,r in _meses_ano)
            _sha=sum(r["h_ativos"] for _,r in _meses_ano); _sht=sum(r["h_todos"] for _,r in _meses_ano)
            _pl=_shl/_sht if _sht>0 else 0; _pc=_shc/_sht if _sht>0 else 0
            _plo=_shl/_sha if _sha>0 else 0
            _tA=round(sum(r["tot_A"] for _,r in _meses_ano)/_n,1)
            _tB=round(sum(r["tot_B"] for _,r in _meses_ano)/_n,1)
            _tC=round(sum(r["tot_C"] for _,r in _meses_ano)/_n,1)
            _tf=round(sum(r["total"] for _,r in _meses_ano)/_n,1)
            _pico=max(_meses_ano,key=lambda x:x[1]["total"]); _vale=min(_meses_ano,key=lambda x:x[1]["total"])
            _max_tot=max(r["total"] for _,r in _meses_ano)

            # HEADER
            st.markdown(f'''
<div style="background:linear-gradient(135deg,#1F4D19,#0D2A0D);border-radius:12px;padding:16px 22px;margin-bottom:18px;border-left:5px solid #FFDE00;">
  <div style="font-size:12px;color:#7BC67A;font-weight:700;text-transform:uppercase;letter-spacing:.08em;margin-bottom:4px;">📅 Visão Anual</div>
  <div style="font-size:20px;font-weight:800;color:#FFFFFF;">{_n} meses calculados · {_d} dias trabalhados</div>
  <div style="font-size:12px;color:#AAAAAA;margin-top:4px;">
    Pico: <b style="color:#FFDE00">{_pico[0][:3].upper()} ({_pico[1]["total"]} func.)</b> &nbsp;·&nbsp;
    Vale: <b style="color:#7BC67A">{_vale[0][:3].upper()} ({_vale[1]["total"]} func.)</b>
  </div>
</div>''', unsafe_allow_html=True)

            # KPI CARDS
            c1,c2,c3,c4=st.columns(4)
            with c1:
                st.markdown(f'''<div class="kpi-card destaque">
  <div class="kpi-icon">⭐</div>
  <div class="kpi-label">Labor Total Anual</div>
  <div class="kpi-value">{_pl:.1%}</div>
  <div class="kpi-sub">Produtividade real de toda a equipe</div>
</div>''', unsafe_allow_html=True)
            with c2:
                st.markdown(f'''<div class="kpi-card">
  <div class="kpi-icon">👷</div>
  <div class="kpi-label">Média Funcionários / Mês</div>
  <div class="kpi-value">{_tf:.0f}</div>
  <div class="kpi-sub"><span class="tp tpA">A {_tA:.0f}</span><span class="tp tpB">B {_tB:.0f}</span><span class="tp tpC">C {_tC:.0f}</span></div>
</div>''', unsafe_allow_html=True)
            with c3:
                st.markdown(f'''<div class="kpi-card">
  <div class="kpi-icon">🔄</div>
  <div class="kpi-label">Ciclo Total Anual</div>
  <div class="kpi-value">{_pc:.1%}</div>
  <div class="kpi-sub">Labor Operacional: {_plo:.1%}</div>
</div>''', unsafe_allow_html=True)
            with c4:
                _var=_max_tot-min(r["total"] for _,r in _meses_ano)
                st.markdown(f'''<div class="kpi-card">
  <div class="kpi-icon">📈</div>
  <div class="kpi-label">Variação Anual</div>
  <div class="kpi-value">{_var}</div>
  <div class="kpi-sub">func. entre pico e vale</div>
</div>''', unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)

            # GAUGE + BARRAS
            col_g,col_b=st.columns([1,2])
            with col_g:
                st.markdown('<div class="jd-sub">Labor Total Anual</div>',unsafe_allow_html=True)
                pct_int=int(_pl*100)
                cor_g="#69F0AE" if _pl>=0.45 else ("#FFDE00" if _pl>=0.30 else "#FF5252")
                ang=min(180,int(_pl*180)); rad=math.radians(ang)
                ex=80+60*math.cos(math.pi-rad); ey=80-60*math.sin(rad)
                lg="1" if ang>90 else "0"
                st.markdown(f'''<div class="gauge-wrap">
  <svg viewBox="0 0 160 100" width="180">
    <path d="M 20 80 A 60 60 0 0 1 140 80" fill="none" stroke="#2A2A2A" stroke-width="12" stroke-linecap="round"/>
    <path d="M 20 80 A 60 60 0 {lg} 1 {ex:.1f} {ey:.1f}" fill="none" stroke="{cor_g}" stroke-width="12" stroke-linecap="round"/>
    <text x="80" y="76" text-anchor="middle" font-size="20" font-weight="900" fill="{cor_g}">{pct_int}%</text>
    <text x="80" y="92" text-anchor="middle" font-size="9" fill="#888">Labor Total</text>
  </svg>
  <div style="font-size:11px;color:#7BC67A;">Meta: acima de 40%</div>
</div>''', unsafe_allow_html=True)
                pct2=int(_pc*100); ang2=min(180,int(_pc*180)); rad2=math.radians(ang2)
                ex2=80+60*math.cos(math.pi-rad2); ey2=80-60*math.sin(rad2); lg2="1" if ang2>90 else "0"
                st.markdown(f'''<div class="gauge-wrap" style="margin-top:8px;">
  <svg viewBox="0 0 160 100" width="140">
    <path d="M 20 80 A 60 60 0 0 1 140 80" fill="none" stroke="#2A2A2A" stroke-width="10" stroke-linecap="round"/>
    <path d="M 20 80 A 60 60 0 {lg2} 1 {ex2:.1f} {ey2:.1f}" fill="none" stroke="#7BC67A" stroke-width="10" stroke-linecap="round"/>
    <text x="80" y="76" text-anchor="middle" font-size="18" font-weight="800" fill="#7BC67A">{pct2}%</text>
    <text x="80" y="92" text-anchor="middle" font-size="9" fill="#888">Ciclo Total</text>
  </svg>
</div>''', unsafe_allow_html=True)

            with col_b:
                st.markdown('<div class="jd-sub">Funcionários por Mês</div>',unsafe_allow_html=True)
                _mb=max(r["total"] for _,r in _meses_ano) or 1
                bars="".join([f'''<div class="mes-row">
  <div class="mes-nome">{m[:3].upper()}</div>
  <div style="flex:1;display:flex;align-items:center;gap:6px;">
    <div class="mes-bar"><div class="mes-bar-fill" style="width:{r["total"]/_mb*100:.0f}%"></div></div>
    <div style="display:flex;gap:3px;min-width:90px;"><span class="tp tpA">{r["tot_A"]}</span><span class="tp tpB">{r["tot_B"]}</span><span class="tp tpC">{r["tot_C"]}</span></div>
  </div>
  <div class="mes-num">{r["total"]}</div>
  <div class="mes-labor" style="color:{"#69F0AE" if r["prod_labor_tot"]>=0.45 else ("#FFDE00" if r["prod_labor_tot"]>=0.30 else "#FF5252")}">{r["prod_labor_tot"]:.0%}</div>
</div>''' for m,r in _meses_ano])
                st.markdown(bars,unsafe_allow_html=True)

            st.markdown("<br>",unsafe_allow_html=True)
            st.markdown('<div class="jd-sub">Tabela detalhada</div>',unsafe_allow_html=True)
            _rows=[]
            for _m,_r in _meses_ano:
                _rows.append({"Mês":_m,"Dias":_r["dias"],"Turno A":_r["tot_A"],"Turno B":_r["tot_B"],"Turno C":_r["tot_C"],"Total":_r["total"],"Labor Total":f'{_r["prod_labor_tot"]:.1%}',"Labor Op.":f'{_r["prod_labor_op"]:.1%}',"Ciclo Total":f'{_r["prod_ciclo_tot"]:.1%}'})
            _rows.append({"Mês":"📅 MÉDIA ANO","Dias":_d,"Turno A":_tA,"Turno B":_tB,"Turno C":_tC,"Total":_tf,"Labor Total":f'{_pl:.1%}',"Labor Op.":f'{_plo:.1%}',"Ciclo Total":f'{_pc:.1%}'})
            def _sty_ano(row):
                if "ANO" in str(row["Mês"]): return [f"background-color:{JD_AMARELO};color:{JD_VERDE_ESC};font-weight:700"]*len(row)
                styles=[""]*len(row)
                try:
                    lv=float(str(row["Labor Total"]).strip("%"))/100
                    cor="#003D10" if lv>=0.45 else ("#3D2D00" if lv>=0.30 else "#3D0000")
                    txt="#B9F6CA" if lv>=0.45 else ("#FFE57F" if lv>=0.30 else "#FF8A80")
                    for i,col in enumerate(pd.DataFrame(_rows).columns):
                        if col=="Labor Total": styles[i]=f"background-color:{cor};color:{txt};font-weight:600"
                except: pass
                return styles
            st.dataframe(pd.DataFrame(_rows).style.apply(_sty_ano,axis=1),use_container_width=True,hide_index=True)

    elif mes_r and res_base.get(mes_r):
        r=res_base[mes_r]
        st.markdown(f'<div class="aviso-ok">📋 <b>{mes_r}</b> — {r["dias"]} dias &nbsp;|&nbsp; A: <b>{r["tot_A"]}</b> &nbsp;|&nbsp; B: <b>{r["tot_B"]}</b> &nbsp;|&nbsp; C: <b>{r["tot_C"]}</b> &nbsp;|&nbsp; <b>Total: {r["total"]} func.</b></div>',unsafe_allow_html=True)
        show_tabela(r)

# ── TAB 5 CENÁRIOS
if _aba == "🎯 Cenários":
    if "cenarios" not in st.session_state: st.session_state.cenarios={}
    st.markdown('<div class="jd-section">Simulador de cenários</div>',unsafe_allow_html=True)
    st.markdown('<div class="aviso-ok">🎯 <b>Como usar:</b> dê um nome, escolha mês ou ANO, ajuste os turnos por centro à vontade (sem travar a tela) e clique em <b>Salvar</b>. Até 4 cenários podem ser comparados no gráfico ao mesmo tempo.</div>', unsafe_allow_html=True)

    # ── Funções de lookup — definidas fora do expander para cache funcionar corretamente
    @st.cache_data(show_spinner=False)
    def _get_ocup_mes(_pmp_hash, mes, _centros_json):
        """Lê ocupação e ativação por centro para um mês. Cacheia por hash dos dados."""
        import json
        centros_data = json.loads(_centros_json)
        return centros_data

    def _build_ocup_ref(mes):
        if not res_base.get(mes): return {}
        df_c = res_base[mes]["centros"]
        ref = {}
        for cen in sorted(df_c.centro.tolist()):
            row_ = df_c[df_c.centro == cen]
            if not row_.empty:
                r_ = row_.iloc[0]
                ref[cen] = {"oA": r_.ocup_A, "oB": r_.ocup_B, "oC": r_.ocup_C,
                             "aA": int(r_.ativo_A), "aB": int(r_.ativo_B), "aC": int(r_.ativo_C)}
        return ref

    def _build_ocup_ref_ano(meses_lista):
        """Lê ocupação e ativos base diretamente do AnoFY26 (1° bloco RESUMO DA LOTAÇÃO)."""
        _fb = st.session_state.get("_fb_anual")
        if _fb:
            try:
                _wb = openpyxl.load_workbook(BytesIO(_fb), read_only=True, data_only=True)
                if "AnoFY26" in _wb.sheetnames:
                    _ws = _wb["AnoFY26"]
                    _rows = list(_ws.rows)
                    # Total min disponíveis
                    _l2 = [c.value for c in _rows[1]]
                    _minA = float(_l2[12]) if _l2[12] else 96300.0
                    _minB = float(_l2[13]) if _l2[13] else 182970.0
                    _minC = float(_l2[14]) if _l2[14] else 250380.0
                    # Somar min_ciclo por centro
                    from collections import defaultdict as _dd
                    _cen_mc = _dd(float)
                    for _row in _rows[6:]:
                        _vals = [c.value for c in _row[:18]]
                        if _vals[0] and str(_vals[0]).startswith("CEN"):
                            _cen_mc[str(_vals[0])] += float(_vals[15] or 0)
                    # Ler ativo base do 1° bloco do resumo
                    _cen_base = {}
                    for _row in _rows:
                        _vals = [c.value for c in _row[:22]]
                        _cen = _vals[11] if len(_vals) > 11 else None
                        if _cen and str(_cen).startswith("CEN"):
                            try:
                                _cen_base[str(_cen)] = {
                                    "aA": int(_vals[15] or 0),
                                    "aB": int(_vals[16] or 0),
                                    "aC": int(_vals[17] or 0),
                                }
                            except: pass
                    _wb.close()
                    ref = {}
                    for cen in sorted(_cen_mc.keys()):
                        _mc = _cen_mc[cen]
                        _oA = _mc / _minA if _minA > 0 else 0
                        _oB = _mc / _minB if _minB > 0 else 0
                        _oC = _mc / _minC if _minC > 0 else 0
                        _b = _cen_base.get(cen, {})
                        ref[cen] = {"oA": _oA, "oB": _oB, "oC": _oC,
                                    "aA": _b.get("aA", 1 if _oA > 0.40 else 0),
                                    "aB": _b.get("aB", 1 if _oA > 0.75 else 0),
                                    "aC": _b.get("aC", 1 if _oB > 0.75 else 0)}
                    if ref: return ref
                _wb.close()
            except: pass
        # Fallback: média dos meses
        centros_set = set()
        for _m in meses_lista:
            if res_base.get(_m):
                centros_set.update(res_base[_m]["centros"].centro.tolist())
        ref = {}
        for cen in sorted(centros_set):
            vA,vB,vC,aA,aB,aC = [],[],[],[],[],[]
            for _m in meses_lista:
                if not res_base.get(_m): continue
                df_c = res_base[_m]["centros"]; row_ = df_c[df_c.centro == cen]
                if not row_.empty:
                    r_ = row_.iloc[0]
                    vA.append(r_.ocup_A); vB.append(r_.ocup_B); vC.append(r_.ocup_C)
                    aA.append(int(r_.ativo_A)); aB.append(int(r_.ativo_B)); aC.append(int(r_.ativo_C))
            if vA:
                ref[cen] = {"oA": np.mean(vA), "oB": np.mean(vB), "oC": np.mean(vC),
                             "aA": round(np.mean(aA)), "aB": round(np.mean(aB)), "aC": round(np.mean(aC))}
        return ref

    def _render_grade_form(centros_ref, prefix):
        """Grade de centros×turnos dentro de um st.form. Retorna dict {centro: {A,B,C}}."""
        # Instrução clara no topo
        st.markdown("""
<div style="background:#0D2A0D;border-left:4px solid #FFDE00;border-radius:6px;padding:10px 14px;margin-bottom:12px;font-size:13px;color:#FAFAFA;">
✏️ <b>Como usar:</b> para cada máquina, defina quantos <b>funcionários</b> devem trabalhar em cada turno.<br>
Use os botões <b>+</b> e <b>−</b> para ajustar. O valor <b>0</b> significa que o turno não estará ativo para aquela máquina.
</div>
""", unsafe_allow_html=True)

        # Legenda de cores
        st.markdown("""
<div style="display:flex;gap:18px;font-size:12px;color:#AAAAAA;margin-bottom:10px;padding:6px 0;">
  <span>🟢 Carga abaixo de 85% — turno confortável</span>
  <span>🟡 Carga entre 85% e 100% — atenção</span>
  <span>🔴 Carga acima de 100% — sobrecarga</span>
</div>
""", unsafe_allow_html=True)

        # Cabeçalho da tabela
        cols_h = st.columns([3, 1, 1, 1])
        cols_h[0].markdown("**Máquina — % de ocupação atual (Turno A / B / C)**")
        cols_h[1].markdown("**Nº Funcionários<br>Turno A (manhã)**", unsafe_allow_html=True)
        cols_h[2].markdown("**Nº Funcionários<br>Turno B (tarde)**", unsafe_allow_html=True)
        cols_h[3].markdown("**Nº Funcionários<br>Turno C (noite)**", unsafe_allow_html=True)

        st.markdown("<hr style='border-color:#333;margin:4px 0 8px 0;'>", unsafe_allow_html=True)

        ov = {}
        for cen, ref in centros_ref.items():
            eA = "🔴" if ref["oA"] > 1 else ("🟡" if ref["oA"] >= 0.85 else "🟢")
            eB = "🔴" if ref["oB"] > 1 else ("🟡" if ref["oB"] >= 0.85 else "🟢")
            eC = "🔴" if ref["oC"] > 1 else ("🟡" if ref["oC"] >= 0.85 else "🟢")
            c0, c1, c2, c3 = st.columns([3, 1, 1, 1])
            c0.markdown(
                f"<div style='padding:6px 0;font-size:13px;'>"
                f"<b style='color:#FFDE00;font-size:14px;'>{cen}</b><br>"
                f"<span style='font-size:12px;color:#AAAAAA;'>Ocupação: "
                f"{eA} {ref['oA']:.0%} (A) &nbsp;|&nbsp; "
                f"{eB} {ref['oB']:.0%} (B) &nbsp;|&nbsp; "
                f"{eC} {ref['oC']:.0%} (C)</span></div>",
                unsafe_allow_html=True
            )
            vA = c1.number_input(f"Funcionários Turno A — {cen}", 0, 5, ref["aA"],
                                  key=f"{prefix}_{cen}_A", label_visibility="collapsed",
                                  help=f"{cen}: número de funcionários no Turno A (manhã). Ocupação atual: {ref['oA']:.0%}")
            vB = c2.number_input(f"Funcionários Turno B — {cen}", 0, 5, ref["aB"],
                                  key=f"{prefix}_{cen}_B", label_visibility="collapsed",
                                  help=f"{cen}: número de funcionários no Turno B (tarde). Ocupação atual: {ref['oB']:.0%}")
            vC = c3.number_input(f"Funcionários Turno C — {cen}", 0, 5, ref["aC"],
                                  key=f"{prefix}_{cen}_C", label_visibility="collapsed",
                                  help=f"{cen}: número de funcionários no Turno C (noite). Ocupação atual: {ref['oC']:.0%}")
            ov[cen] = {"A": vA, "B": vB, "C": vC}

        st.markdown("<hr style='border-color:#333;margin:8px 0 4px 0;'>", unsafe_allow_html=True)
        st.markdown(
            "<div style='font-size:11px;color:#666;text-align:right;'>"
            "💡 Passe o mouse sobre os campos para ver detalhes de cada máquina e turno."
            "</div>", unsafe_allow_html=True
        )
        return ov

    _meses_disponiveis = [m for m in MESES if res_base.get(m)]

    with st.expander("➕ Criar novo cenário", expanded=st.session_state.pop("_cen_criar_open", len(st.session_state.cenarios) == 0)):
        _cen_form_id = st.session_state.get("_cen_form_id", 0)
        col_nome, col_btn = st.columns([4, 1])
        novo_nome = col_nome.text_input("Nome do cenário", placeholder="Ex: Redução B nov + Aumento A mar",
                                         key=f"cen_novo_nome_{_cen_form_id}")
        _btn_criar = col_btn.button("✅ Criar", type="primary", use_container_width=True,
                                    key="btn_criar_cenario",
                                    help="Clique para criar o cenário com o nome digitado")

        _opcoes_periodo = _meses_disponiveis + ["📅 ANO FY26"]
        meses_sel_raw = st.multiselect(
            "Período(s) a configurar", _opcoes_periodo,
            default=[],
            key=f"cen_meses_sel_{_cen_form_id}",
            help="Selecione ANO FY26 para configurar apenas a visão anual consolidada, ou meses individuais."
        )
        eh_ano_novo  = "📅 ANO FY26" in meses_sel_raw
        meses_sel    = [m for m in meses_sel_raw if m != "📅 ANO FY26"]  # meses reais selecionados
        # Se só ANO foi selecionado, usa todos os meses como base de cálculo mas exibe como ANO
        meses_calc   = meses_sel if meses_sel else (_meses_disponiveis if eh_ano_novo else [])

        if not novo_nome.strip():
            st.info("👆 Digite o nome do cenário e clique em 'Criar' para continuar.")
        elif not meses_sel_raw:
            st.warning("Selecione ao menos um período para configurar.")
        else:
            if eh_ano_novo and not meses_sel:
                st.markdown('<div class="aviso-warn">📅 <b>ANO FY26</b> — ajuste os funcionários por turno e veja o impacto direto nas produtividades anuais.</div>', unsafe_allow_html=True)
            elif eh_ano_novo and meses_sel:
                meses_str = ", ".join(m[:3].upper() for m in meses_sel)
                st.markdown(f'<div class="aviso-warn">📅 <b>ANO FY26 + meses individuais</b> — ANO e também: {meses_str}.</div>', unsafe_allow_html=True)
            else:
                n = len(meses_sel)
                meses_str = ", ".join(m[:3].upper() for m in meses_sel)
                st.markdown(f'<div class="aviso-ok">🗓️ <b>{n} mês(es)</b>: {meses_str}. Cada período tem overrides independentes.</div>', unsafe_allow_html=True)

            # Botão de abertura: tanto o botão "Criar" do topo quanto o botão "Configurar" abrem a grade
            _btn_key = f"btn_abrir_grade_{novo_nome}_{'_'.join(meses_sel_raw)}"
            if _btn_key not in st.session_state:
                st.session_state[_btn_key] = False
            # Botão "Criar" do topo também abre a grade
            if _btn_criar and novo_nome.strip():
                st.session_state[_btn_key] = True; st.rerun()
            if not st.session_state[_btn_key]:
                if st.button("✏️ Configurar funcionários por turno →", type="primary", key=f"open_{_btn_key}", use_container_width=True):
                    st.session_state[_btn_key] = True; st.rerun()
                st.stop()
            st.markdown('<div class="aviso-ok" style="margin-bottom:8px;">✏️ <b>Edite à vontade</b> — nada recalcula enquanto você ajusta. Só roda quando clicar em <b>Salvar</b>.</div>', unsafe_allow_html=True)

            # Pré-calcular referências de ocupação
            _refs = {}
            if eh_ano_novo:
                _refs["__ano__"] = _build_ocup_ref_ano(_meses_disponiveis)
            for _m in meses_sel:
                _refs[_m] = _build_ocup_ref(_m)

            _form_key = f"form_cen_{'ANO_' if eh_ano_novo else ''}{'_'.join(meses_sel)}_{novo_nome}"

            with st.form(key=_form_key):
                novo_ov_por_mes = {}

                # Abas: ANO FY26 primeiro (se selecionado), depois meses individuais
                _periodos_form = (["📅 ANO FY26"] if eh_ano_novo else []) + meses_sel

                if len(_periodos_form) == 1:
                    _p = _periodos_form[0]
                    st.markdown(f"**Configurando: {_p}**")
                    if _p == "📅 ANO FY26":
                        _ov_ano = _render_grade_form(_refs["__ano__"], f"fa_{novo_nome}")
                        novo_ov_por_mes["__ano__"] = _ov_ano
                    else:
                        novo_ov_por_mes[_p] = _render_grade_form(_refs[_p], f"fm_{novo_nome}_{_p}")
                else:
                    _tabs = st.tabs([f"📅 ANO" if p == "📅 ANO FY26" else f"📅 {p[:3].upper()}" for p in _periodos_form])
                    for _tab, _p in zip(_tabs, _periodos_form):
                        with _tab:
                            if _p == "📅 ANO FY26":
                                _ov_ano = _render_grade_form(_refs["__ano__"], f"fa_{novo_nome}")
                                novo_ov_por_mes["__ano__"] = _ov_ano
                            else:
                                novo_ov_por_mes[_p] = _render_grade_form(_refs[_p], f"fm_{novo_nome}_{_p}")

                salvar = st.form_submit_button("💾 Salvar cenário", type="primary", use_container_width=True)

            if salvar:
                if novo_nome in st.session_state.cenarios:
                    st.warning(f"Já existe um cenário chamado '{novo_nome}'. Escolha outro nome.")
                else:
                    with st.spinner(f"Calculando '{novo_nome}'..."):
                        _ov_ano_direto = novo_ov_por_mes.get("__ano__", {})
                        _ov_meses = {k: v for k, v in novo_ov_por_mes.items() if k != "__ano__"}
                        # ANO-only: meses não mudam, só o cálculo anual usa o override
                        if eh_ano_novo and not meses_sel:
                            res_cen = res_base
                        else:
                            res_cen = calcular(pmp, tempo, dist, aplic, dias, horas_turno, thresholds, suporte_cfg,
                                               horas_efetivas=horas_efetivas, overrides=_ov_meses)
                        # ANO FY26: calcular visão anual consolidada com o override do ANO
                        _res_ano_fy26 = None
                        _cp_data_ano  = None
                        if eh_ano_novo:
                            _dias_map = {m: res_base[m]["dias"] for m in MESES if res_base.get(m)}
                            # Sempre calcula cp_data (necessário para layout da aba ANO no export)
                            _cp_data_ano, _res_ano_fy26_meses = build_cp_data_from_meses(
                                res_base, tempo, dist, aplic, pmp, _dias_map,
                                horas_turno, horas_efetivas,
                                overrides_ano=_ov_ano_direto, suporte_cfg=suporte_cfg
                            )
                            # Preferir aba AnoFY26 do Excel; fallback = cálculo dos meses
                            _res_ano_fy26 = calcular_ano_fy26(
                                st.session_state.get("_fb_anual"),
                                _ov_ano_direto, horas_efetivas, suporte_cfg, horas_turno
                            ) or _res_ano_fy26_meses
                    _label_periodo = ("ANO FY26" if eh_ano_novo and not meses_sel else
                                      f"ANO FY26 + {','.join(m[:3].upper() for m in meses_sel)}" if eh_ano_novo else
                                      f"{len(meses_sel)} mês(es)")
                    st.session_state.cenarios[novo_nome] = {
                        "resultados": res_cen,
                        "res_ano_fy26": _res_ano_fy26,
                        "cp_data_ano": _cp_data_ano,
                        "mes": ("__ano__" if eh_ano_novo and not meses_sel else
                                meses_sel[0] if meses_sel else _meses_disponiveis[0]),
                        "meses_configurados": list(meses_sel_raw),
                        "overrides": novo_ov_por_mes,
                        "eh_ano": eh_ano_novo,
                    }
                    st.success(f"✅ '{novo_nome}' salvo — {_label_periodo} configurado(s)!")
                    st.session_state["_cen_criar_open"] = True
                    st.session_state["_cen_form_id"] = _cen_form_id + 1
                    st.rerun()

    if st.session_state.cenarios:
        _has_ano_cen = any(_v.get("eh_ano") for _v in st.session_state.cenarios.values())
        _res_base_chart = res_base
        if _has_ano_cen:
            _all_m_base = [m for m in MESES if res_base.get(m)]
            _base_ano = agregar_ano(res_base, _all_m_base)
            if _base_ano:
                _res_base_chart = dict(res_base); _res_base_chart["__ANO__"] = _base_ano
        todos={"📌 Base":_res_base_chart}
        for _k,_v in st.session_state.cenarios.items():
            _res=_v["resultados"]
            if _v.get("eh_ano"):
                if _v.get("res_ano_fy26") is None:
                    _ck_a=f"_ano_exp_{_k}_{_file_id}"
                    if _ck_a not in st.session_state or st.session_state[_ck_a][1] is None:
                        _dm_a={m:res_base[m]["dias"] for m in MESES if res_base.get(m)}
                        _ov_a=_v.get("overrides",{}).get("__ano__",{})
                        st.session_state[_ck_a]=build_cp_data_from_meses(
                            res_base,tempo,dist,aplic,pmp,_dm_a,
                            horas_turno,horas_efetivas,overrides_ano=_ov_a,suporte_cfg=suporte_cfg)
                    if st.session_state.get(_ck_a) and st.session_state[_ck_a][1] is not None:
                        _v["res_ano_fy26"]=st.session_state[_ck_a][1]
                        _v["cp_data_ano"]=st.session_state[_ck_a][0]
                if _v.get("res_ano_fy26"):
                    _res=dict(_res); _res["__ANO__"]=_v["res_ano_fy26"]
            todos[_k]=_res
        st.plotly_chart(grafico_cenarios(todos),use_container_width=True)

        # ── Resumo dos cenários salvos
        st.markdown('<div class="jd-sub">📋 Cenários salvos</div>',unsafe_allow_html=True)
        for nm,v in st.session_state.cenarios.items():
            _meses_conf=v.get("meses_configurados",([v.get("mes","")] if not v.get("eh_ano") else [m for m in MESES if res_base.get(m)]))
            _tag="ANO FY26" if v.get("eh_ano") else (", ".join(m[:3].upper() for m in _meses_conf))
            st.markdown(f'<div class="aviso-ok" style="margin:2px 0;padding:6px 12px;">📌 <b>{nm}</b> &nbsp;—&nbsp; meses com override: <b>{_tag}</b></div>', unsafe_allow_html=True)

        st.markdown('<div class="jd-sub">📊 Comparação detalhada</div>',unsafe_allow_html=True)

        # Monta lista de meses relevantes = união de todos os meses modificados em qualquer cenário
        _meses_modificados = set()
        for v in st.session_state.cenarios.values():
            if v.get("eh_ano"):
                _meses_modificados.update(m for m in MESES if res_base.get(m))
            else:
                _meses_modificados.update(v.get("meses_configurados", [v.get("mes","")]))
        _meses_relevantes = [m for m in MESES if m in _meses_modificados and res_base.get(m)]
        _opcoes_cmp = _meses_relevantes + (["📅 ANO FY26 (consolidado)"] if len(_meses_relevantes) > 1 else [])

        if not _opcoes_cmp:
            st.info("Nenhum mês modificado identificado nos cenários.")
        else:
            # Contexto: explica o que o seletor significa
            if len(_meses_relevantes) < len([m for m in MESES if res_base.get(m)]):
                _ignorados = [m[:3].upper() for m in MESES if res_base.get(m) and m not in _meses_modificados]
                st.markdown(f'<div class="aviso-warn">🗓️ Exibindo apenas os <b>{len(_meses_relevantes)} meses que têm overrides</b> em algum cenário. Meses sem override ({", ".join(_ignorados)}) são iguais à base em todos os cenários.</div>', unsafe_allow_html=True)

            mes_cmp = st.selectbox("Mês para comparar", _opcoes_cmp, key="mes_cmp_r",
                                   help="Apenas meses com overrides em algum cenário aparecem aqui.")
            eh_ano_cmp = mes_cmp == "📅 ANO FY26 (consolidado)"
            meses_cmp_lista = [m for m in MESES if res_base.get(m)] if eh_ano_cmp else ([mes_cmp] if res_base.get(mes_cmp) else [])

        if _opcoes_cmp and meses_cmp_lista:
            r_base_agg = agregar_ano(res_base, meses_cmp_lista)
            sufixo = " (ANO)" if eh_ano_cmp else ""
            cmp_rows = []
            for nm, res in todos.items():
                is_base = "Base" in nm
                dados_c = st.session_state.cenarios.get(nm, {})
                if eh_ano_cmp and not is_base and dados_c.get("eh_ano") and dados_c.get("res_ano_fy26"):
                    r_agg = dados_c["res_ano_fy26"]
                else:
                    r_agg = agregar_ano(res, meses_cmp_lista)
                if not r_agg or not r_base_agg: continue
                # Verifica se este cenário tem override no(s) mês(es) selecionado(s)
                if not is_base:
                    dados_c = st.session_state.cenarios.get(nm, {})
                    _meses_c = dados_c.get("meses_configurados", []) if not dados_c.get("eh_ano") else meses_cmp_lista
                    tem_override = eh_ano_cmp or any(m in _meses_c for m in meses_cmp_lista)
                else:
                    tem_override = True
                dA = round(r_agg["tot_A"]-r_base_agg["tot_A"],1) if not is_base else "—"
                dB = round(r_agg["tot_B"]-r_base_agg["tot_B"],1) if not is_base else "—"
                dC = round(r_agg["tot_C"]-r_base_agg["tot_C"],1) if not is_base else "—"
                dT = round(r_agg["total"]-r_base_agg["total"],1) if not is_base else "—"
                dL = f'{r_agg["prod_labor_tot"]-r_base_agg["prod_labor_tot"]:+.1%}' if not is_base else "—"
                _nm_label = nm if tem_override else f"{nm} ⚠️ sem override aqui"
                cmp_rows.append({"Cenário":_nm_label,
                    f"Turno A{sufixo}":r_agg["tot_A"],f"Turno B{sufixo}":r_agg["tot_B"],
                    f"Turno C{sufixo}":r_agg["tot_C"],f"Total{sufixo}":r_agg["total"],
                    "Labor Tot.":f'{r_agg["prod_labor_tot"]:.1%}',"Ciclo Tot.":f'{r_agg["prod_ciclo_tot"]:.1%}',
                    "ΔA":f"{dA:+.1f}" if isinstance(dA,float) else dA,
                    "ΔB":f"{dB:+.1f}" if isinstance(dB,float) else dB,
                    "ΔC":f"{dC:+.1f}" if isinstance(dC,float) else dC,
                    "Δ Total":f"{dT:+.1f}" if isinstance(dT,float) else dT,"Δ Labor":dL})
            df_cmp = pd.DataFrame(cmp_rows)
            def _sty_cmp(row):
                nm_r = str(row["Cenário"])
                is_base = "Base" in nm_r
                sem_ov = "⚠️ sem override" in nm_r
                if is_base: return [f"background-color:{JD_VERDE_ESC};color:#FFFFFF;font-weight:700"]*len(row)
                if sem_ov: return ["background-color:#1A1A1A;color:#666;font-style:italic"]*len(row)
                styles = [""]*len(row)
                try:
                    d = float(str(row["Δ Total"]).replace("+",""))
                    cd = "#003D10" if d<0 else ("#3D0000" if d>0 else "")
                    td = "#B9F6CA" if d<0 else ("#FF8A80" if d>0 else "")
                    for i,col in enumerate(df_cmp.columns):
                        if col in ("ΔA","ΔB","ΔC","Δ Total","Δ Labor"):
                            styles[i] = f"background-color:{cd};color:{td};font-weight:600"
                except: pass
                return styles
            st.dataframe(df_cmp.style.apply(_sty_cmp,axis=1), use_container_width=True, hide_index=True)

            for nome_cen, dados_cen in st.session_state.cenarios.items():
                r_cen_res = dados_cen["resultados"]
                _meses_c = dados_cen.get("meses_configurados",[dados_cen.get("mes","")]) if not dados_cen.get("eh_ano") else [m for m in MESES if res_base.get(m)]
                tem_ov_aqui = eh_ano_cmp or any(m in _meses_c for m in meses_cmp_lista)

                with st.expander(f"🔍 Detalhamento — {nome_cen} vs Base"):
                    if not tem_ov_aqui:
                        st.markdown(f'<div class="aviso-warn">⚠️ Este cenário não tem overrides em <b>{mes_cmp}</b> — os valores são idênticos à base. Os meses configurados são: <b>{", ".join(m[:3].upper() for m in _meses_c)}</b>.</div>', unsafe_allow_html=True)

                    _m_ref = meses_cmp_lista[0] if meses_cmp_lista else None
                    _meses_prod = meses_cmp_lista if eh_ano_cmp else ([_m_ref] if _m_ref else [])

                    def _calc_prod(res_d, meses_l, _res_ano=None):
                        # Se resultado do ANO FY26 direto disponível, usa ele
                        if _res_ano:
                            r = _res_ano
                            shc=r["h_ciclo"]; shl=r["h_labor"]
                            sha=r["h_ativos"]; sht=r["h_todos"]
                        else:
                            rr = [res_d.get(m) for m in meses_l if res_d.get(m)]
                            if not rr: return None
                            shc=sum(r["h_ciclo"] for r in rr); shl=sum(r["h_labor"] for r in rr)
                            sha=sum(r["h_ativos"] for r in rr); sht=sum(r["h_todos"] for r in rr)
                        return {"ciclo_op":shc/sha if sha>0 else 0,"ciclo_tot":shc/sht if sht>0 else 0,
                                "labor_op":shl/sha if sha>0 else 0,"labor_tot":shl/sht if sht>0 else 0}

                    _res_ano_cen = dados_cen.get("res_ano_fy26") if eh_ano_cmp and dados_cen.get("eh_ano") and not dados_cen.get("meses_configurados", [None])[0].startswith("📅 ANO") else None
                    _res_ano_base = read_horas_anual(st.session_state.get("_fb_anual"))
                    _prod_base_ano = None
                    if eh_ano_cmp and _res_ano_base:
                        _h = _res_ano_base
                        _shc=_h["h_ciclo"]; _shl=_h["h_labor"]; _sha=_h["h_ativos"]; _sht=_h["h_todos"]
                        _prod_base_ano = {"ciclo_op":_shc/_sha if _sha>0 else 0,"ciclo_tot":_shc/_sht if _sht>0 else 0,
                                          "labor_op":_shl/_sha if _sha>0 else 0,"labor_tot":_shl/_sht if _sht>0 else 0}
                    prod_b = _prod_base_ano if _prod_base_ano else _calc_prod(res_base, _meses_prod)
                    prod_c = _calc_prod(r_cen_res, _meses_prod, _res_ano=dados_cen.get("res_ano_fy26") if eh_ano_cmp else None)

                    if prod_b and prod_c:
                        st.markdown('<div class="jd-sub">Produtividades — Base vs Cenário</div>',unsafe_allow_html=True)
                        _items=[
                            ("Ciclo Operacional",prod_b["ciclo_op"],prod_c["ciclo_op"],False),
                            ("Ciclo Total",prod_b["ciclo_tot"],prod_c["ciclo_tot"],False),
                            ("Labor Operacional",prod_b["labor_op"],prod_c["labor_op"],False),
                            ("⭐ Labor Total",prod_b["labor_tot"],prod_c["labor_tot"],True),
                        ]
                        parts=[]
                        for lbl,vb,vc,dest in _items:
                            delta=vc-vb
                            arrow="↑" if delta>0 else ("↓" if delta<0 else "→")
                            cor_d="#69F0AE" if delta>0 else ("#FF5252" if delta<0 else "#888888")
                            bg="linear-gradient(135deg,#1F4D19,#0D2A0D)" if dest else "linear-gradient(135deg,#151525,#0D0D1A)"
                            brd="#FFDE00" if dest else "#2A3A4A"
                            parts.append(
                                '<div style="background:'+bg+';border:1.5px solid '+brd+';border-radius:10px;padding:12px 14px;">'
                                +'<div style="font-size:9px;color:#7BC67A;text-transform:uppercase;letter-spacing:.05em;font-weight:600;margin-bottom:8px;">'+lbl+'</div>'
                                +'<div style="display:flex;justify-content:space-between;align-items:flex-end;">'
                                +'<div><div style="font-size:9px;color:#888;margin-bottom:1px;">Base</div>'
                                +'<div style="font-size:19px;font-weight:800;color:#AAAAAA;">'+f"{vb:.1%}"+'</div></div>'
                                +'<div style="font-size:16px;color:#444;padding-bottom:3px;">→</div>'
                                +'<div style="text-align:right"><div style="font-size:9px;color:#FFDE00;margin-bottom:1px;">Cenário</div>'
                                +'<div style="font-size:19px;font-weight:800;color:#FFFFFF;">'+f"{vc:.1%}"+'</div></div>'
                                +'</div>'
                                +'<div style="margin-top:6px;padding-top:6px;border-top:1px solid #333;display:flex;align-items:center;gap:5px;">'
                                +'<span style="font-size:13px;">'+arrow+'</span>'
                                +'<span style="font-size:13px;font-weight:700;color:'+cor_d+';">'+f"{delta:+.1%}"+'</span>'
                                +'<span style="font-size:10px;color:#666;">vs base</span>'
                                +'</div></div>'
                            )
                        st.markdown('<div style="display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin-bottom:14px;">'+"".join(parts)+'</div>',unsafe_allow_html=True)
                        _prod_rows=[]
                        for lbl,vb,vc,dest in _items:
                            delta=vc-vb
                            _prod_rows.append({"Indicador":lbl,"Base":f"{vb:.1%}","Cenário":f"{vc:.1%}","Δ":f"{delta:+.1%}","":("✅" if delta>0 else ("⚠️" if delta<0 else "➖"))})
                        _df_pr=pd.DataFrame(_prod_rows)
                        def _sty_pr(row):
                            dest2="Labor Total" in str(row["Indicador"])
                            bg_d=f"background-color:{JD_AMARELO};color:{JD_VERDE_ESC};font-weight:700" if dest2 else ""
                            try:
                                dv=float(str(row["Δ"]).replace("+","").replace("%",""))/100
                                cd2="background-color:#003D10;color:#B9F6CA;font-weight:700" if dv>0 else ("background-color:#3D0000;color:#FF8A80;font-weight:700" if dv<0 else "background-color:#222;color:#888")
                            except: cd2=""
                            return [bg_d if dest2 else "" for col in _df_pr.columns[:-2]] + [cd2, bg_d if dest2 else ""]
                        st.dataframe(_df_pr.style.apply(_sty_pr,axis=1),use_container_width=True,hide_index=True)

                    st.markdown('<div class="jd-sub">Ativação por centro</div>',unsafe_allow_html=True)
                    det_rows=[]
                    if _m_ref and res_base.get(_m_ref) and r_cen_res.get(_m_ref):
                        centros_set2=sorted(set(res_base[_m_ref]["centros"].centro.tolist()+r_cen_res[_m_ref]["centros"].centro.tolist()))
                        for cen in centros_set2:
                            rb_c=res_base[_m_ref]["centros"]; rc_c=r_cen_res[_m_ref]["centros"]
                            rb_r=rb_c[rb_c.centro==cen]; rc_r=rc_c[rc_c.centro==cen]
                            if rb_r.empty or rc_r.empty: continue
                            rb=rb_r.iloc[0]; rc=rc_r.iloc[0]
                            mA=int(rb.ativo_A)!=int(rc.ativo_A); mB=int(rb.ativo_B)!=int(rc.ativo_B); mC=int(rb.ativo_C)!=int(rc.ativo_C)
                            det_rows.append({"Centro":cen,"Ocup.A":f"{rb.ocup_A:.0%}","Base A":int(rb.ativo_A),"Cen A":int(rc.ativo_A),"Ocup.B":f"{rb.ocup_B:.0%}","Base B":int(rb.ativo_B),"Cen B":int(rc.ativo_B),"Ocup.C":f"{rb.ocup_C:.0%}","Base C":int(rb.ativo_C),"Cen C":int(rc.ativo_C),"Mudou":"✅ Igual" if not(mA or mB or mC) else ("A " if mA else "")+("B " if mB else "")+("C" if mC else "")+"alterado(s)"})
                    if det_rows:
                        df_det=pd.DataFrame(det_rows)
                        def _sty_det(row):
                            if "alterado" in str(row["Mudou"]): return ["background-color:#3D2D00;color:#FFE57F"]*len(row)
                            return [""]*len(row)
                        st.dataframe(df_det.style.apply(_sty_det,axis=1),use_container_width=True,hide_index=True)

        st.markdown("---")
        if st.session_state.cenarios:
            st.markdown('<div class="jd-sub">📥 Baixar cenários</div>',unsafe_allow_html=True)
            for nm_dl, v_dl in st.session_state.cenarios.items():
                _is_ano_only_dl = v_dl.get("eh_ano") and not [m for m in v_dl.get("meses_configurados",[]) if not m.startswith("📅")]
                if _is_ano_only_dl:
                    _meses_dl = []
                    _res_ano_c_dl = v_dl.get("res_ano_fy26")
                    _cp_fb_dl = v_dl.get("cp_data_ano")
                    if _res_ano_c_dl is None or _cp_fb_dl is None:
                        _ck = f"_ano_exp_{nm_dl}_{_file_id}"
                        if _ck not in st.session_state or st.session_state[_ck][1] is None:
                            _dm = {m: res_base[m]["dias"] for m in MESES if res_base.get(m)}
                            _ov = v_dl.get("overrides", {}).get("__ano__", {})
                            st.session_state[_ck] = build_cp_data_from_meses(
                                res_base, tempo, dist, aplic, pmp, _dm,
                                horas_turno, horas_efetivas,
                                overrides_ano=_ov, suporte_cfg=suporte_cfg
                            )
                        _cp_fb_dl = _cp_fb_dl or st.session_state[_ck][0]
                        _res_ano_c_dl = _res_ano_c_dl or st.session_state[_ck][1]
                    _eh_ano_dl = _res_ano_c_dl is not None
                else:
                    _meses_dl = [m for m in v_dl.get("meses_configurados",[v_dl.get("mes","")]) if m and not m.startswith("📅")]
                    if not _meses_dl: _meses_dl = [m for m in MESES if res_base.get(m)]
                    _eh_ano_dl = False
                    _res_ano_c_dl = None
                    _cp_fb_dl = None
                _hash_dl = hash(str(v_dl["resultados"]) + str(res_base) + nm_dl + str(_meses_dl) + "cen")
                st.download_button(
                    f"📥 {nm_dl} vs Base",
                    data=exportar_cenario_vs_base_cached(
                        _hash_dl, res_base, v_dl["resultados"], _meses_dl, nm_dl,
                        _res_ano_fy26_b=res_base if _eh_ano_dl else None,
                        _res_ano_fy26_c=_res_ano_c_dl if _eh_ano_dl else None,
                        _cp_data_fallback=_cp_fb_dl,
                        _file_bytes_ano=file_bytes
                    ),
                    file_name=f"cenario_vs_base_{nm_dl.replace(' ','_')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"cen_dl_{nm_dl}"
                )
            st.markdown("---")
            dn = st.selectbox("Remover cenário", list(st.session_state.cenarios.keys()), key="del_c")
            if st.button("🗑️ Remover", type="secondary", key="btn_del_cen"):
                del st.session_state.cenarios[dn]; st.rerun()
    else:
        st.info("Nenhum cenário criado ainda.")

# ── TAB 6 COMPARAÇÃO
if _aba == "🔄 Comparar com Excel":
    st.markdown('<div class="jd-section">Comparação com o Excel atual</div>',unsafe_allow_html=True)
    st.markdown('<div class="aviso-warn">🔄 Esta aba compara automaticamente os resultados calculados pelo app com as abas mensais do seu arquivo Excel (ex: <b>NovFY26, DezFY26…</b>). A comparação usa cache — só recalcula quando você muda thresholds, horas de turno ou sobe um novo arquivo.</div>', unsafe_allow_html=True)
    cache_key=f"cmp_{hash(str(dias))}_{hash(str(thresholds))}_{hash(str(horas_turno))}"
    if st.session_state.get("cmp_cache_key")!=cache_key:
        with st.spinner("Comparando com o Excel... (executado uma única vez por configuração)"):
            _res_hash = hash(str(res_base))
            _file_hash = hash(file_bytes)
            _r,_d,_e=comparar_com_excel_cached(_res_hash, res_base, _file_hash, file_bytes, tempo, dist, aplic, pmp, dias, horas_turno, thresholds, suporte_cfg)
        st.session_state["cmp_cache_key"]=cache_key; st.session_state["cmp_cache_resumo"]=_r; st.session_state["cmp_cache_detalhe"]=_d; st.session_state["cmp_cache_err"]=_e
    else:
        st.caption("✅ Resultado em cache — nenhum recálculo necessário.")
    df_resumo=st.session_state["cmp_cache_resumo"]; err=st.session_state["cmp_cache_err"]
    if err: st.error(err)
    elif df_resumo is not None and len(df_resumo)>0:
        n_ok=(df_resumo["Status"].str.startswith("✅")).sum() if "Status" in df_resumo else 0
        n_warn=(df_resumo["Status"].str.startswith("🟡")).sum() if "Status" in df_resumo else 0
        n_err=(df_resumo["Status"].str.startswith("🔴")).sum() if "Status" in df_resumo else 0
        c1,c2,c3=st.columns(3)
        c1.metric("✅ Meses iguais",n_ok); c2.metric("🟡 Pequena diferença",n_warn); c3.metric("🔴 Com divergência",n_err)
        def build_vis(df):
            rows=[]
            for _,r in df.iterrows():
                def cell(app,excel,delta):
                    try:
                        d_n=int(str(delta).replace("+",""))
                        icon="✅" if d_n==0 else ("🟡" if abs(d_n)<=2 else "🔴")
                        return f"{icon} App={app} | Excel={excel} ({delta})" if d_n!=0 else f"✅ {app}"
                    except: return f"App={app} / Excel={excel}"
                rows.append({"Mês":r["Mês"],"Status":r["Status"],"Turno A":cell(r.get("CEN A App","?"),r.get("CEN A Excel","?"),r.get("Δ A","?")),"Turno B":cell(r.get("CEN B App","?"),r.get("CEN B Excel","?"),r.get("Δ B","?")),"Turno C":cell(r.get("CEN C App","?"),r.get("CEN C Excel","?"),r.get("Δ C","?")),"Total":cell(r.get("Total App","?"),r.get("Total Excel","?"),r.get("Δ Total","?")),"Labor":f"App={r.get('Labor App','?')} / Excel={r.get('Labor Excel','?')}"})
            return pd.DataFrame(rows)
        df_vis=build_vis(df_resumo)
        def sty_res(row):
            st_v=str(row.get("Status",""))
            if "✅" in st_v: base="background-color:#003D10;color:#B9F6CA"
            elif "🟡" in st_v: base="background-color:#3D2D00;color:#FFE57F"
            elif "🔴" in st_v: base="background-color:#3D0000;color:#FF8A80"
            else: base="background-color:#2D1A00;color:#FFD54F"
            styles=[]
            for col in row.index:
                if col=="Status": styles.append(base)
                elif col in("Turno A","Turno B","Turno C","Total"):
                    val=str(row[col])
                    if "🔴" in val: styles.append("background-color:#3D0000;color:#FF8A80")
                    elif "🟡" in val: styles.append("background-color:#3D2D00;color:#FFE57F")
                    elif "✅" in val: styles.append("background-color:#003D10;color:#B9F6CA")
                    else: styles.append("")
                else: styles.append("")
            return styles
        st.dataframe(df_vis.astype(str).style.apply(sty_res,axis=1),use_container_width=True,hide_index=True)
    else:
        st.warning("Nenhuma aba mensal encontrada (NovFY26, DezFY26…)")

# ── TAB 7 EXPORTAR
if _aba == "📥 Exportar":
    st.markdown('<div class="jd-section">Exportação</div>',unsafe_allow_html=True)

    sub_tab, sub_res = st.tabs(["📋 Tabelona completa — layout INPUTDISTRIBUIÇÃO", "📊 Resultados"])

    # ══════════════════════════════════════════
    # SUB-ABA 1 — TABELONA COMPLETA COM COMPARAÇÃO
    # ══════════════════════════════════════════
    with sub_tab:
        st.markdown('<div class="jd-sub">📋 Tabelona completa — layout idêntico ao INPUTDISTRIBUIÇÃO</div>', unsafe_allow_html=True)
        st.markdown("""
Gera a **tabelona completa** no mesmo layout do seu Excel — colunas A até os modelos todos —
com as colunas de % ocupação calculadas pelo App. Células em **vermelho** = divergência com o Excel de referência.

Inclui também, no mesmo Excel: **totais de minutos/horas/dias** por turno lá em cima, e o bloco de
**DADOS AUTOMÁTICOS** (% ocupação, ativo/inativo, horas disponíveis, suporte e produtividades) a partir da coluna F linha 66.
        """)
        if st.button("📋 Gerar tabelona completa", key="btn_tabelona"):
            with st.spinner("Gerando tabelona... (~10s)"):
                import openpyxl as _opx
                from openpyxl.styles import PatternFill as _PF, Font as _Ft, Alignment as _Al, Border as _Bd, Side as _Sd

                _F_VERDE=_PF("solid",fgColor="66FF66"); _F_AMAR=_PF("solid",fgColor="FFFF00")
                _F_AZUL=_PF("solid",fgColor="00B0F0"); _F_PRETO=_PF("solid",fgColor="000000")
                _F_CINZA=_PF("solid",fgColor="D9D9D9"); _F_CINZA2=_PF("solid",fgColor="BFBFBF")
                _F_BRANCO=_PF("solid",fgColor="FFFFFF"); _F_VERM=_PF("solid",fgColor="FF0000")
                _F_VERM_S=_PF("solid",fgColor="FFCCCC"); _F_VERDE_JD=_PF("solid",fgColor="1F4D19")
                _BRD=_Bd(left=_Sd(style="thin",color="AAAAAA"),right=_Sd(style="thin",color="AAAAAA"),
                        top=_Sd(style="thin",color="AAAAAA"),bottom=_Sd(style="thin",color="AAAAAA"))

                def _ec(ws,r,c,val,fill=None,bold=False,color="000000",size=8,center=True,wrap=False):
                    cell=ws.cell(row=r,column=c,value=val)
                    cell.font=_Ft(name="Arial",bold=bold,color=color,size=size)
                    cell.fill=fill or _F_BRANCO
                    cell.alignment=_Al(horizontal="center" if center else "left",vertical="center",wrap_text=wrap)
                    cell.border=_BRD
                    return cell
                def _ec_pct(ws,r,c,val,fill=None,bold=False,color="000000",size=8):
                    cell=_ec(ws,r,c,val,fill,bold,color,size)
                    cell.number_format="0.0000000000%"
                    return cell

                def _cor_pct(v):
                    try:
                        f=float(v)
                        if f>=1.06: return _PF("solid",fgColor="FF0000")
                        if f>=1.00: return _PF("solid",fgColor="FFFF00")
                        if f>=0.40: return _PF("solid",fgColor="92D050")
                        return _F_BRANCO
                    except: return _F_BRANCO

                _CANDS_T={"Novembro":["NovFY26","NOV","Nov","NOVEMBRO"],"Dezembro":["DezFY26","DEZ","Dez","DEZEMBRO"],
                          "Janeiro":["JanFY26","JAN","Jan","JANEIRO"],"Fevereiro":["FevFY26","FEV","Fev","FEVEREIRO"],
                          "Março":["MarFY26","MAR","Mar","MARÇO"],"Abril":["AbrFY26","ABR","Abr","ABRIL"],
                          "Maio":["MaiFY26","MAI","Mai","MAIO"],"Junho":["JunFY26","JUN","Jun","JUNHO"],
                          "Julho":["JulFY26","JUL","Jul","JULHO"],"Agosto":["AgoFY26","AGO","Ago","AGOSTO"],
                          "Setembro":["SetFY26","SET","Set","SETEMBRO"],"Outubro":["OutFY26","OUT","Out","OUTUBRO"]}

                hA_t=horas_turno["A"]; hB_t=horas_turno["B"]; hC_t=horas_turno["C"]
                thr_A_t=thresholds["A"]/100; thr_B_t=thresholds["B"]/100; thr_C_t=thresholds["C"]/100

                try:
                    df_all_t=(aplic.merge(pmp,on="modelo").merge(tempo,on=["centro","peca"]).merge(dist,on=["centro","peca"]))
                    if "pc_trat_x" in df_all_t.columns and "pc_trat_y" in df_all_t.columns:
                        df_all_t["pc_trat"]=df_all_t["pc_trat_y"].fillna(df_all_t["pc_trat_x"]).fillna(1.0); df_all_t.drop(columns=["pc_trat_x","pc_trat_y"],inplace=True)
                    elif "pc_trat_y" in df_all_t.columns: df_all_t.rename(columns={"pc_trat_y":"pc_trat"},inplace=True)
                    elif "pc_trat_x" in df_all_t.columns: df_all_t.rename(columns={"pc_trat_x":"pc_trat"},inplace=True)
                    df_all_t["pc_trat"]=pd.to_numeric(df_all_t.get("pc_trat",1.0),errors="coerce").fillna(1.0).clip(lower=1.0)
                    if "vol_int" not in df_all_t.columns: df_all_t["vol_int"] = 1.0
                    df_all_t["vol_int"]    = pd.to_numeric(df_all_t["vol_int"],    errors="coerce").fillna(1.0)
                    df_all_t["div_carga"]  = pd.to_numeric(df_all_t["div_carga"],  errors="coerce").fillna(0.0)
                    df_all_t["div_volume"] = pd.to_numeric(df_all_t["div_volume"], errors="coerce").fillna(0.0)
                    df_all_t["disponib"]   = pd.to_numeric(df_all_t["disponib"],   errors="coerce").fillna(1.0)
                    if "perf_op" not in df_all_t.columns: df_all_t["perf_op"]=1.0
                    df_all_t["perf_op"]=pd.to_numeric(df_all_t["perf_op"],errors="coerce").fillna(1.0)
                    df_all_t["indice_ciclo"]=(df_all_t.t_ciclo*df_all_t.div_carga*df_all_t.div_volume*df_all_t.vol_int)/(df_all_t.disponib*df_all_t["perf_op"])
                    df_all_t["min_ciclo"]=df_all_t.indice_ciclo*df_all_t.qtd
                    df_all_t["min_labor"]=df_all_t.t_labor*df_all_t.div_carga*df_all_t.qtd*df_all_t.pc_trat
                    agg_cp_t=df_all_t.groupby(["centro","peca","mes"])[["min_ciclo","min_labor"]].sum()
                    # Índice por (centro,peca) usando INPUTTEMPO — fonte única da verdade para t_ciclo/t_labor/pc_trat
                    _tempo_idx_t = {(r.centro, r.peca): r for r in tempo.itertuples()}
                except Exception as _e_merge:
                    st.error(f"Erro ao preparar dados: {_e_merge}"); st.stop()

                try:
                    wb_r=_opx.load_workbook(BytesIO(file_bytes),read_only=True,data_only=True)
                    MAPA_T={m: find_aba(wb_r.sheetnames, cands) for m, cands in _CANDS_T.items()}
                    MAPA_T={m: a for m, a in MAPA_T.items() if a}
                    _aba_ref_t=next((a for a in MAPA_T.values() if a), None)
                    if _aba_ref_t is None:
                        st.error("❌ Nenhuma aba mensal (NovFY26, DEZ, JAN etc.) encontrada no arquivo. "
                                 "A exportação precisa de pelo menos uma aba mensal para ler o layout de referência.")
                        wb_r.close()
                    else:
                        ws_nov_t=wb_r[_aba_ref_t]
                        # ── Auto-detecta colunas extras antes de MÁQUINA (ex: coluna SEQ) ──
                        _xl_offset = 0; _xl_hdr_row = 6
                        for _sr in range(5, min(18, ws_nov_t.max_row + 1)):
                            _rv = [ws_nov_t.cell(_sr, _sc).value for _sc in range(1, 9)]
                            if not any(_rv): continue
                            for _si, _xv in enumerate(_rv):
                                if _xv and str(_xv).strip().upper().startswith("CEN"):
                                    _xl_offset = _si; _xl_hdr_row = max(1, _sr - 1); break
                            if any(_xv and str(_xv).strip().upper().startswith("CEN") for _xv in _rv): break
                        _xl_mod_start = 19 + _xl_offset   # col (1-idx) onde começam os modelos
                        _xl_main_end  = 18 + _xl_offset   # col (1-idx) fim do bloco principal
                        base_rows_t=list(ws_nov_t.iter_rows(min_row=_xl_hdr_row+1,max_row=_xl_hdr_row+57,min_col=1,max_col=87+_xl_offset,values_only=True))
                        base_rows_t=[r for r in base_rows_t if len(r)>_xl_offset+1 and r[_xl_offset] and r[_xl_offset+1]]
                        modelos_xl_t=[str(ws_nov_t.cell(_xl_hdr_row,c).value) for c in range(_xl_mod_start,88+_xl_offset)
                                      if ws_nov_t.cell(_xl_hdr_row,c).value and str(ws_nov_t.cell(_xl_hdr_row,c).value).startswith("MODELO")]
                        modelo_col_idx={str(ws_nov_t.cell(_xl_hdr_row,c).value):(c-_xl_mod_start) for c in range(_xl_mod_start,88+_xl_offset)
                                        if ws_nov_t.cell(_xl_hdr_row,c).value and str(ws_nov_t.cell(_xl_hdr_row,c).value).startswith("MODELO")}

                        dados_mes_t={}
                        for mes_t,aba_t in MAPA_T.items():
                            if aba_t not in wb_r.sheetnames: continue
                            ws_m_t=wb_r[aba_t]
                            dados_mes_t[mes_t]={
                                "main":list(ws_m_t.iter_rows(min_row=_xl_hdr_row+1,max_row=_xl_hdr_row+57,min_col=1,max_col=_xl_main_end,values_only=True)),
                                "vols":list(ws_m_t.iter_rows(min_row=_xl_hdr_row+1,max_row=_xl_hdr_row+57,min_col=_xl_mod_start,max_col=87+_xl_offset,values_only=True))}
                        wb_r.close()

                        try:
                            _aba_aplic_t=_get_aba("INPUTAPLICAÇÃO")
                            aplic_orig=pd.read_excel(BytesIO(file_bytes),sheet_name=_aba_aplic_t,header=0)
                            aplic_orig=aplic_orig.rename(columns={aplic_orig.columns[0]:"centro",aplic_orig.columns[1]:"peca"})
                        except: aplic_orig=aplic.copy()

                        wb_out=_opx.Workbook(); primeira_t=True
                        for mes_t in MESES:
                            d_t=dias.get(mes_t,0)
                            if d_t==0: continue
                            minA_t=d_t*hA_t*60; minB_t=d_t*hB_t*60; minC_t=d_t*hC_t*60
                            heA_t=horas_efetivas["A"]; heB_t=horas_efetivas["B"]; heC_t=horas_efetivas["C"]
                            dm_t=dados_mes_t.get(mes_t,{}); pmp_mes_t=pmp[pmp.mes==mes_t]

                            if primeira_t: ws_out=wb_out.active; ws_out.title=mes_t[:10]; primeira_t=False
                            else: ws_out=wb_out.create_sheet(mes_t[:10])
                            ws_out.freeze_panes="F7"

                            _F_CINZA_H=_PF("solid",fgColor="D9D9D9"); _F_VD_H=_PF("solid",fgColor="1F4D19")
                            ws_out.merge_cells("A1:O1")
                            _ec(ws_out,1,1,f"TOTAIS — {mes_t.upper()}",_F_VD_H,True,"FFFFFF",9,True)
                            for ci_h,txt_h,f_h in [(17,"TURNO A",_F_VERDE),(18,"TURNO B",_F_AMAR),(19,"TURNO C",_F_AZUL)]:
                                _ec(ws_out,1,ci_h,txt_h,f_h,True,"000000",8,True)
                            ws_out.row_dimensions[1].height=14
                            ws_out.merge_cells("A2:O2")
                            _ec(ws_out,2,1,"TOTAL DE MINUTOS",_F_CINZA_H,True,"000000",8,False)
                            _ec(ws_out,2,17,minA_t,_F_VERDE,True,"000000",8)
                            _ec(ws_out,2,18,minB_t,_F_AMAR,True,"000000",8)
                            _ec(ws_out,2,19,minC_t,_F_AZUL,True,"000000",8)
                            ws_out.row_dimensions[2].height=13
                            ws_out.merge_cells("A3:O3")
                            _ec(ws_out,3,1,"TOTAL DE HORAS",_F_CINZA_H,True,"000000",8,False)
                            _ec(ws_out,3,17,minA_t/60,_F_VERDE,True,"000000",8)
                            _ec(ws_out,3,18,minB_t/60,_F_AMAR,True,"000000",8)
                            _ec(ws_out,3,19,minC_t/60,_F_AZUL,True,"000000",8)
                            ws_out.row_dimensions[3].height=13
                            ws_out.merge_cells("A4:O4")
                            _ec(ws_out,4,1,"Nº DIAS TRABALHADOS",_F_CINZA_H,True,"000000",8,False)
                            _ec(ws_out,4,17,d_t,_F_VERDE,True,"FF0000",9)
                            _ec(ws_out,4,18,d_t,_F_AMAR,True,"FF0000",9)
                            _ec(ws_out,4,19,d_t,_F_AZUL,True,"FF0000",9)
                            ws_out.row_dimensions[4].height=13

                            ws_out.merge_cells(f"A5:{get_column_letter(20+len(modelos_xl_t))}5")
                            _ec(ws_out,5,1,f"RESUMO DA CARGA — {mes_t.upper()} ({d_t} dias)",_F_VERDE_JD,True,"FFFFFF",10,True)
                            ws_out.row_dimensions[5].height=18

                            hdrs_f=[("Máquina",_F_CINZA2,"000000"),("PEÇA",_F_CINZA2,"000000"),("DESCRIÇÃO",_F_CINZA2,"000000"),
                                    ("PÇ/TRAT",_F_CINZA2,"000000"),("UM",_F_CINZA2,"000000"),
                                    ("Tempo Ciclo (min)",_F_PRETO,"FFFFFF"),("Tempo Labor (min)",_F_PRETO,"FFFFFF"),
                                    ("Div. Carga",_PF("solid",fgColor="FF0000"),"FFFF00"),("Vol. Interna",_F_CINZA2,"000000"),
                                    ("Div. Volume",_PF("solid",fgColor="FF0000"),"FFFF00"),("Disponib.",_F_CINZA2,"000000"),
                                    ("Perf. Op.",_F_CINZA2,"000000"),("Indice Ciclo",_F_CINZA2,"000000"),
                                    ("JA.A",_F_VERDE,"000000"),("JA.B",_F_AMAR,"000000"),("JA.C",_F_AZUL,"000000"),
                                    ("TOTAL CICLOS (MIN)",_F_CINZA,"000000"),("TOTAL LABOR (MIN)",_F_CINZA,"000000"),
                                    ("TOTAL PECAS",_F_CINZA,"000000"),("PECAS\n(Excel)",_PF("solid",fgColor="BBDEFB"),"000000")]
                            largs_t=[9,8,16,6,5,9,9,8,8,8,8,8,9,8,8,8,12,12,8,8]
                            for ci_t,(h_t,f_t,cor_t) in enumerate(hdrs_f,1):
                                _ec(ws_out,6,ci_t,h_t,f_t,True,cor_t,8,True,True)
                                ws_out.column_dimensions[get_column_letter(ci_t)].width=largs_t[ci_t-1]
                            for mi_t,mod_t in enumerate(modelos_xl_t):
                                ci_t=21+mi_t
                                _ec(ws_out,6,ci_t,mod_t,_F_CINZA,True,"000000",7,True,True)
                                ws_out.column_dimensions[get_column_letter(ci_t)].width=7
                            ws_out.row_dimensions[6].height=42

                            main_data_t=dm_t.get("main",[]); vols_data_t=dm_t.get("vols",[])
                            for ri_t_idx,base_row_t in enumerate(base_rows_t):
                                _o=_xl_offset
                                cen_t=str(base_row_t[_o]).strip(); peca_t=str(base_row_t[_o+1]).strip()
                                ri_t=7+ri_t_idx
                                tc_xl_t=base_row_t[_o+5]; tl_xl_t=base_row_t[_o+6]
                                dc_xl_t=base_row_t[_o+7]; vi_xl_t=base_row_t[_o+8]; dv_xl_t=base_row_t[_o+9]
                                di_xl_t=base_row_t[_o+10]; idx_xl_t=base_row_t[_o+11]

                                mrow_t=main_data_t[ri_t_idx] if ri_t_idx<len(main_data_t) else [None]*(_xl_main_end)
                                xl_pA_t=mrow_t[_o+12] if len(mrow_t)>_o+12 else None
                                xl_pB_t=mrow_t[_o+13] if len(mrow_t)>_o+13 else None
                                xl_ciclo_t=mrow_t[_o+15] if len(mrow_t)>_o+15 else None
                                xl_pecas_t=mrow_t[_o+17] if len(mrow_t)>_o+17 else None
                                # Lê os valores de distribuição do MÊS CORRENTE (não da aba estrutural)
                                # para comparar corretamente com o INPUT
                                dc_xl_t = mrow_t[_o+7] if len(mrow_t)>_o+7 else base_row_t[_o+7]
                                vi_xl_t = mrow_t[_o+8] if len(mrow_t)>_o+8 else base_row_t[_o+8]
                                dv_xl_t = mrow_t[_o+9] if len(mrow_t)>_o+9 else base_row_t[_o+9]
                                di_xl_t = mrow_t[_o+10] if len(mrow_t)>_o+10 else base_row_t[_o+10]
                                idx_xl_t = mrow_t[_o+11] if len(mrow_t)>_o+11 else base_row_t[_o+11]
                                vrow_t=dm_t.get("vols",[])[ri_t_idx] if dm_t.get("vols") and ri_t_idx<len(dm_t["vols"]) else []

                                try: mc_t=float(agg_cp_t.loc[(cen_t,peca_t,mes_t),"min_ciclo"])
                                except: mc_t=0.0
                                try: ml_t=float(agg_cp_t.loc[(cen_t,peca_t,mes_t),"min_labor"])
                                except: ml_t=0.0

                                pA_t=mc_t/minA_t if minA_t>0 else 0
                                pB_t=mc_t/minB_t if minB_t>0 else 0
                                pC_t=mc_t/minC_t if minC_t>0 else 0

                                app_mod_v={}
                                for mod_t2 in modelos_xl_t:
                                    qtd_t=int(pmp_mes_t[pmp_mes_t.modelo==mod_t2]["qtd"].sum()) if mod_t2 in pmp_mes_t.modelo.values else 0
                                    fr_t=aplic_orig[(aplic_orig.centro==cen_t)&(aplic_orig.peca==peca_t)]
                                    flag_t=int(fr_t[mod_t2].values[0]) if len(fr_t)>0 and mod_t2 in fr_t.columns and not pd.isna(fr_t[mod_t2].values[0] if len(fr_t)>0 else 0) else 0
                                    app_mod_v[mod_t2]=qtd_t*flag_t
                                # Calcula total de peças direto do aplic+pmp (independente dos nomes do Excel)
                                _aplic_models_cp = set(aplic[(aplic.centro==cen_t)&(aplic.peca==peca_t)].modelo.unique())
                                app_tot_t = int(sum(int(pmp_mes_t[pmp_mes_t.modelo==m]["qtd"].sum()) for m in _aplic_models_cp))

                                def _df(a,b,tol=0.02):
                                    if b is None: return False
                                    try: return abs(float(a or 0)-float(b))>tol
                                    except: return False

                                div_A_t=_df(pA_t,xl_pA_t,0.02); div_B_t=_df(pB_t,xl_pB_t,0.02)
                                div_c_t=_df(mc_t,xl_ciclo_t,1); div_p_t=_df(app_tot_t,xl_pecas_t,0.5)

                                # Lê direto do dist (INPUTDISTRIBUIÇÃO) — fonte única da verdade
                                _dist_row = dist[(dist.centro==cen_t)&(dist.peca==peca_t)]
                                if not _dist_row.empty:
                                    dc_inp = float(_dist_row.iloc[0].div_carga)
                                    vi_inp = float(_dist_row.iloc[0].vol_int)
                                    dv_inp = float(_dist_row.iloc[0].div_volume)
                                    di_inp = float(_dist_row.iloc[0].disponib)
                                    po_inp = float(_dist_row.iloc[0].perf_op) if hasattr(_dist_row.iloc[0], "perf_op") else 1.0
                                else:
                                    dc_inp = float(dc_xl_t or 0)
                                    vi_inp = float(vi_xl_t or 1)
                                    dv_inp = float(dv_xl_t or 0)
                                    di_inp = float(di_xl_t or 1)
                                    po_inp = 1.0
                                # Usa INPUTTEMPO como fonte única da verdade — fallback para referência se ausente
                                _tr_t = _tempo_idx_t.get((cen_t, peca_t))
                                tc_inp = float(_tr_t.t_ciclo) if _tr_t is not None else float(tc_xl_t or 0)
                                tl_inp = float(_tr_t.t_labor) if _tr_t is not None else float(tl_xl_t or 0)
                                pc_trat_t = (float(_tr_t.pc_trat) if _tr_t is not None and hasattr(_tr_t,"pc_trat") else None) or (float(base_row_t[3]) if base_row_t[3] else 1.0)
                                idx_app_t = (tc_inp * dc_inp * dv_inp * vi_inp) / (di_inp * po_inp) if (di_inp and po_inp) else 0.0
                                div_idx_t = abs(float(idx_xl_t or 0) - float(idx_app_t or 0)) > 0.5
                                # Recompute mc_t e ml_t com t_ciclo/t_labor do INPUTTEMPO (evita inflação e outdated ref)
                                mc_t = idx_app_t * app_tot_t
                                ml_t = tl_inp * dc_inp * pc_trat_t * app_tot_t
                                pA_t = mc_t/minA_t if minA_t>0 else 0
                                pB_t = mc_t/minB_t if minB_t>0 else 0
                                pC_t = mc_t/minC_t if minC_t>0 else 0
                                div_A_t=_df(pA_t,xl_pA_t,0.02); div_B_t=_df(pB_t,xl_pB_t,0.02)
                                div_c_t=_df(mc_t,xl_ciclo_t,1)

                                # Vermelho = valor no arquivo mensal difere do INPUT
                                def _dif_val(xl_val, inp_val, tol=0.001):
                                    if xl_val is None: return False
                                    try:
                                        tol_rel = max(tol, abs(float(inp_val)) * 0.001)
                                        return abs(float(xl_val) - float(inp_val)) > tol_rel
                                    except: return False

                                div_dc_t = _dif_val(dc_xl_t, dc_inp)
                                div_vi_t = _dif_val(vi_xl_t, vi_inp)
                                div_dv_t = _dif_val(dv_xl_t, dv_inp)
                                div_di_t = _dif_val(di_xl_t, di_inp)

                                _fill_dc = _F_VERM if div_dc_t else _F_BRANCO
                                _fill_vi = _F_VERM if div_vi_t else _F_BRANCO
                                _fill_dv = _F_VERM if div_dv_t else _F_BRANCO
                                _fill_di = _F_VERM if div_di_t else _F_BRANCO

                                _ec(ws_out,ri_t,1,cen_t,_F_BRANCO,False,"000000",8,False)
                                _ec(ws_out,ri_t,2,peca_t,_F_BRANCO,False,"000000",8,False)
                                _ec(ws_out,ri_t,3,base_row_t[_o+2],_F_BRANCO,False,"000000",8,False)
                                _ec(ws_out,ri_t,4,base_row_t[_o+3],_F_BRANCO,False,"000000",8)
                                _ec(ws_out,ri_t,5,base_row_t[_o+4],_F_BRANCO,False,"000000",8)
                                # Col 6-7: mostra valor do INPUTTEMPO; vermelho se diferir do arquivo de referência
                                _fill_tc = _F_VERM if _dif_val(tc_xl_t, tc_inp, 0.01) else _F_PRETO
                                _fill_tl = _F_VERM if _dif_val(tl_xl_t, tl_inp, 0.01) else _F_PRETO
                                _tc_color = "000000" if _fill_tc == _F_VERM else "FFFFFF"
                                _tl_color = "000000" if _fill_tl == _F_VERM else "FFFFFF"
                                _ec(ws_out,ri_t,6,tc_inp,_fill_tc,False,_tc_color,8)
                                _ec(ws_out,ri_t,7,tl_inp,_fill_tl,False,_tl_color,8)
                                # Colunas 8-11: SEMPRE valor do INPUT — branco=correto, vermelho=difere do mês
                                _ec(ws_out,ri_t,8,dc_inp,_fill_dc,False,"000000",8)
                                _ec(ws_out,ri_t,9,vi_inp,_fill_vi,False,"000000",8)
                                _ec(ws_out,ri_t,10,dv_inp,_fill_dv,False,"000000",8)
                                _ec(ws_out,ri_t,11,di_inp,_fill_di,False,"000000",8)
                                _fill_po = _F_VERM if abs(po_inp-1.0)>0.001 else _F_BRANCO
                                _ec(ws_out,ri_t,12,po_inp,_fill_po,False,"000000",8)
                                _ec(ws_out,ri_t,13,float(idx_app_t),_F_VERM if div_idx_t else _F_BRANCO,False,"000000",8)
                                _ec_pct(ws_out,ri_t,14,pA_t,_F_VERM if div_A_t else _cor_pct(pA_t))
                                _ec_pct(ws_out,ri_t,15,pB_t,_F_VERM if div_B_t else _cor_pct(pB_t))
                                _ec_pct(ws_out,ri_t,16,pC_t,_cor_pct(pC_t))
                                _ec(ws_out,ri_t,17,mc_t,_F_VERM if div_c_t else _F_BRANCO,False,"000000",8)
                                _ec(ws_out,ri_t,18,ml_t,_F_BRANCO,False,"000000",8)
                                _ec(ws_out,ri_t,19,app_tot_t,_F_VERM if div_p_t else _F_BRANCO,False,"000000",8)
                                _xl_pec_ref = int(float(xl_pecas_t)) if xl_pecas_t is not None else ""
                                _ec(ws_out,ri_t,20,_xl_pec_ref,_F_VERM if div_p_t else _PF("solid",fgColor="BBDEFB"),False,"000000",8)
                                for mi_t2,mod_t2 in enumerate(modelos_xl_t):
                                    ci_t2=21+mi_t2
                                    v_app_t=app_mod_v.get(mod_t2,0)
                                    col_idx=modelo_col_idx.get(mod_t2, mi_t2)
                                    v_xl_t=vrow_t[col_idx] if vrow_t and col_idx<len(vrow_t) else None
                                    if v_xl_t is not None:
                                        try: div_vm=abs(float(v_app_t)-float(v_xl_t or 0))>0.5
                                        except: div_vm=False
                                        fill_vm=_F_VERM if div_vm else (_F_CINZA if v_app_t else _F_BRANCO)
                                    else:
                                        fill_vm=_F_CINZA if v_app_t else _F_BRANCO
                                    _ec(ws_out,ri_t,ci_t2,v_app_t if v_app_t else None,fill_vm,False,"000000",7)
                                ws_out.row_dimensions[ri_t].height=13

                            nota_rt=7+len(base_rows_t)+1
                            ws_out.merge_cells(f"A{nota_rt}:{get_column_letter(19+len(modelos_xl_t))}{nota_rt}")
                            nt=ws_out.cell(nota_rt,1,"🔴 VERMELHO (células): valor no arquivo de referência difere do INPUTTEMPO/INPUTDISTRIBUIÇÃO — cálculo usa sempre o INPUT  |  🔴 JA.A/JA.B vermelho = % ocupação difere do Excel  |  🔴 Rosa = total de ciclos ou peças difere  |  Cinza = presente no App")
                            nt.font=_Ft(name="Arial",bold=True,size=8,color="CC0000")
                            nt.fill=_PF("solid",fgColor="FFEEEE")
                            nt.alignment=_Al(horizontal="left",vertical="center")
                            ws_out.row_dimensions[nota_rt].height=14

                            r_auto = res_base.get(mes_t)
                            _COL_F = 6; _ROW_START = 66
                            if r_auto:
                                _F_VD_JD = _PF("solid",fgColor="1F4D19"); _F_AMAR_JD = _PF("solid",fgColor="FFDE00")
                                _F_CINZA_D = _PF("solid",fgColor="D9D9D9")
                                ws_out.merge_cells(start_row=_ROW_START,start_column=_COL_F,end_row=_ROW_START,end_column=_COL_F+9)
                                _ec(ws_out,_ROW_START,_COL_F,"DADOS AUTOMÁTICOS",_F_CINZA_D,True,"000000",9,True)
                                ws_out.row_dimensions[_ROW_START].height=14
                                _r67=_ROW_START+1
                                _ec(ws_out,_r67,_COL_F,"PERÍODO:",_F_CINZA_D,True,"000000",8,False)
                                _ec(ws_out,_r67,_COL_F+1,mes_t,_F_BRANCO,True,"FF0000",9,False)
                                _ec(ws_out,_r67,_COL_F+3,"DATA DE REVISÃO:",_F_CINZA_D,True,"000000",8,False)
                                _ec(ws_out,_r67,_COL_F+4,datetime.now().strftime("%d/%m/%Y"),_F_BRANCO,True,"FF0000",9)
                                _ec(ws_out,_r67,_COL_F+6,"HORAS POR TURNO DE TRABALHO",_F_CINZA_D,True,"000000",8,True)
                                ws_out.row_dimensions[_r67].height=14
                                _r68=_ROW_START+2
                                _ec(ws_out,_r68,_COL_F+6,heA_t,_F_CINZA_D,True,"000000",8)
                                _ec(ws_out,_r68,_COL_F+7,heB_t,_F_CINZA_D,True,"000000",8)
                                _ec(ws_out,_r68,_COL_F+8,heC_t,_F_CINZA_D,True,"000000",8)
                                ws_out.row_dimensions[_r68].height=14
                                _r69=_ROW_START+3
                                for _ci_g,_txt_g,_fg in [(_COL_F+1,"TURNO A",_F_VERDE),(_COL_F+2,"TURNO B",_F_AMAR),(_COL_F+3,"TURNO C",_F_AZUL),(_COL_F+4,"TURNO A",_F_VERDE),(_COL_F+5,"TURNO B",_F_AMAR),(_COL_F+6,"TURNO C",_F_AZUL),(_COL_F+7,"TURNO A",_F_VERDE),(_COL_F+8,"TURNO B",_F_AMAR),(_COL_F+9,"TURNO C",_F_AZUL)]:
                                    _ec(ws_out,_r69,_ci_g,_txt_g,_fg,True,"000000",8,True)
                                ws_out.row_dimensions[_r69].height=14
                                _r70=_ROW_START+4
                                _ec(ws_out,_r70,_COL_F,"Centro",_F_PRETO,True,"FFFFFF",8)
                                for _ci_s,_txt_s in [(_COL_F+1,"% Ocup"),(_COL_F+2,"% Ocup"),(_COL_F+3,"% Ocup"),(_COL_F+4,"Ativo"),(_COL_F+5,"Ativo"),(_COL_F+6,"Ativo"),(_COL_F+7,"Horas"),(_COL_F+8,"Horas"),(_COL_F+9,"Horas")]:
                                    _ec(ws_out,_r70,_ci_s,_txt_s,_F_PRETO,True,"FFFFFF",8)
                                ws_out.row_dimensions[_r70].height=14
                                _ri_c=_ROW_START+5
                                for _,_crow in r_auto["centros"].iterrows():
                                    def _cbg_t(v):
                                        if v>=1.06: return _PF("solid",fgColor="FF0000")
                                        if v>=1.00: return _PF("solid",fgColor="FFFF00")
                                        if v>=0.40: return _PF("solid",fgColor="92D050")
                                        return _F_BRANCO
                                    _ec(ws_out,_ri_c,_COL_F,_crow.centro,_F_BRANCO,False,"000000",8,False)
                                    _ec_pct(ws_out,_ri_c,_COL_F+1,_crow.ocup_A,_cbg_t(_crow.ocup_A))
                                    _ec_pct(ws_out,_ri_c,_COL_F+2,_crow.ocup_B,_cbg_t(_crow.ocup_B))
                                    _ec_pct(ws_out,_ri_c,_COL_F+3,_crow.ocup_C,_cbg_t(_crow.ocup_C))
                                    _ec(ws_out,_ri_c,_COL_F+4,int(_crow.ativo_A),_F_VERDE if _crow.ativo_A else _F_AMAR,True,"000000",8)
                                    _ec(ws_out,_ri_c,_COL_F+5,int(_crow.ativo_B),_F_VERDE if _crow.ativo_B else _F_AMAR,True,"000000",8)
                                    _ec(ws_out,_ri_c,_COL_F+6,int(_crow.ativo_C),_F_AZUL if _crow.ativo_C else _F_CINZA,True,"000000",8)
                                    _ec(ws_out,_ri_c,_COL_F+7,_crow.horas_disp_A if _crow.ativo_A else 0,_F_VERDE if _crow.ativo_A else _F_BRANCO,True,"000000",8)
                                    _ec(ws_out,_ri_c,_COL_F+8,_crow.horas_disp_B if _crow.ativo_B else 0,_F_AMAR if _crow.ativo_B else _F_BRANCO,True,"000000",8)
                                    _ec(ws_out,_ri_c,_COL_F+9,_crow.horas_disp_C if _crow.ativo_C else 0,_F_AZUL if _crow.ativo_C else _F_BRANCO,True,"000000",8)
                                    ws_out.row_dimensions[_ri_c].height=13; _ri_c+=1
                                _sup_a=r_auto["suporte"]
                                for _snm,_skey in [("TOTAL DE OPERADORES",None),("LAVADORA E INSPEÇÃO","lavadora"),("GRAVAÇÃO E ESTANQUEIDADE","gravacao"),("PRESET","preset"),("CORINGA","coringa"),("FACILITADOR","facilitador"),("TOTAL POR TURNO",None),("TOTAL FUNCIONÁRIOS",None)]:
                                    _bold_s="TOTAL" in _snm; _bg_s=_F_AMAR_JD if _bold_s else _F_BRANCO; _fg_s="1F4D19" if _bold_s else "000000"
                                    _ec(ws_out,_ri_c,_COL_F,_snm,_bg_s,_bold_s,_fg_s,8,False)
                                    if _skey:
                                        _sv=_sup_a[_skey]
                                        for _ci_sv,_tk in [(_COL_F+4,"A"),(_COL_F+5,"B"),(_COL_F+6,"C")]:
                                            _ec(ws_out,_ri_c,_ci_sv,_sv[_tk],_F_VERDE if _tk=="A" else (_F_AMAR if _tk=="B" else _F_AZUL),True,"000000",8)
                                        for _ci_hv,_tk,_hef in [(_COL_F+7,"A",heA_t),(_COL_F+8,"B",heB_t),(_COL_F+9,"C",heC_t)]:
                                            _hv=_sv[_tk]*_hef*d_t
                                            _ec(ws_out,_ri_c,_ci_hv,_hv if _hv else 0,_F_VERDE if _tk=="A" else (_F_AMAR if _tk=="B" else _F_AZUL),True,"000000",8)
                                    elif "TOTAL DE OPERADORES" in _snm:
                                        for _ci_sv,_vv in [(_COL_F+4,r_auto["op_A"]),(_COL_F+5,r_auto["op_B"]),(_COL_F+6,r_auto["op_C"])]:
                                            _ec(ws_out,_ri_c,_ci_sv,_vv,_F_AMAR_JD,True,"1F4D19",8)
                                        for _ci_hv,_vv,_hef in [(_COL_F+7,r_auto["op_A"],heA_t),(_COL_F+8,r_auto["op_B"],heB_t),(_COL_F+9,r_auto["op_C"],heC_t)]:
                                            _ec(ws_out,_ri_c,_ci_hv,_vv*_hef*d_t,_F_AMAR_JD,True,"1F4D19",8)
                                    elif "TOTAL POR TURNO" in _snm:
                                        for _ci_sv,_vv in [(_COL_F+4,r_auto["tot_A"]),(_COL_F+5,r_auto["tot_B"]),(_COL_F+6,r_auto["tot_C"])]:
                                            _ec(ws_out,_ri_c,_ci_sv,_vv,_F_AMAR_JD,True,"1F4D19",8)
                                        for _ci_hv,_vv,_hef in [(_COL_F+7,r_auto["tot_A"],heA_t),(_COL_F+8,r_auto["tot_B"],heB_t),(_COL_F+9,r_auto["tot_C"],heC_t)]:
                                            _ec(ws_out,_ri_c,_ci_hv,_vv*_hef*d_t,_F_AMAR_JD,True,"1F4D19",8)
                                    elif "FUNCIONÁRIOS" in _snm:
                                        _ec(ws_out,_ri_c,_COL_F+4,r_auto["total"],_F_AMAR_JD,True,"1F4D19",9)
                                        _th=r_auto["tot_A"]*heA_t*d_t+r_auto["tot_B"]*heB_t*d_t+r_auto["tot_C"]*heC_t*d_t
                                        _ec(ws_out,_ri_c,_COL_F+7,_th,_F_AMAR_JD,True,"1F4D19",9)
                                    ws_out.row_dimensions[_ri_c].height=13; _ri_c+=1
                                _ri_c+=1
                                for _pnm,_pv,_dest in [("PRODUTIVIDADE POR TEMPO DE CICLO OPERACIONAL",r_auto["prod_ciclo_op"],False),("PRODUTIVIDADE POR TEMPO DE CICLO TOTAL",r_auto["prod_ciclo_tot"],False),("PRODUTIVIDADE POR TEMPO DE LABOR OPERACIONAL",r_auto["prod_labor_op"],False),("PRODUTIVIDADE POR TEMPO DE LABOR TOTAL ★",r_auto["prod_labor_tot"],True)]:
                                    ws_out.merge_cells(start_row=_ri_c,start_column=_COL_F,end_row=_ri_c,end_column=_COL_F+8)
                                    _ec(ws_out,_ri_c,_COL_F,_pnm,_F_AMAR_JD if _dest else _F_BRANCO,_dest,"1F4D19" if _dest else "000000",8,False)
                                    _ec_pct(ws_out,_ri_c,_COL_F+9,_pv,_F_AMAR_JD if _dest else _F_BRANCO)
                                    ws_out.row_dimensions[_ri_c].height=14; _ri_c+=1
                                for _ci_w,_ww in [(_COL_F,14),(_COL_F+1,8),(_COL_F+2,8),(_COL_F+3,8),(_COL_F+4,8),(_COL_F+5,8),(_COL_F+6,8),(_COL_F+7,10),(_COL_F+8,10),(_COL_F+9,10)]:
                                    ws_out.column_dimensions[get_column_letter(_ci_w)].width=_ww

                        try:
                            _cp_ano_t = build_cp_data_anual(res_base, tempo, dist, aplic, pmp, file_bytes=file_bytes)
                        except: _cp_ano_t = None
                        _horas_ano_t = read_horas_anual(file_bytes)
                        gerar_aba_anual(wb_out, res_base, label="ANO", cp_data=_cp_ano_t, horas_anual=_horas_ano_t)
                        tabelona_buf=BytesIO(); wb_out.save(tabelona_buf); tabelona_buf.seek(0)
                        st.session_state["tabelona_buf"] = tabelona_buf

                except Exception as _e_tab:
                    st.error(f"Erro ao gerar tabelona: {_e_tab}")

        if st.session_state.get("tabelona_buf"):
            st.download_button(
                "📋 Baixar tabelona completa (layout INPUTDISTRIBUIÇÃO + divergências)",
                data=st.session_state["tabelona_buf"],
                file_name="tabelona_por_mes.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_tabelona"
            )

    # ══════════════════════════════════════════
    # SUB-ABA 2 — RESULTADOS
    # ══════════════════════════════════════════
    with sub_res:
        st.markdown('<div class="jd-sub">📋 Resultado completo e base tratada</div>', unsafe_allow_html=True)
        st.markdown("Exportações dos **resultados reais** calculados pelo App, sem comparação com o Excel de referência.")

        st.markdown('<div class="jd-sub">📊 Tabelona — apenas resultados (sem vermelhinho)</div>', unsafe_allow_html=True)
        st.markdown("""
Gera a **tabelona no layout do INPUTDISTRIBUIÇÃO** usando **apenas os dados calculados pelo App** —
sem precisar das abas mensais do Excel de referência. Sem comparação, sem vermelho.
Inclui totais de minutos/horas/dias, bloco de DADOS AUTOMÁTICOS e aba ANO.
        """)
        if st.button("📊 Gerar tabelona de resultados", key="btn_tab_pura"):
            with st.spinner("Gerando tabelona de resultados..."):
                tab_pura_buf = gerar_tabelona_pura(res_base, tempo, dist, aplic, pmp, dias, horas_turno, horas_efetivas, thresholds)
                st.session_state["tab_pura_buf"] = tab_pura_buf
        if st.session_state.get("tab_pura_buf"):
            st.download_button("📥 Baixar tabelona de resultados", data=st.session_state["tab_pura_buf"],
                file_name="tabelona_resultados.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_tab_pura")

        st.markdown("---")
        c1,c2=st.columns(2)
        with c1:
            st.markdown("**Resultado completo (todas as abas)**")
            _exp_hash = hash(str(res_base) + str(horas_turno) + str(thresholds) + str(hash(file_bytes)))
            st.download_button("📥 Baixar resultado base",data=exportar_cached(_exp_hash, res_base, tempo, dist, aplic, pmp, _file_bytes=file_bytes),file_name="resultado_usinagem.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",key="dl_res_base")
        with c2:
            st.markdown("**Base tratada (pós-JOIN)**")
            _base_cache_key = f"bt_{_exp_hash}"
            if st.session_state.get("_base_cache_key") != _base_cache_key:
                _buf_base = BytesIO(); df_interm.to_excel(_buf_base, index=False); _buf_base.seek(0)
                st.session_state["base_tratada_cache"] = _buf_base.read()
                st.session_state["_base_cache_key"] = _base_cache_key
            st.download_button("📥 Baixar base tratada", data=st.session_state["base_tratada_cache"], file_name="base_tratada.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="dl_base")
        if st.session_state.get("cenarios"):
            st.markdown('<div class="jd-sub">Cenários salvos</div>',unsafe_allow_html=True)
            for nm,v in st.session_state.cenarios.items():
                _is_ano_only_e = v.get("eh_ano") and not [m for m in v.get("meses_configurados",[]) if not m.startswith("📅")]
                if _is_ano_only_e:
                    _meses_e = []
                    _res_ano_c_e = v.get("res_ano_fy26")
                    _cp_fb_e = v.get("cp_data_ano")
                    if _res_ano_c_e is None or _cp_fb_e is None:
                        _ck_e = f"_ano_exp_{nm}_{_file_id}"
                        if _ck_e not in st.session_state or st.session_state[_ck_e][1] is None:
                            _dm_e = {m: res_base[m]["dias"] for m in MESES if res_base.get(m)}
                            _ov_e = v.get("overrides", {}).get("__ano__", {})
                            st.session_state[_ck_e] = build_cp_data_from_meses(
                                res_base, tempo, dist, aplic, pmp, _dm_e,
                                horas_turno, horas_efetivas,
                                overrides_ano=_ov_e, suporte_cfg=suporte_cfg
                            )
                        _cp_fb_e = _cp_fb_e or st.session_state[_ck_e][0]
                        _res_ano_c_e = _res_ano_c_e or st.session_state[_ck_e][1]
                    _eh_ano_e = _res_ano_c_e is not None
                else:
                    _meses_e = [m for m in v.get("meses_configurados",[v.get("mes","")]) if m and not m.startswith("📅")]
                    if not _meses_e: _meses_e = [m for m in MESES if res_base.get(m)]
                    _eh_ano_e = False
                    _res_ano_c_e = None
                    _cp_fb_e = None
                _cen_hash = hash(str(v["resultados"]) + str(res_base) + nm + str(_meses_e))
                st.download_button(
                    f"📥 Cenário vs Base: {nm}",
                    data=exportar_cenario_vs_base_cached(
                        _cen_hash, res_base, v["resultados"], _meses_e, nm,
                        _res_ano_fy26_b=res_base if _eh_ano_e else None,
                        _res_ano_fy26_c=_res_ano_c_e if _eh_ano_e else None,
                        _cp_data_fallback=_cp_fb_e,
                        _file_bytes_ano=file_bytes
                    ),
                    file_name=f"cenario_vs_base_{nm.replace(' ','_')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"exp_vsb_{nm}"
                )
